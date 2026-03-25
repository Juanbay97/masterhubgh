import os
import json

import frappe


def _xlsx_library():
	try:
		import openpyxl  # type: ignore

		return openpyxl
	except Exception:
		frappe.throw("La librería openpyxl no está instalada en el entorno.")


SHEET_MAP = {
	"Banco": "Banco Siesa",
	"EPS": "Entidad EPS Siesa",
	"AFP": "Entidad AFP Siesa",
	"Cesantias": "Entidad Cesantias Siesa",
	"CCF": "Entidad CCF Siesa",
	"Tipo Cotizante": "Tipo Cotizante Siesa",
	"Unidad Negocio": "Unidad Negocio Siesa",
	"Centro Costos": "Centro Costos Siesa",
	"Centro Trabajo": "Centro Trabajo Siesa",
	"Grupo Empleados": "Grupo Empleados Siesa",
}


SHEET_ALIASES = {
	"Entidad AFP": "Entidad AFP Siesa",
	"Entidad AFC": "Entidad Cesantias Siesa",
	"Banco": "Banco Siesa",
	"Cotizante": "Tipo Cotizante Siesa",
	"Unidad Negocio": "Unidad Negocio Siesa",
	"ID CO": "Centro Costos Siesa",
}


JSON_KEY_MAP = {
	"bancos": "Banco Siesa",
	"entidades_pension": "Entidad AFP Siesa",
	"entidades_cesantias": "Entidad Cesantias Siesa",
	"entidades_eps": "Entidad EPS Siesa",
	"entidades_salud": "Entidad EPS Siesa",
	"tipos_cotizante": "Tipo Cotizante Siesa",
	"unidades_negocio": "Unidad Negocio Siesa",
	"centros_costo": "Centro Costos Siesa",
}


def _get_value(row, *keys):
	for key in keys:
		if key in row and row[key] not in (None, ""):
			return row[key]
	return None


def _header_map(raw_headers):
	mapped = []
	for h in raw_headers:
		label = str(h).strip() if h is not None else ""
		mapped.append(label.lower())
	return mapped


def _normalize_code_desc(row):
	code = _get_value(
		row,
		"code",
		"codigo",
		"código",
	)
	desc = _get_value(
		row,
		"description",
		"descripcion",
		"descripción",
		"nombre",
	)
	if code is None or desc is None:
		return None, None
	return str(code).strip(), str(desc).strip()


def _upsert_catalog_row(doctype, code, desc):
	name = frappe.db.get_value(doctype, {"code": str(code)})
	if name:
		doc = frappe.get_doc(doctype, name)
		doc.description = str(desc)
		doc.enabled = 1
		doc.save(ignore_permissions=True)
		return 0, 1

	doc = frappe.get_doc({
		"doctype": doctype,
		"code": str(code),
		"description": str(desc),
		"enabled": 1,
	})
	doc.insert(ignore_permissions=True)
	return 1, 0


def _parse_sheet_sections(ws):
	headers = _header_map([c.value for c in ws[1]])
	rows = []
	for r in ws.iter_rows(min_row=2, values_only=True):
		if not any(r):
			continue
		row = {headers[i]: r[i] for i in range(min(len(headers), len(r)))}
		code, desc = _normalize_code_desc(row)
		if code and desc:
			rows.append((code, desc))
	return rows


def _parse_embedded_sections(ws, section_markers):
	"""
	Parse sheets that contain multiple catalogs embedded one below another.
	`section_markers` is an ordered list of tuples: (marker_text, doctype).
	Rows are assigned to the current active doctype until next marker appears.
	"""
	headers = _header_map([c.value for c in ws[1]])
	markers = [(m.lower().strip(), d) for m, d in section_markers]
	active_doctype = None
	result = {d: [] for _, d in section_markers}

	for r in ws.iter_rows(min_row=2, values_only=True):
		if not any(r):
			continue
		row = {headers[i]: r[i] for i in range(min(len(headers), len(r)))}
		code, desc = _normalize_code_desc(row)
		if not code or not desc:
			continue

		lc_code = code.lower().strip()
		lc_desc = desc.lower().strip()
		found_marker = False
		for marker, doctype in markers:
			if lc_code == marker or lc_desc == marker:
				active_doctype = doctype
				found_marker = True
				break
		if found_marker:
			continue

		if not active_doctype:
			active_doctype = section_markers[0][1]

		result[active_doctype].append((code, desc))

	return result


def _reset_target_doctypes(doctypes):
	for dt in doctypes:
		if frappe.db.exists("DocType", dt):
			frappe.db.sql(f"delete from `tab{dt}`")


def _import_from_workbook(wb):
	created = 0
	updated = 0

	effective_sheet_map = dict(SHEET_MAP)
	effective_sheet_map.update(SHEET_ALIASES)

	# Special embedded catalogs in this workbook format
	if "Entidad AFP" in wb.sheetnames:
		ws = wb["Entidad AFP"]
		embedded = _parse_embedded_sections(
			ws,
			[
				("Entidad AFP", "Entidad AFP Siesa"),
				("Entidad EPS", "Entidad EPS Siesa"),
			],
		)
		for doctype, pairs in embedded.items():
			for code, desc in pairs:
				c, u = _upsert_catalog_row(doctype, code, desc)
				created += c
				updated += u
		effective_sheet_map.pop("Entidad AFP", None)

	if "Entidad AFC" in wb.sheetnames:
		ws = wb["Entidad AFC"]
		embedded = _parse_embedded_sections(
			ws,
			[
				("Entidad AFC", "Entidad Cesantias Siesa"),
				("Entidad CCF", "Entidad CCF Siesa"),
			],
		)
		for doctype, pairs in embedded.items():
			for code, desc in pairs:
				c, u = _upsert_catalog_row(doctype, code, desc)
				created += c
				updated += u
		effective_sheet_map.pop("Entidad AFC", None)

	for sheet_name, doctype in effective_sheet_map.items():
		if sheet_name not in wb.sheetnames:
			continue

		ws = wb[sheet_name]
		for code, desc in _parse_sheet_sections(ws):
			c, u = _upsert_catalog_row(doctype, code, desc)
			created += c
			updated += u

	return created, updated


def _import_from_json(json_path):
	if not os.path.exists(json_path):
		return 0, 0

	with open(json_path, "r", encoding="utf-8") as f:
		payload = json.load(f)

	if not isinstance(payload, dict):
		return 0, 0

	created = 0
	updated = 0
	for key, doctype in JSON_KEY_MAP.items():
		rows = payload.get(key) or []
		if not isinstance(rows, list):
			continue

		for row in rows:
			if not isinstance(row, dict):
				continue
			code = row.get("codigo") or row.get("code")
			desc = row.get("descripcion") or row.get("description")
			if code in (None, "") or desc in (None, ""):
				continue
			c, u = _upsert_catalog_row(doctype, str(code).strip(), str(desc).strip())
			created += c
			updated += u

	return created, updated


@frappe.whitelist()
def import_madre_codigos(
	file_path="Archivos siesa/madre_codigos.xlsx",
	json_fallback_path="Archivos siesa/Arquitectura refactorizacion siesa modulo contratación/codigos_siesa_completo.json",
	reset_targets=0,
):
	"""
	Ejecutar desde bench console o llamada whitelisted interna.
	"""
	openpyxl = _xlsx_library()
	absolute = file_path
	if not os.path.isabs(absolute):
		absolute = f"/workspace/{file_path}"
	if not os.path.exists(absolute):
		frappe.throw(f"No se encontró el archivo de códigos: {absolute}")

	if int(reset_targets or 0) == 1:
		_reset_target_doctypes(
			[
				"Banco Siesa",
				"Entidad AFP Siesa",
				"Entidad Cesantias Siesa",
				"Entidad EPS Siesa",
				"Entidad CCF Siesa",
				"Tipo Cotizante Siesa",
				"Unidad Negocio Siesa",
				"Centro Costos Siesa",
			]
		)

	wb = openpyxl.load_workbook(absolute, data_only=True)
	created_xlsx, updated_xlsx = _import_from_workbook(wb)

	json_absolute = json_fallback_path
	if json_absolute and not os.path.isabs(json_absolute):
		json_absolute = f"/workspace/{json_absolute}"
	created_json, updated_json = _import_from_json(json_absolute) if json_absolute else (0, 0)

	created = created_xlsx + created_json
	updated = updated_xlsx + updated_json

	frappe.db.commit()
	return {
		"ok": True,
		"created": created,
		"updated": updated,
		"sources": {
			"xlsx": {"created": created_xlsx, "updated": updated_xlsx},
			"json": {"created": created_json, "updated": updated_json},
		},
	}
