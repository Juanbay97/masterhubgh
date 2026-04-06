import json
from pathlib import Path

import frappe
from frappe.sessions import get_csrf_token

_PROCEDENCIA_WORKBOOK_CACHE = None

_DEFAULT_BANCOS_SIESA = [
	{"code": "0002", "description": "BANCOLOMBIA"},
	{"code": "1001", "description": "BANCO DE BOGOTA"},
	{"code": "1002", "description": "BANCO POPULAR"},
	{"code": "1006", "description": "BBVA"},
	{"code": "1007", "description": "BANCOLOMBIA AHORROS"},
	{"code": "1012", "description": "BANCO GNB SUDAMERIS"},
	{"code": "1013", "description": "BANCO DAVIVIENDA"},
	{"code": "1014", "description": "BANCO DE OCCIDENTE"},
	{"code": "1019", "description": "SCOTIABANK COLPATRIA"},
	{"code": "1032", "description": "BANCO CAJA SOCIAL"},
	{"code": "1052", "description": "BANCO AV VILLAS"},
	{"code": "1060", "description": "BANCO AGRARIO"},
	{"code": "1551", "description": "BANCO W"},
	{"code": "1801", "description": "NEQUI"},
	{"code": "1803", "description": "DAVIPLATA"},
]

_DEFAULT_PROCEDENCIA_PAISES = [
	{"codigo": "169", "nombre": "Colombia", "aliases": ["CO", "Colombia", "169"]},
	{"codigo": "850", "nombre": "Venezuela", "aliases": ["VE", "Venezuela", "850"]},
]

_DEFAULT_PROCEDENCIA_DEPARTAMENTOS_VENEZUELA = [
	{"codigo": "01", "nombre": "Departamento Venezuela", "pais_codigo": "850"},
]


def _procedencia_countries_from_frappe() -> list[dict]:
	"""Retorna únicamente Colombia (169) y Venezuela (850)."""
	return list(_DEFAULT_PROCEDENCIA_PAISES)


def _catalog_json_path() -> Path | None:
	rel_path = Path("Archivos siesa") / "Arquitectura refactorizacion siesa modulo contratación" / "codigos_siesa_completo.json"
	for parent in Path(__file__).resolve().parents:
		candidate = parent / rel_path
		if candidate.exists():
			return candidate
	return None


def _catalog_workbook_path() -> Path | None:
	rel_path = Path("Archivos siesa") / "madre_codigos.xlsx"
	for parent in Path(__file__).resolve().parents:
		candidate = parent / rel_path
		if candidate.exists():
			return candidate
	return None


def _bundled_procedencia_catalog_path() -> Path | None:
	for parent in Path(__file__).resolve().parents:
		candidate = parent / "hubgh" / "data" / "procedencia_siesa_fallback.json"
		if candidate.exists():
			return candidate
	return None


def _load_json_catalog(path: Path | None) -> dict:
	if not path:
		return {}
	try:
		with path.open("r", encoding="utf-8") as f:
			payload = json.load(f)
	except Exception:
		return {}
	return payload if isinstance(payload, dict) else {}


def _load_procedencia_catalog_from_workbook() -> dict:
	global _PROCEDENCIA_WORKBOOK_CACHE
	if _PROCEDENCIA_WORKBOOK_CACHE is not None:
		return _PROCEDENCIA_WORKBOOK_CACHE

	workbook_path = _catalog_workbook_path()
	if not workbook_path:
		_PROCEDENCIA_WORKBOOK_CACHE = {"paises": [], "departamentos": [], "ciudades": []}
		return _PROCEDENCIA_WORKBOOK_CACHE

	try:
		import openpyxl  # type: ignore
		wb = openpyxl.load_workbook(str(workbook_path), data_only=True)
	except Exception:
		_PROCEDENCIA_WORKBOOK_CACHE = {"paises": [], "departamentos": [], "ciudades": []}
		return _PROCEDENCIA_WORKBOOK_CACHE

	paises = []
	departamentos = []
	ciudades = []

	if "Pais" in wb.sheetnames:
		ws = wb["Pais"]
		for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
			code = str((row[0] if len(row) > 0 else "") or "").strip()
			name = str((row[1] if len(row) > 1 else "") or "").strip()
			if not code or not name:
				continue
			if code in {"169", "850"}:
				paises.append({"codigo": code, "nombre": name})
				continue
			if code.lower() in {"departamento", "pais"}:
				continue
			# Hoja Pais mezcla filas informativas con departamentos.
			if code.isdigit() and len(code) <= 3 and idx > 3:
				departamentos.append({"codigo": code.zfill(2), "nombre": name, "pais_codigo": "169"})

	if "Cuidad" in wb.sheetnames:
		ws = wb["Cuidad"]
		for row in ws.iter_rows(min_row=2, values_only=True):
			code = str((row[0] if len(row) > 0 else "") or "").strip()
			name = str((row[1] if len(row) > 1 else "") or "").strip()
			if not code or not name:
				continue
			if code.lower() == "cuidad":
				continue
			ciudades.append({"codigo": code.zfill(3), "nombre": name, "pais_codigo": "169"})

	_PROCEDENCIA_WORKBOOK_CACHE = {
		"paises": paises,
		"departamentos": departamentos,
		"ciudades": ciudades,
	}
	return _PROCEDENCIA_WORKBOOK_CACHE


def _load_procedencia_catalog() -> dict:
	json_path = _catalog_json_path()
	paises_default = _procedencia_countries_from_frappe() or list(_DEFAULT_PROCEDENCIA_PAISES)
	workbook_catalog = _load_procedencia_catalog_from_workbook()
	bundled_catalog = _load_json_catalog(_bundled_procedencia_catalog_path())
	if not json_path:
		return {
			"paises": workbook_catalog.get("paises") or bundled_catalog.get("paises") or paises_default,
			"departamentos": (workbook_catalog.get("departamentos") or bundled_catalog.get("departamentos") or []) + list(_DEFAULT_PROCEDENCIA_DEPARTAMENTOS_VENEZUELA),
			"ciudades": workbook_catalog.get("ciudades") or bundled_catalog.get("ciudades") or [],
		}

	payload = _load_json_catalog(json_path)

	paises = payload.get("paises") or payload.get("países") or payload.get("paises_muestra") or []
	if not isinstance(paises, list) or not paises:
		paises = paises_default
	else:
		allowed = {"169", "850"}
		filtered = []
		for row in paises:
			if not isinstance(row, dict):
				continue
			code = str(row.get("codigo") or row.get("code") or "").strip()
			if code in allowed:
				filtered.append(row)
		paises = filtered or paises_default

	departamentos = payload.get("departamentos") or []
	if not isinstance(departamentos, list):
		departamentos = []
	departamentos = [
		{**row, "pais_codigo": str(row.get("pais_codigo") or row.get("pais") or "169").strip() or "169"}
		for row in departamentos
		if isinstance(row, dict)
	]

	ciudades = payload.get("ciudades") or payload.get("ciudades_muestra") or []
	if not isinstance(ciudades, list):
		ciudades = []
	if workbook_catalog.get("ciudades"):
		ciudades = workbook_catalog.get("ciudades") or ciudades
	elif bundled_catalog.get("ciudades"):
		ciudades = bundled_catalog.get("ciudades") or ciudades

	return {
		"paises": paises or bundled_catalog.get("paises") or paises_default,
		"departamentos": (workbook_catalog.get("departamentos") or departamentos or bundled_catalog.get("departamentos") or []) + list(_DEFAULT_PROCEDENCIA_DEPARTAMENTOS_VENEZUELA),
		"ciudades": ciudades,
	}


def _normalize_catalog_rows(rows: list[dict], *, code_keys=("codigo", "code"), label_keys=("nombre", "name", "descripcion", "description")) -> list[dict[str, str]]:
	seen = set()
	normalized = []
	for row in rows or []:
		if not isinstance(row, dict):
			continue
		code = ""
		label = ""
		for key in code_keys:
			candidate = str(row.get(key) or "").strip()
			if candidate:
				code = candidate
				break
		for key in label_keys:
			candidate = str(row.get(key) or "").strip()
			if candidate:
				label = candidate
				break
		if not code or not label:
			continue
		if code in seen:
			continue
		seen.add(code)
		normalized.append({"code": code, "name": label})
	return normalized


def _normalize_catalog_code(raw_value: str | None, rows: list[dict], *, code_keys=("codigo", "code"), label_keys=("nombre", "name", "descripcion", "description")) -> str:
	value = (raw_value or "").strip()
	if not value:
		return ""

	def _norm(v):
		return " ".join(str(v or "").strip().lower().split())

	lookup = _norm(value)
	by_code = {}
	by_label = {}
	for row in rows or []:
		if not isinstance(row, dict):
			continue
		code = ""
		for key in code_keys:
			candidate = str(row.get(key) or "").strip()
			if candidate:
				code = candidate
				break
		if not code:
			continue
		by_code[_norm(code)] = code
		for key in label_keys:
			label = str(row.get(key) or "").strip()
			if label:
				by_label[_norm(label)] = code
		aliases = row.get("aliases") or []
		if isinstance(aliases, (list, tuple)):
			for alias in aliases:
				alias_norm = _norm(alias)
				if alias_norm:
					by_label[alias_norm] = code

	if lookup in by_code:
		return by_code[lookup]
	if lookup in by_label:
		return by_label[lookup]
	if " - " in value:
		left, right = [p.strip() for p in value.split(" - ", 1)]
		for part in (left, right):
			norm_part = _norm(part)
			if norm_part in by_code:
				return by_code[norm_part]
			if norm_part in by_label:
				return by_label[norm_part]

	return ""


def _normalize_colombia_city_code(raw_value: str | None, rows: list[dict]) -> str:
	"""Accepts SIESA city catalog code and common Colombian 5-digit municipality variants.

	SIESA sample uses 3-digit municipality codes (e.g. 001), while some clients may
	send a 5-digit value (e.g. 11001). In that case, we normalize to the last 3 digits
	when it matches the loaded city catalog.
	"""
	value = str(raw_value or "").strip()
	if not value:
		return ""

	direct = _normalize_catalog_code(value, rows)
	if direct:
		return direct

	just_digits = "".join(ch for ch in value if ch.isdigit())
	if len(just_digits) == 5:
		candidate = just_digits[-3:]
		fallback = _normalize_catalog_code(candidate, rows)
		if fallback:
			return fallback

	return ""


def _resolve_procedencia_codes(data: dict) -> dict[str, str]:
	catalog = _load_procedencia_catalog()
	raw_pais = str(data.get("procedencia_pais") or "").strip()
	raw_departamento = str(data.get("procedencia_departamento") or "").strip()
	raw_ciudad = str(data.get("procedencia_ciudad") or "").strip()

	pais = _normalize_catalog_code(raw_pais, catalog.get("paises") or [])
	if raw_pais and pais not in {"169", "850"}:
		frappe.throw("Procedencia país inválida. Solo aplica Colombia (169) o Venezuela (850).", frappe.ValidationError)
	departamento = _normalize_catalog_code(raw_departamento, catalog.get("departamentos") or [])
	ciudades_catalog = catalog.get("ciudades") or []
	ciudad = _normalize_catalog_code(raw_ciudad, ciudades_catalog)
	if pais == "169" and raw_ciudad and not ciudad:
		ciudad = _normalize_colombia_city_code(raw_ciudad, ciudades_catalog)

	if raw_pais and not pais:
		frappe.throw("Procedencia país inválida. Debes seleccionar un país válido del catálogo SIESA.", frappe.ValidationError)
	if raw_departamento and not departamento:
		frappe.throw("Procedencia departamento inválida. Debes seleccionar un departamento válido del catálogo SIESA.", frappe.ValidationError)
	if raw_ciudad and not ciudad:
		frappe.throw("Procedencia ciudad inválida. Debes seleccionar una ciudad válida del catálogo SIESA.", frappe.ValidationError)

	is_foreign = bool(pais == "850")

	if pais == "169":
		if not raw_departamento:
			frappe.throw("Para Colombia debes seleccionar departamento de procedencia.", frappe.ValidationError)
		if not raw_ciudad:
			frappe.throw("Para Colombia debes seleccionar municipio/ciudad de procedencia.", frappe.ValidationError)
	elif pais == "850":
		# Para Venezuela no forzamos depto/municipio de Colombia.
		departamento = ""
		ciudad = ""
	else:
		if not raw_departamento:
			departamento = ""
		if not raw_ciudad:
			ciudad = ""

	return {
		"pais": pais,
		"departamento": departamento,
		"ciudad": ciudad,
		"is_foreign": 1 if is_foreign else 0,
	}


def _resolve_banco_siesa_name(raw_value: str | None) -> str:
	banco = (raw_value or "").strip()
	if not banco:
		return ""

	if frappe.db.exists("Banco Siesa", banco):
		return banco

	if " - " in banco:
		left = banco.split(" - ", 1)[0].strip()
		if left and frappe.db.exists("Banco Siesa", left):
			return left

	match_by_description = frappe.db.get_value("Banco Siesa", {"description": banco}, "name")
	if match_by_description:
		return match_by_description

	upper_description = frappe.db.get_value("Banco Siesa", {"description": ["like", banco.upper()]}, "name")
	if upper_description:
		return upper_description

	return ""


def _default_bank_map():
	return {
		row["code"]: row["description"]
		for row in _DEFAULT_BANCOS_SIESA
	}


def _seed_default_bancos_if_missing() -> int:
	inserted = 0
	for row in _DEFAULT_BANCOS_SIESA:
		code = (row.get("code") or "").strip()
		description = (row.get("description") or "").strip()
		if not code or not description:
			continue
		if frappe.db.exists("Banco Siesa", code):
			continue
		frappe.get_doc(
			{
				"doctype": "Banco Siesa",
				"code": code,
				"description": description,
				"enabled": 1,
			}
		).insert(ignore_permissions=True)
		inserted += 1
	return inserted


def _ensure_default_bank_exists(code: str) -> str:
	code = (code or "").strip()
	if not code:
		return ""
	if frappe.db.exists("Banco Siesa", code):
		return code
	description = _default_bank_map().get(code)
	if not description:
		return ""
	frappe.get_doc(
		{
			"doctype": "Banco Siesa",
			"code": code,
			"description": description,
			"enabled": 1,
		}
	).insert(ignore_permissions=True)
	return code


def _split_apellidos(apellidos: str | None) -> tuple[str, str]:
	parts = [p.strip() for p in str(apellidos or "").split() if p and p.strip()]
	if not parts:
		return "", ""
	if len(parts) == 1:
		return parts[0], ""
	return parts[0], " ".join(parts[1:]).strip()

from hubgh.hubgh.onboarding_security import (
	enforce_onboarding_rate_limit,
	validate_candidate_duplicates,
	validate_onboarding_captcha,
)


def get_context(context):
	context.no_cache = 1
	context.page_title = "Candidato"


@frappe.whitelist(allow_guest=True)
def create_candidate(payload):
	logger = frappe.logger("hubgh.onboarding")
	if isinstance(payload, dict):
		data = payload
	elif isinstance(payload, str):
		try:
			data = json.loads(payload or "{}")
		except Exception:
			frappe.throw("Payload inválido: JSON malformado.", frappe.ValidationError)
	else:
		frappe.throw("Payload inválido: formato no soportado.", frappe.ValidationError)

	if not isinstance(data, dict):
		frappe.throw("Payload inválido: se esperaba un objeto JSON.", frappe.ValidationError)

	logger.info(
		"create_candidate:start",
		extra={
			"numero_documento": data.get("numero_documento"),
			"email": (data.get("email") or "").strip().lower(),
			"has_disponibilidad": 1 if data.get("disponibilidad") else 0,
		},
	)

	try:
		enforce_onboarding_rate_limit(identifier=data.get("numero_documento") or data.get("email"))
		validate_onboarding_captcha(data)
		validate_candidate_duplicates(
			numero_documento=data.get("numero_documento"),
			email=data.get("email"),
		)
	except Exception as exc:
		logger.warning(
			"create_candidate:pre_insert_validation_failed",
			extra={
				"numero_documento": data.get("numero_documento"),
				"email": (data.get("email") or "").strip().lower(),
				"error_type": type(exc).__name__,
				"error": str(exc),
			},
		)
		raise

	allowed = {
		"nombres",
		"apellidos",
		"primer_apellido",
		"segundo_apellido",
		"email",
		"celular",
		"telefono_fijo",
		"contacto_emergencia_nombre",
		"contacto_emergencia_telefono",
		"tipo_documento",
		"numero_documento",
		"fecha_nacimiento",
		"fecha_expedicion",
		"banco_siesa",
		"tipo_cuenta_bancaria",
		"numero_cuenta_bancaria",
		"grupo_sanguineo",
		"tiene_alergias",
		"descripcion_alergias",
		"talla_camisa",
		"talla_pantalon",
		"numero_zapatos",
		"talla_delantal",
		"personas_a_cargo",
		"ciudad",
		"localidad",
		"localidad_otras",
		"barrio",
		"direccion",
		"procedencia_pais",
		"procedencia_departamento",
		"procedencia_ciudad",
		"es_extranjero",
		"prefijo_cuenta_extranjero",
		"pais_residencia_siesa",
		"departamento_residencia_siesa",
		"ciudad_residencia_siesa",
		"disponibilidad",
	}

	banco = (data.get("banco_siesa") or "").strip()
	banco_resuelto = _resolve_banco_siesa_name(banco)
	if banco and not banco_resuelto:
		banco_resuelto = _ensure_default_bank_exists(banco)
	if banco and not banco_resuelto and " - " in banco:
		left_code = banco.split(" - ", 1)[0].strip()
		banco_resuelto = _ensure_default_bank_exists(left_code) or _resolve_banco_siesa_name(left_code)
	banco_exists_by_name = 1 if (banco and frappe.db.exists("Banco Siesa", banco)) else 0
	banco_exists_by_description = 1 if (banco and frappe.db.exists("Banco Siesa", {"description": banco})) else 0
	logger.info(
		"create_candidate:banco_received",
		extra={
			"banco_siesa": banco,
			"banco_siesa_resuelto": banco_resuelto,
			"banco_exists_by_name": banco_exists_by_name,
			"banco_exists_by_description": banco_exists_by_description,
			"tipo": type(data.get("banco_siesa")).__name__,
			"numero_documento": data.get("numero_documento"),
		},
	)
	if banco and not banco_resuelto:
		sample = frappe.get_all(
			"Banco Siesa",
			fields=["name", "description"],
			order_by="modified desc",
			limit_page_length=10,
			ignore_permissions=True,
		)
		logger.warning(
			"create_candidate:banco_not_found",
			extra={
				"banco_siesa": banco,
				"sample_catalog": sample,
			},
		)
		frappe.throw(
			"El banco seleccionado no existe en el catálogo de Banco Siesa. Selecciona un banco válido de la lista.",
			frappe.ValidationError,
		)

	if banco_resuelto:
		data["banco_siesa"] = banco_resuelto

	procedencia_codes = _resolve_procedencia_codes(data)
	if procedencia_codes.get("pais"):
		data["procedencia_pais"] = procedencia_codes["pais"]
	if procedencia_codes.get("departamento"):
		data["procedencia_departamento"] = procedencia_codes["departamento"]
	if procedencia_codes.get("ciudad"):
		data["procedencia_ciudad"] = procedencia_codes["ciudad"]

	if procedencia_codes.get("pais"):
		is_foreign = int(procedencia_codes.get("is_foreign") or 0)
		data["es_extranjero"] = is_foreign
		prefijo = (data.get("prefijo_cuenta_extranjero") or "").strip()
		if is_foreign:
			if not prefijo:
				frappe.throw(
					"Debes diligenciar el prefijo de cuenta para candidatos extranjeros.",
					frappe.ValidationError,
				)
			data["prefijo_cuenta_extranjero"] = prefijo
		else:
			data["prefijo_cuenta_extranjero"] = "NO APLICA"

	# Consistencia con exportadores SIESA actuales (residencia en códigos).
	if procedencia_codes.get("pais") == "169" and not (data.get("pais_residencia_siesa") or "").strip():
		data["pais_residencia_siesa"] = procedencia_codes["pais"]
	if procedencia_codes.get("pais") == "169" and procedencia_codes.get("departamento") and not (data.get("departamento_residencia_siesa") or "").strip():
		data["departamento_residencia_siesa"] = procedencia_codes["departamento"]
	if procedencia_codes.get("pais") == "169" and procedencia_codes.get("ciudad") and not (data.get("ciudad_residencia_siesa") or "").strip():
		data["ciudad_residencia_siesa"] = procedencia_codes["ciudad"]
	if procedencia_codes.get("pais") == "850" and not (data.get("pais_residencia_siesa") or "").strip():
		data["pais_residencia_siesa"] = "850"

	# Backfill de campos geográficos SIESA de nacimiento/expedición para evitar vacíos
	# en Captura Final cuando el candidato sí reportó procedencia.
	if procedencia_codes.get("pais") == "169":
		if not (data.get("pais_nacimiento_siesa") or "").strip():
			data["pais_nacimiento_siesa"] = "169"
		if not (data.get("departamento_nacimiento_siesa") or "").strip() and procedencia_codes.get("departamento"):
			data["departamento_nacimiento_siesa"] = procedencia_codes["departamento"]
		if not (data.get("ciudad_nacimiento_siesa") or "").strip() and procedencia_codes.get("ciudad"):
			data["ciudad_nacimiento_siesa"] = procedencia_codes["ciudad"]

		if not (data.get("pais_expedicion_siesa") or "").strip():
			data["pais_expedicion_siesa"] = "169"
		if not (data.get("departamento_expedicion_siesa") or "").strip() and (
			data.get("departamento_nacimiento_siesa") or procedencia_codes.get("departamento")
		):
			data["departamento_expedicion_siesa"] = (
				data.get("departamento_nacimiento_siesa") or procedencia_codes.get("departamento")
			)
		if not (data.get("ciudad_expedicion_siesa") or "").strip() and (
			data.get("ciudad_nacimiento_siesa") or procedencia_codes.get("ciudad")
		):
			data["ciudad_expedicion_siesa"] = data.get("ciudad_nacimiento_siesa") or procedencia_codes.get("ciudad")

	primer_apellido = (data.get("primer_apellido") or "").strip()
	segundo_apellido = (data.get("segundo_apellido") or "").strip()
	apellidos = (data.get("apellidos") or "").strip()
	if apellidos and (not primer_apellido or not segundo_apellido):
		primer_split, segundo_split = _split_apellidos(apellidos)
		if not primer_apellido:
			primer_apellido = primer_split
		if not segundo_apellido:
			segundo_apellido = segundo_split
		data["primer_apellido"] = primer_apellido
		data["segundo_apellido"] = segundo_apellido

	doc = frappe.get_doc({"doctype": "Candidato"})
	for key in allowed:
		if key in data:
			doc.set(key, data.get(key))

	table_field = next(
		(df.fieldname for df in doc.meta.fields if df.fieldtype == "Table" and df.options == "Candidato Disponibilidad"),
		None,
	)
	if table_field and data.get("disponibilidad"):
		doc.set(table_field, [])
		for row in data.get("disponibilidad"):
			doc.append(
				table_field,
				{
					"dia": row.get("dia"),
					"hora_inicio": row.get("hora_inicio"),
					"hora_fin": row.get("hora_fin"),
				},
			)

	try:
		doc.insert(ignore_permissions=True)
	except Exception:
		frappe.log_error(
			title="Onboarding create_candidate insert failed",
			message=(
				f"banco_payload={banco}\n"
				f"banco_resuelto={banco_resuelto}\n"
				f"banco_exists_by_name={banco_exists_by_name}\n"
				f"banco_exists_by_description={banco_exists_by_description}\n"
				f"traceback=\n{frappe.get_traceback()}"
			),
		)
		raise
	return {
		"name": doc.name,
		"user": doc.user,
		"login_user": getattr(doc.flags, "onboarding_login_user", None) or doc.user or data.get("email") or data.get("numero_documento"),
		"initial_password": getattr(doc.flags, "onboarding_initial_password", None),
		"password_must_be_changed": 1,
		"user_created": 1 if getattr(doc.flags, "onboarding_user_created", False) else 0,
	}


@frappe.whitelist(allow_guest=True)
def get_bancos_siesa():
	logger = frappe.logger("hubgh.onboarding")
	seeded = _seed_default_bancos_if_missing()
	if seeded:
		logger.info("get_bancos_siesa:seed_defaults", extra={"inserted": seeded})

	bancos = frappe.get_all(
		"Banco Siesa",
		filters={"enabled": 1},
		fields=["name", "description"],
		order_by="description asc",
		ignore_permissions=True,
	)
	if bancos:
		return bancos

	all_bancos = frappe.get_all(
		"Banco Siesa",
		fields=["name", "description"],
		order_by="description asc",
		ignore_permissions=True,
	)
	if all_bancos:
		return all_bancos

	# Fallback sin escritura en GET: exponer lista mínima cuando la tabla está vacía.
	return [{"name": row["code"], "description": row["description"]} for row in _DEFAULT_BANCOS_SIESA]


@frappe.whitelist(allow_guest=True)
def get_onboarding_csrf_token():
	"""Retorna CSRF token vigente para el flujo guest de onboarding."""
	return {"csrf_token": get_csrf_token()}


@frappe.whitelist(allow_guest=True)
def get_procedencia_siesa_catalog():
	catalog = _load_procedencia_catalog()
	paises = _normalize_catalog_rows(catalog.get("paises") or [])
	if not paises:
		paises = [{"code": row["codigo"], "name": row["nombre"]} for row in _DEFAULT_PROCEDENCIA_PAISES]
	paises = [row for row in paises if str(row.get("code") or "") in {"169", "850"}]

	departamentos = []
	for row in catalog.get("departamentos") or []:
		if not isinstance(row, dict):
			continue
		code = str(row.get("codigo") or row.get("code") or "").strip()
		name = str(row.get("nombre") or row.get("name") or row.get("descripcion") or row.get("description") or "").strip()
		pais_codigo = str(row.get("pais_codigo") or row.get("pais") or "").strip() or "169"
		if not code or not name:
			continue
		departamentos.append({"code": code, "name": name, "pais_codigo": pais_codigo})
	ciudades = []
	for row in catalog.get("ciudades") or []:
		if not isinstance(row, dict):
			continue
		code = str(row.get("codigo") or row.get("code") or "").strip()
		name = str(row.get("nombre") or row.get("name") or row.get("descripcion") or row.get("description") or "").strip()
		departamento_codigo = str(row.get("departamento_codigo") or row.get("departamento") or "").strip()
		pais_codigo = str(row.get("pais_codigo") or row.get("pais") or "").strip() or "169"
		if not code or not name:
			continue
		ciudades.append(
			{
				"code": code,
				"name": name,
				"departamento_codigo": departamento_codigo,
				"pais_codigo": pais_codigo,
			}
		)

	departamentos.sort(key=lambda row: row.get("name") or "")
	ciudades.sort(key=lambda row: row.get("name") or "")

	return {
		"paises": paises,
		"departamentos": departamentos,
		"ciudades": ciudades,
	}
