"""
Payroll Export Prenomina - Generate Excel export in Prenomina format.

Exports TP-approved payroll data to Excel format matching the target template
(Prenomina_Febrero_Home_Burgers.xlsx). Handles column mapping, employee totals,
and period consolidation for payroll processing.

Sprint 4: Core Prenomina export with production-ready format.
"""

import frappe
from frappe import _
from frappe.utils import now_datetime, getdate, flt, cstr, get_site_path
from typing import List, Dict, Any, Optional
import json
import os
from datetime import datetime

from hubgh.hubgh.payroll_employee_compat import (
	build_employee_parametrization_message,
	get_payroll_employee_context,
	normalize_tipo_jornada,
)
from hubgh.hubgh.payroll_permissions import enforce_payroll_access

try:
	import openpyxl
	from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
	from openpyxl.utils.dataframe import dataframe_to_rows
except ImportError:
	openpyxl = None
	Font = Alignment = Border = Side = PatternFill = None


# Target Prenomina column mapping based on template analysis
PRENOMINA_COLUMNS = {
	"Documento": "document_number",
	"Nombre Empleado": "employee_name", 
	"Punto de Venta": "pdv_location",
	"HD": "horas_diurnas",
	"HN": "horas_nocturnas",
	"HED": "horas_extras_diurnas",
	"HEN": "horas_extras_nocturnas",
	"Auxilio Transporte": "aux_transporte",
	"Auxilio Dominical": "aux_dominical",
	"HOME 12": "subsidio_home12",
	"Recargo Nocturno": "recargo_nocturno",
	"Recargo Dominical": "recargo_dominical",
	"Bonificaciones": "bonificaciones",
	"Comisiones": "comisiones",
	"Total Devengado": "total_devengado",
	"Sanitas": "desc_sanitas",
	"Libranzas": "desc_libranzas",
	"Payflow": "desc_payflow",
	"Embargos": "desc_embargos",
	"Otros Descuentos": "desc_otros",
	"Total Deducciones": "total_deducciones",
	"Neto a Pagar": "neto_a_pagar"
}

# Excel styling constants
HEADER_STYLE = {
	"font": Font(bold=True, color="FFFFFF") if Font else None,
	"fill": PatternFill(start_color="366092", end_color="366092", fill_type="solid") if PatternFill else None,
	"alignment": Alignment(horizontal="center", vertical="center") if Alignment else None,
}

MONEY_FORMAT = "#,##0"
HOUR_FORMAT = "0.00"


class PrenominaExportService:
	"""Service for generating Prenomina Excel exports."""
	
	def __init__(self):
		if not openpyxl:
			raise ImportError("openpyxl library is required for Excel export")
		
		self.default_values = {
			"pdv_location": "Sin PDV",
			"aux_transporte": 140606,  # Standard transport allowance
			"aux_dominical": 0,
			"subsidio_home12": 0,
			"recargo_nocturno": 0,
			"recargo_dominical": 0
		}

	def get_exports_directory(self, output_dir: str = None) -> str:
		if output_dir:
			return os.path.realpath(output_dir)
		return os.path.realpath(os.path.join(get_site_path(), "private", "files", "prenomina_exports"))
	
	def generate_prenomina_export(
		self, batch_name: str, output_dir: str = None, jornada_filter: str = None
	) -> Dict[str, Any]:
		"""
		Generate Prenomina Excel export for a specific batch.
		
		Args:
			batch_name: Name of the PayrollImportBatch
			output_dir: Optional output directory (defaults to site files)
			
		Returns:
			Dict with export status, file path, and summary
		"""
		
		try:
			# Get batch information
			batch_doc = frappe.get_doc("Payroll Import Batch", batch_name)
			if not batch_doc:
				return {"status": "error", "message": f"Lote {batch_name} no encontrado"}
			
			# Get TP-approved lines for this batch
			lines, jornada_context = self._get_tp_approved_lines(batch_name, jornada_filter=jornada_filter)
			if not lines:
				message = f"No hay líneas aprobadas TP en el lote {batch_name}"
				warning = self._build_jornada_filter_warning(jornada_context)
				if warning:
					message = f"{message}. {warning}"
				return {"status": "error", "message": message, "jornada_filter": jornada_context.get("canonical_filter") or "Todas"}
			
			# Consolidate by employee for Prenomina format
			employee_data = self._consolidate_employee_data(lines)
			
			# Generate Excel file
			file_path = self._create_excel_export(employee_data, batch_doc, output_dir)
			
			# Calculate summary statistics
			summary = self._calculate_export_summary(employee_data, batch_doc)
			
			return {
				"status": "success",
				"message": f"Prenomina generada exitosamente para {len(employee_data)} empleados",
				"file_path": file_path,
				"batch_name": batch_name,
				"employee_count": len(employee_data),
				"period": batch_doc.nomina_period or "Sin Período",
				"summary": summary,
				"jornada_filter": jornada_context.get("canonical_filter") or "Todas",
				"jornada_filter_warning": self._build_jornada_filter_warning(jornada_context),
			}
			
		except Exception as e:
			frappe.log_error(f"Error generating prenomina export for {batch_name}: {str(e)}")
			return {
				"status": "error",
				"message": f"Error generando prenomina: {str(e)}",
				"file_path": None
			}
	
	def _get_tp_approved_lines(self, batch_name: str, jornada_filter: str = None):
		"""Get all TP-approved lines for the batch."""
		
		filters = {
			"batch": batch_name,
			"status": ["in", ["Válido", "Procesado"]],
			"tc_status": "Aprobado",
			"tp_status": "Aprobado"
		}
		
		lines = frappe.get_all("Payroll Import Line",
			filters=filters,
			fields=[
				"name", "matched_employee", "employee_id", "employee_name",
				"novedad_type", "novedad_date", "quantity", "amount",
				"rule_applied", "rule_notes", "source_row_data"
			],
			order_by="employee_name asc, novedad_type asc"
		)
		return self._filter_lines_by_jornada(lines, jornada_filter)

	def _filter_lines_by_jornada(self, lines: List[Dict], jornada_filter: str = None):
		canonical_filter = normalize_tipo_jornada(jornada_filter)
		context = {
			"canonical_filter": canonical_filter,
			"missing_employee_count": 0,
			"missing_employee_labels": [],
		}
		if not lines:
			return lines, context

		filtered_lines = []
		missing_labels = set()
		employee_cache = {}
		for line in lines:
			emp_key = line.get("matched_employee") or line.get("employee_id")
			if not emp_key:
				filtered_lines.append(line)
				continue

			if emp_key not in employee_cache:
				employee_context = get_payroll_employee_context(emp_key)
				employee_cache[emp_key] = {
					"tipo_jornada": normalize_tipo_jornada(employee_context.get("tipo_jornada")),
					"label": employee_context.get("employee_name") or emp_key,
				}

			employee_info = employee_cache[emp_key]
			tipo_jornada = employee_info.get("tipo_jornada")
			if not tipo_jornada:
				missing_labels.add(str(employee_info.get("label") or emp_key))

			if canonical_filter and tipo_jornada != canonical_filter:
				continue

			filtered_lines.append(line)

		context["missing_employee_labels"] = sorted(missing_labels)
		context["missing_employee_count"] = len(missing_labels)
		return filtered_lines, context

	def _build_jornada_filter_warning(self, jornada_context: Dict[str, Any]) -> str | None:
		missing_count = jornada_context.get("missing_employee_count", 0)
		if not missing_count:
			return None

		canonical_filter = jornada_context.get("canonical_filter")
		base_message = (
			f"{missing_count} empleado(s) no tienen Tipo de Jornada canónico parametrizado en Ficha Empleado."
		)
		if canonical_filter:
			return f"{base_message} No se incluyen cuando filtrás por {canonical_filter}."
		return base_message
	
	def _consolidate_employee_data(self, lines: List[Dict]) -> List[Dict[str, Any]]:
		"""
		Consolidate lines by employee into Prenomina format.
		
		Returns list of employee records ready for Excel export.
		"""
		
		employee_map = {}
		
		for line in lines:
			emp_key = line.get("matched_employee") or line.get("employee_id")
			if not emp_key:
				continue
			
			if emp_key not in employee_map:
				# Initialize employee record with Prenomina structure
				employee_map[emp_key] = self._create_employee_record(line)
			
			# Add this line's data to the employee record
			self._add_line_to_employee(employee_map[emp_key], line)
		
		# Calculate totals for each employee
		result = []
		for emp_record in employee_map.values():
			self._calculate_employee_totals(emp_record)
			result.append(emp_record)
		
		# Sort by employee name for consistent export
		result.sort(key=lambda x: x["employee_name"])
		return result
	
	def _create_employee_record(self, line: Dict) -> Dict[str, Any]:
		"""Create initial employee record structure."""
		
		emp_key = line.get("matched_employee") or line.get("employee_id")
		employee_context = get_payroll_employee_context(emp_key)
		
		# Get employee master data for additional fields
		document_number = emp_key  # Fallback
		pdv_location = self.default_values["pdv_location"]
		base_hourly_rate = 15000
		parameterization_warning = build_employee_parametrization_message(
			employee_context,
			required_fields=["document_number", "pdv", "contrato", "salary", "monthly_hours"],
		)
		
		document_number = employee_context.get("document_number") or emp_key
		branch = employee_context.get("branch") or ""
		if branch:
			pdv_location = branch
		salary = flt(employee_context.get("salary") or 0)
		monthly_hours = flt(employee_context.get("monthly_hours") or 220)
		if salary > 0 and monthly_hours > 0:
			base_hourly_rate = salary / monthly_hours
		
		return {
			"document_number": document_number,
			"employee_name": line.get("employee_name", "Sin Nombre"),
			"employee_id": emp_key,
			"pdv_location": pdv_location,
			"base_hourly_rate": base_hourly_rate,
			"parameterization_warning": parameterization_warning,
			
			# Hour columns
			"horas_diurnas": 0,
			"horas_nocturnas": 0, 
			"horas_extras_diurnas": 0,
			"horas_extras_nocturnas": 0,
			
			# Auxilio columns
			"aux_transporte": self.default_values["aux_transporte"],
			"aux_dominical": 0,
			"subsidio_home12": 0,
			
			# Recargo columns
			"recargo_nocturno": 0,
			"recargo_dominical": 0,
			
			# Other devengos
			"bonificaciones": 0,
			"comisiones": 0,
			
			# Deduction columns
			"desc_sanitas": 0,
			"desc_libranzas": 0,
			"desc_payflow": 0,
			"desc_embargos": 0,
			"desc_otros": 0,
			
			# Calculated totals (will be computed)
			"total_devengado": 0,
			"total_deducciones": 0,
			"neto_a_pagar": 0,
			
			# Metadata for tracking
			"line_count": 0,
			"novelty_types": set()
		}
	
	def _add_line_to_employee(self, emp_record: Dict, line: Dict):
		"""Add a single line's data to the employee record."""
		
		novelty_type = line.get("novedad_type", "")
		quantity = flt(line.get("quantity", 0))
		amount = flt(line.get("amount", 0))
		
		emp_record["line_count"] += 1
		emp_record["novelty_types"].add(novelty_type)
		
		# Map novelty types to Prenomina columns
		if novelty_type == "HD":
			emp_record["horas_diurnas"] += quantity
		elif novelty_type == "HN":
			emp_record["horas_nocturnas"] += quantity
			# Calculate nocturnal recargo
			emp_record["recargo_nocturno"] += self._calculate_nocturnal_recargo(
				quantity,
				emp_record.get("base_hourly_rate"),
			)
		elif novelty_type == "HED":
			emp_record["horas_extras_diurnas"] += quantity
		elif novelty_type == "HEN":
			emp_record["horas_extras_nocturnas"] += quantity
			# Extra nocturnal has both recargos
			emp_record["recargo_nocturno"] += self._calculate_nocturnal_recargo(
				quantity,
				emp_record.get("base_hourly_rate"),
			)
		
		# Auxiliary mapping
		elif novelty_type == "AUX-TRANSPORTE":
			emp_record["aux_transporte"] += amount
		elif novelty_type == "AUX-DOMINICAL":
			emp_record["aux_dominical"] += amount
		elif novelty_type == "AUX-HOME12" or "HOME12" in novelty_type:
			emp_record["subsidio_home12"] += amount
		
		# Bonifications and commissions
		elif novelty_type in ["BONIFICACION", "BONIF", "PRIMA"]:
			emp_record["bonificaciones"] += amount
		elif novelty_type in ["COMISION", "COMIS"]:
			emp_record["comisiones"] += amount
		
		# Deductions mapping
		elif "SANITAS" in novelty_type:
			emp_record["desc_sanitas"] += abs(amount)  # Ensure positive for deductions
		elif "LIBRANZA" in novelty_type:
			emp_record["desc_libranzas"] += abs(amount)
		elif "PAYFLOW" in novelty_type:
			emp_record["desc_payflow"] += abs(amount)
		elif "EMBARGO" in novelty_type:
			emp_record["desc_embargos"] += abs(amount)
		elif novelty_type.startswith("DESC-") or "DEDUC" in novelty_type:
			emp_record["desc_otros"] += abs(amount)
		
		# Dominical recargo calculation
		if self._is_dominical_line(line):
			emp_record["recargo_dominical"] += self._calculate_dominical_recargo(
				quantity,
				novelty_type,
				emp_record.get("base_hourly_rate"),
			)
	
	def _calculate_nocturnal_recargo(self, hours: float, base_rate: float = None) -> float:
		"""Calculate nocturnal recargo (25% additional)."""
		base_rate = flt(base_rate or 15000)
		return hours * base_rate * 0.25
	
	def _calculate_dominical_recargo(self, hours: float, novelty_type: str, base_rate: float = None) -> float:
		"""Calculate dominical recargo (100% additional)."""
		base_rate = flt(base_rate or 15000)
		if novelty_type in ["HD", "HN", "HED", "HEN"]:
			return hours * base_rate * 1.0
		return 0
	
	def _is_dominical_line(self, line: Dict) -> bool:
		"""Check if the line represents dominical work."""
		try:
			novelty_date = line.get("novedad_date")
			if not novelty_date:
				return False
			
			date_obj = getdate(novelty_date)
			return date_obj.weekday() == 6  # Sunday
		except Exception:
			return False
	
	def _calculate_employee_totals(self, emp_record: Dict):
		"""Calculate total devengado, deducciones, and neto for employee."""
		
		# Total devengado = all positive amounts
		devengado = (
			emp_record["aux_transporte"] +
			emp_record["aux_dominical"] +
			emp_record["subsidio_home12"] +
			emp_record["recargo_nocturno"] +
			emp_record["recargo_dominical"] +
			emp_record["bonificaciones"] +
			emp_record["comisiones"]
		)
		
		# Add base salary equivalent for hours worked
		# This should come from employee master, using simplified calculation
		total_hours = (
			emp_record["horas_diurnas"] +
			emp_record["horas_nocturnas"] +
			emp_record["horas_extras_diurnas"] + 
			emp_record["horas_extras_nocturnas"]
		)
		base_rate = flt(emp_record.get("base_hourly_rate") or 15000)
		hour_salary = total_hours * base_rate
		devengado += hour_salary
		
		# Total deducciones = all deduction amounts
		deducciones = (
			emp_record["desc_sanitas"] +
			emp_record["desc_libranzas"] +
			emp_record["desc_payflow"] +
			emp_record["desc_embargos"] +
			emp_record["desc_otros"]
		)
		
		# Neto = devengado - deducciones
		neto = devengado - deducciones
		
		emp_record["total_devengado"] = devengado
		emp_record["total_deducciones"] = deducciones
		emp_record["neto_a_pagar"] = neto
		
		# Convert novelty_types set to list for JSON serialization
		emp_record["novelty_types"] = list(emp_record["novelty_types"])
	
	def _create_excel_export(self, employee_data: List[Dict], batch_doc, output_dir: str = None) -> str:
		"""
		Create Excel file with Prenomina format.
		
		Returns the file path of the generated Excel file.
		"""
		
		# Determine output path
		output_dir = self.get_exports_directory(output_dir)
		
		# Ensure output directory exists
		os.makedirs(output_dir, exist_ok=True)
		
		# Generate filename
		period = batch_doc.nomina_period or "Sin_Periodo"
		timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
		filename = f"Prenomina_{period}_{batch_doc.name}_{timestamp}.xlsx"
		file_path = os.path.join(output_dir, filename)
		
		# Create workbook
		wb = openpyxl.Workbook()
		ws = wb.active
		ws.title = f"Prenomina {period}"
		
		# Write headers
		headers = list(PRENOMINA_COLUMNS.keys())
		for col, header in enumerate(headers, 1):
			cell = ws.cell(row=1, column=col, value=header)
			cell.font = HEADER_STYLE["font"]
			cell.fill = HEADER_STYLE["fill"]
			cell.alignment = HEADER_STYLE["alignment"]
		
		# Write employee data
		for row, emp_data in enumerate(employee_data, 2):
			for col, column_key in enumerate(PRENOMINA_COLUMNS.keys(), 1):
				field_name = PRENOMINA_COLUMNS[column_key]
				value = emp_data.get(field_name, "")
				
				# Format values appropriately
				if isinstance(value, (int, float)) and value != 0:
					if field_name in ["horas_diurnas", "horas_nocturnas", "horas_extras_diurnas", "horas_extras_nocturnas"]:
						# Hour formatting
						cell = ws.cell(row=row, column=col, value=value)
						cell.number_format = HOUR_FORMAT
					else:
						# Money formatting
						cell = ws.cell(row=row, column=col, value=value)
						cell.number_format = MONEY_FORMAT
				else:
					# Text or zero values
					display_value = value if value != 0 else ""
					ws.cell(row=row, column=col, value=display_value)
		
		# Auto-adjust column widths
		for column in ws.columns:
			max_length = 0
			column_letter = column[0].column_letter
			for cell in column:
				try:
					if len(str(cell.value)) > max_length:
						max_length = len(str(cell.value))
				except:
					pass
			adjusted_width = min(max_length + 2, 20)
			ws.column_dimensions[column_letter].width = adjusted_width
		
		# Add summary row at the bottom
		summary_row = len(employee_data) + 3
		ws.cell(row=summary_row, column=1, value="TOTALES").font = Font(bold=True)
		
		# Calculate and add column totals
		total_devengado = sum(emp.get("total_devengado", 0) for emp in employee_data)
		total_deducciones = sum(emp.get("total_deducciones", 0) for emp in employee_data)
		total_neto = sum(emp.get("neto_a_pagar", 0) for emp in employee_data)
		
		# Find total columns and add totals
		for col, column_key in enumerate(PRENOMINA_COLUMNS.keys(), 1):
			if column_key == "Total Devengado":
				cell = ws.cell(row=summary_row, column=col, value=total_devengado)
				cell.font = Font(bold=True)
				cell.number_format = MONEY_FORMAT
			elif column_key == "Total Deducciones":
				cell = ws.cell(row=summary_row, column=col, value=total_deducciones)
				cell.font = Font(bold=True)
				cell.number_format = MONEY_FORMAT
			elif column_key == "Neto a Pagar":
				cell = ws.cell(row=summary_row, column=col, value=total_neto)
				cell.font = Font(bold=True)
				cell.number_format = MONEY_FORMAT
		
		# Add metadata sheet
		metadata_ws = wb.create_sheet("Metadata")
		metadata = [
			["Generado", now_datetime().strftime("%Y-%m-%d %H:%M:%S")],
			["Lote", batch_doc.name],
			["Período", batch_doc.nomina_period or "Sin Período"],
			["Empleados", len(employee_data)],
			["Total Devengado", total_devengado],
			["Total Deducciones", total_deducciones],
			["Neto Total", total_neto],
			["Usuario", frappe.session.user]
		]
		
		for row, (label, value) in enumerate(metadata, 1):
			metadata_ws.cell(row=row, column=1, value=label).font = Font(bold=True)
			metadata_ws.cell(row=row, column=2, value=value)
		
		# Save the file
		wb.save(file_path)
		
		return file_path
	
	def _calculate_export_summary(self, employee_data: List[Dict], batch_doc) -> Dict[str, Any]:
		"""Calculate summary statistics for the export."""
		
		total_devengado = sum(emp.get("total_devengado", 0) for emp in employee_data)
		total_deducciones = sum(emp.get("total_deducciones", 0) for emp in employee_data)
		total_neto = sum(emp.get("neto_a_pagar", 0) for emp in employee_data)
		
		# Hour totals
		total_hours = {
			"diurnas": sum(emp.get("horas_diurnas", 0) for emp in employee_data),
			"nocturnas": sum(emp.get("horas_nocturnas", 0) for emp in employee_data),
			"extras_diurnas": sum(emp.get("horas_extras_diurnas", 0) for emp in employee_data),
			"extras_nocturnas": sum(emp.get("horas_extras_nocturnas", 0) for emp in employee_data)
		}
		
		# Top employees by neto
		top_employees = sorted(employee_data, 
							  key=lambda x: x.get("neto_a_pagar", 0), 
							  reverse=True)[:5]
		
		return {
			"employee_count": len(employee_data),
			"total_devengado": total_devengado,
			"total_deducciones": total_deducciones,
			"total_neto": total_neto,
			"average_neto": total_neto / len(employee_data) if employee_data else 0,
			"total_hours": total_hours,
			"top_employees": [
				{
					"name": emp.get("employee_name"),
					"neto": emp.get("neto_a_pagar", 0)
				}
				for emp in top_employees
			],
			"generation_date": now_datetime().strftime("%Y-%m-%d %H:%M:%S"),
			"batch": batch_doc.name,
			"period": batch_doc.nomina_period or "Sin Período",
			"parameterization_warnings": [
				emp.get("parameterization_warning")
				for emp in employee_data
				if emp.get("parameterization_warning")
			],
		}


# =============================================================================
# Public API Functions
# =============================================================================

@frappe.whitelist()
def generate_prenomina_export(batch_name, output_dir=None, jornada_filter=None):
	"""
	API endpoint to generate Prenomina export for a batch.
	
	Args:
		batch_name: Name of the PayrollImportBatch
		output_dir: Optional output directory
	"""
	enforce_payroll_access("tp_tray")
	
	try:
		service = PrenominaExportService()
		return service.generate_prenomina_export(batch_name, output_dir, jornada_filter=jornada_filter)
		
	except ImportError as e:
		return {
			"status": "error",
			"message": "openpyxl library not available. Please install: pip install openpyxl"
		}
	except Exception as e:
		frappe.log_error(f"Error in prenomina export API: {str(e)}")
		return {
			"status": "error",
			"message": f"Error generando prenomina: {str(e)}"
		}


@frappe.whitelist()
def get_prenomina_preview(batch_name, limit=10, jornada_filter=None):
	"""
	API endpoint to get a preview of Prenomina data without generating the file.
	
	Args:
		batch_name: Name of the PayrollImportBatch
		limit: Number of employees to preview
	"""
	enforce_payroll_access("tp_tray")
	
	try:
		service = PrenominaExportService()
		
		# Get TP-approved lines
		lines, jornada_context = service._get_tp_approved_lines(batch_name, jornada_filter=jornada_filter)
		if not lines:
			message = "No hay líneas aprobadas TP"
			warning = service._build_jornada_filter_warning(jornada_context)
			if warning:
				message = f"{message}. {warning}"
			return {"status": "error", "message": message, "jornada_filter": jornada_context.get("canonical_filter") or "Todas"}
		
		# Consolidate employee data
		employee_data = service._consolidate_employee_data(lines)
		
		# Limit for preview
		preview_data = employee_data[:int(limit)]
		
		# Get batch info
		batch_doc = frappe.get_doc("Payroll Import Batch", batch_name)
		summary = service._calculate_export_summary(employee_data, batch_doc)
		
		return {
			"status": "success",
			"preview_data": preview_data,
			"total_employees": len(employee_data),
			"showing_count": len(preview_data),
			"summary": summary,
			"columns": list(PRENOMINA_COLUMNS.keys()),
			"jornada_filter": jornada_context.get("canonical_filter") or "Todas",
			"jornada_filter_warning": service._build_jornada_filter_warning(jornada_context),
		}
		
	except Exception as e:
		frappe.log_error(f"Error in prenomina preview: {str(e)}")
		return {
			"status": "error",
			"message": f"Error generando preview: {str(e)}"
		}


@frappe.whitelist() 
def download_prenomina_file(file_path):
	"""
	API endpoint to download a generated Prenomina file.
	
	Args:
		file_path: Path to the generated Prenomina file
	"""
	enforce_payroll_access("tp_tray")
	
	try:
		service = PrenominaExportService()
		exports_dir = service.get_exports_directory()
		resolved_path = os.path.realpath(file_path)
		if not resolved_path.startswith(exports_dir + os.sep) and resolved_path != exports_dir:
			return {"status": "error", "message": "Ruta de prenomina no valida"}

		if not os.path.exists(resolved_path):
			return {"status": "error", "message": "Archivo no encontrado"}
		
		# Return file for download
		filename = os.path.basename(resolved_path)
		
		with open(resolved_path, 'rb') as f:
			file_content = f.read()
		
		# Set up file download response
		frappe.local.response.filename = filename
		frappe.local.response.filecontent = file_content
		frappe.local.response.type = "download"
		frappe.local.response.headers = {
			"Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
			"Content-Disposition": f"attachment; filename={filename}"
		}
		
		return {"status": "success", "message": "Descarga iniciada"}
		
	except Exception as e:
		frappe.log_error(f"Error downloading prenomina file: {str(e)}")
		return {"status": "error", "message": f"Error descargando archivo: {str(e)}"}


def get_prenomina_service() -> PrenominaExportService:
	"""Get singleton instance of PrenominaExportService."""
	return PrenominaExportService()
