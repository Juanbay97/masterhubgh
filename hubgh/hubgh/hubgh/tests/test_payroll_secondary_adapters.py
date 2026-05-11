"""Tests de los adapters secundarios: Payflow, Fincomercio, FONGIGA.

Sin Frappe: cada test arma un workbook openpyxl en memoria con la
estructura observada en los archivos reales y verifica matches(),
detect_period() y parse().
"""

from __future__ import annotations

import unittest
from datetime import datetime

from openpyxl import Workbook

from hubgh.hubgh.payroll.adapters import fincomercio, fongiga, payflow


# ──────────────────────────────────────────────────────────────────────
# Payflow
# ──────────────────────────────────────────────────────────────────────

class PayflowTests(unittest.TestCase):
	def _wb(self) -> Workbook:
		wb = Workbook()
		ws = wb.active
		ws.title = "Payflow - Resumen"
		# 3 filas vacías arriba, headers en la 4
		for _ in range(3):
			ws.append([])
		ws.append(["Nombre", "Apellidos", "Cedula", "Id de empleado", "Nombre empresa", "NIT empresa", "Importe total a deducir $"])
		ws.append(["Shirley", "Florez Mora", 586516, "6", "Comidas Varpel SAS", "900808926-9", 702000])
		ws.append(["Ronald", "Garcia", 1116317, "", "Comidas Varpel SAS", "900808926-9", 569000])
		ws.append(["Cero", "Cero", 9999999, "", "Comidas Varpel SAS", "900808926-9", 0])  # se ignora

		ws_d = wb.create_sheet("Payflow - Detalles")
		for _ in range(3):
			ws_d.append([])
		ws_d.append(["Nombre", "Apellidos", "Cedula", "Id de empleado", "Nombre empresa", "NIT empresa", "ID transacción", "Fecha", "Importe"])
		ws_d.append(["Shirley", "Florez Mora", "586516", "6", "Comidas", "900808926-9", "tx1", datetime(2026, 2, 16), 702000])
		ws_d.append(["Ronald", "Garcia", "1116317", "", "Comidas", "900808926-9", "tx2", datetime(2026, 2, 5), 142000])
		return wb

	def test_matches_score(self):
		meta = {"filename": "CO-RD-010226-280226-900808926-9.xlsx", "sheets": ["Payflow - Resumen", "Payflow - Detalles"]}
		self.assertEqual(payflow.matches(meta), 3)

	def test_matches_zero_when_unrelated(self):
		self.assertEqual(payflow.matches({"filename": "x.xlsx", "sheets": ["Hoja1"]}), 0)

	def test_detect_period_from_detalles(self):
		self.assertEqual(payflow.detect_period(self._wb()), (2026, 2))

	def test_parse_from_resumen_skips_zero_values(self):
		results = list(payflow.parse(self._wb()))
		self.assertEqual(len(results), 2)
		self.assertTrue(all(r.tipo_novedad == "ADELANTO_NOMINA_PAYFLOW" for r in results))
		self.assertEqual(results[0].documento_identidad, "586516")
		self.assertEqual(results[0].valor, 702000)
		self.assertEqual(results[0].unidad, "cop")


# ──────────────────────────────────────────────────────────────────────
# Fincomercio
# ──────────────────────────────────────────────────────────────────────

class FincomercioTests(unittest.TestCase):
	def _wb(self) -> Workbook:
		wb = Workbook()
		ws = wb.active
		ws.title = "Descuentos Nomina Agrupado"
		# fila 1-2 vacías, fila 3-6 cabecera del reporte, fila 7 headers reales.
		for i in range(2):
			ws.append([])
		ws.append(["", "", "PAGADURIA: (MICRO) COMIDA"])
		ws.append(["", "", "FECHA DE CORTE: 03/02/2026"])
		ws.append(["", "", "PERIODO: 15/02/2026"])
		ws.append(["", "", "REPORTE No :: 541816"])
		ws.append(["", "IDENTIFICACION", "", "CODIGO INTERNO", "APELLIDOS", "NOMBRES", "CODIGO CONCEPTO", "NOMBRE CONCEPTO", "VALOR DE DESCUENTO"])
		# Empleado A: 2 conceptos (suman)
		ws.append(["", 43781807, "", " ", "ARBOLEDA YEPEZ", "MARGOT OLIVIA", "1", "Afiliación", 20000])
		ws.append(["", 43781807, "", " ", "ARBOLEDA YEPEZ", "MARGOT OLIVIA", "2", "Aportes / Depositos", 102000])
		# Empleado B: 1 concepto
		ws.append(["", 79726391, "", " ", "BERNAL SUA", "EDGAR ORLANDO", "2", "Aportes / Depositos", 191200])
		# Filas TOTAL del reporte se ignoran
		ws.append(["", 79726391, "", "", "", "", "", "TOTAL", 191200])
		return wb

	def test_matches_score(self):
		meta = {"filename": "7898_8153_02152026.xlsx", "sheets": ["Descuentos Nomina Detalle", "Descuentos Nomina Agrupado"]}
		self.assertEqual(fincomercio.matches(meta), 3)

	def test_detect_period_from_header_row5(self):
		self.assertEqual(fincomercio.detect_period(self._wb()), (2026, 2))

	def test_parse_aggregates_by_employee(self):
		results = list(fincomercio.parse(self._wb()))
		self.assertEqual(len(results), 2)
		emp_a = next(r for r in results if r.documento_identidad == "43781807")
		# 20000 + 102000 = 122000 (TOTAL row excluded).
		self.assertEqual(emp_a.valor, 122000)
		self.assertEqual(emp_a.tipo_novedad, "LIBRANZA_FINCOMERCIO")
		emp_b = next(r for r in results if r.documento_identidad == "79726391")
		self.assertEqual(emp_b.valor, 191200)


# ──────────────────────────────────────────────────────────────────────
# FONGIGA
# ──────────────────────────────────────────────────────────────────────

class FongigaTests(unittest.TestCase):
	def _wb(self) -> Workbook:
		wb = Workbook()
		ws = wb.active
		ws.title = "REAL11"
		ws.append(["Tercero", "Descripción", "Descripción Concepto", "Neto a pagar"])
		ws.append([1283443, "PONCE VIERA LISBELIS", "FONDO DE EMPLEADOS FONGIGA", 143600])
		ws.append([1283443, "PONCE VIERA LISBELIS", "PRESTAMO FONGIGA", 302736])
		ws.append([2000001349, "GARCIA ZAMBRANO RAMON", "FONDO DE EMPLEADOS FONGIGA", 142350])
		ws.append([5314314, "MONTERO RODRIGUEZ JOSE", "OTRA COSA NO MAPEADA", 99999])  # se ignora
		return wb

	def test_matches_score(self):
		meta = {"filename": "M HOME FEBRERO 2026.xlsx", "sheets": ["REAL11", "REAL", "FEB 011"]}
		self.assertEqual(fongiga.matches(meta), 3)

	def test_parse_resolves_concepts_and_skips_unknown(self):
		results = list(fongiga.parse(self._wb()))
		self.assertEqual(len(results), 3)
		# Empleado 1283443: dos novedades distintas (fondo + préstamo).
		fondo = next(
			r for r in results
			if r.documento_identidad == "1283443" and r.tipo_novedad == "FONDO_EMPLEADOS_FONGIGA"
		)
		prestamo = next(
			r for r in results
			if r.documento_identidad == "1283443" and r.tipo_novedad == "PRESTAMO_FONGIGA"
		)
		self.assertEqual(fondo.valor, 143600)
		self.assertEqual(prestamo.valor, 302736)
		# Concepto desconocido se filtró.
		self.assertFalse(any(r.documento_identidad == "5314314" for r in results))

	def test_str_id_keeps_leading_zeros_when_string(self):
		# DETALLADO HOME BURGUERS trae cédulas tipo "0001283986" como string;
		# se conservan tras quitar los ceros pero como string.
		from hubgh.hubgh.payroll.adapters.fongiga import _str_id

		self.assertEqual(_str_id("0001283986"), "1283986")
		self.assertEqual(_str_id(1283986), "1283986")
		self.assertEqual(_str_id(1283986.0), "1283986")
		self.assertEqual(_str_id(None), "")


if __name__ == "__main__":
	unittest.main()
