"""Tests del cómputo y export single-sheet del pipeline payroll.

Cadena cubierta sin Frappe: NovedadCanonica → enrich (con stubs) →
compute → export → bytes xlsx → re-leído con openpyxl para validar
celdas.
"""

from __future__ import annotations

import io
import unittest
from datetime import date

from openpyxl import load_workbook

from hubgh.hubgh.payroll.adapters import NovedadCanonica
from hubgh.hubgh.payroll import compute
from hubgh.hubgh.payroll.compute import ausentismos, induccion, literal, recargos
from hubgh.hubgh.payroll.enrichment import (
	ContractRecord,
	EmployeeRecord,
	EnrichmentContext,
	GlobalParams,
	enrich,
)
from hubgh.hubgh.payroll.export import build_single_sheet


def _make_ctx(salario_tc=1_400_000, salario_tp=1_000_000):
	emps = {
		"1001": EmployeeRecord(name="EMP-1", cedula="1001"),
		"1002": EmployeeRecord(name="EMP-2", cedula="1002"),
	}
	cons = {
		"EMP-1": ContractRecord(
			name="C-1", empleado="EMP-1",
			tipo_jornada="Tiempo Completo",
			salario=salario_tc,
			horas_trabajadas_mes=220,
		),
		"EMP-2": ContractRecord(
			name="C-2", empleado="EMP-2",
			tipo_jornada="Tiempo Parcial",
			salario=salario_tp,
			horas_trabajadas_mes=0,
		),
	}
	return EnrichmentContext(
		resolve_employee=lambda doc: emps.get(doc),
		resolve_contract=lambda emp, ps, pe: cons.get(emp),
		params=GlobalParams(),
	)


def _enrich(ctx, doc, tipo, **kwargs):
	# Defaults razonables
	novedad = NovedadCanonica(
		documento_identidad=doc,
		tipo_novedad=tipo,
		cantidad=kwargs.get("cantidad"),
		valor=kwargs.get("valor"),
		unidad=kwargs.get("unidad", "horas"),
		fecha_desde=kwargs.get("fecha_desde"),
		fecha_hasta=kwargs.get("fecha_hasta"),
		raw_payload=kwargs.get("raw_payload", {}),
	)
	return enrich(novedad, date(2026, 1, 16), date(2026, 2, 15), ctx)


class RecargosTests(unittest.TestCase):
	def test_hd_uses_multiplier_one(self):
		ctx = _make_ctx()
		nov = _enrich(ctx, "1001", "HD", cantidad=120.5, unidad="horas")
		compute.compute_novedad(nov, ctx.params)
		expected = round(120.5 * (1_400_000 / 220) * 1.0, 2)
		self.assertEqual(nov.calc_status, "computed")
		self.assertEqual(nov.computed_amount, expected)

	def test_hen_uses_multiplier_175(self):
		ctx = _make_ctx()
		nov = _enrich(ctx, "1001", "HEN", cantidad=10.0, unidad="horas")
		compute.compute_novedad(nov, ctx.params)
		valor_hora = 1_400_000 / 220
		expected = round(10.0 * valor_hora * 1.75, 2)
		self.assertEqual(nov.computed_amount, expected)

	def test_tp_uses_fixed_hour(self):
		ctx = _make_ctx()
		nov = _enrich(ctx, "1002", "HEFN", cantidad=4.0, unidad="horas")
		compute.compute_novedad(nov, ctx.params)
		expected = round(4.0 * 9530 * 2.55, 2)
		self.assertEqual(nov.computed_amount, expected)


class AusentismosTests(unittest.TestCase):
	def test_descanso_full_pay(self):
		ctx = _make_ctx()
		nov = _enrich(ctx, "1001", "DESCANSO", cantidad=2.0, unidad="dias")
		compute.compute_novedad(nov, ctx.params)
		valor_dia = 1_400_000 / 30
		self.assertAlmostEqual(nov.computed_amount, round(2 * valor_dia * 1.0, 2), places=1)

	def test_incapacidad_eg_at_66pct(self):
		ctx = _make_ctx()
		nov = _enrich(ctx, "1001", "INCAPACIDAD_ENFERMEDAD_GENERAL", cantidad=3.0, unidad="dias")
		compute.compute_novedad(nov, ctx.params)
		valor_dia = 1_400_000 / 30
		self.assertAlmostEqual(nov.computed_amount, round(3 * valor_dia * 0.66, 2), places=1)

	def test_ausencia_injustificada_descuenta(self):
		ctx = _make_ctx()
		nov = _enrich(ctx, "1001", "AUSENCIA_INJUSTIFICADA", cantidad=1.0, unidad="dias")
		compute.compute_novedad(nov, ctx.params)
		valor_dia = 1_400_000 / 30
		self.assertAlmostEqual(nov.computed_amount, round(-valor_dia, 2), places=1)
		self.assertLess(nov.computed_amount, 0)

	def test_licencia_no_remunerada_zero(self):
		ctx = _make_ctx()
		nov = _enrich(ctx, "1001", "LICENCIA_NO_REMUNERADA", cantidad=5.0, unidad="dias")
		compute.compute_novedad(nov, ctx.params)
		self.assertEqual(nov.computed_amount, 0.0)

	def test_licencia_luto_capped_at_5(self):
		ctx = _make_ctx()
		nov = _enrich(ctx, "1001", "LICENCIA_LUTO", cantidad=8.0, unidad="dias")
		compute.compute_novedad(nov, ctx.params)
		valor_dia = 1_400_000 / 30
		# Sólo 5 días pagables aunque la cantidad sea 8.
		self.assertAlmostEqual(nov.computed_amount, round(5 * valor_dia * 1.0, 2), places=1)
		self.assertIn("tope luto", nov.calc_notes)


class InduccionTests(unittest.TestCase):
	def test_induccion_tc_1_dia(self):
		ctx = _make_ctx()
		nov = _enrich(ctx, "1001", "INDUCCION", cantidad=1.0, unidad="dias")
		compute.compute_novedad(nov, ctx.params)
		valor_dia = 1_400_000 / 30
		self.assertAlmostEqual(nov.computed_amount, round(valor_dia, 2), places=1)

	def test_induccion_tp_733_horas(self):
		ctx = _make_ctx()
		nov = _enrich(ctx, "1002", "INDUCCION", cantidad=1.0, unidad="dias")
		compute.compute_novedad(nov, ctx.params)
		expected = round(1 * 7.33 * 9530, 2)
		self.assertAlmostEqual(nov.computed_amount, expected, places=1)


class LiteralTests(unittest.TestCase):
	def test_libranza_descuenta_negativo(self):
		ctx = _make_ctx()
		nov = _enrich(ctx, "1001", "LIBRANZA_FINCOMERCIO", valor=50_000, unidad="cop")
		compute.compute_novedad(nov, ctx.params)
		self.assertEqual(nov.computed_amount, -50_000)

	def test_bonificacion_suma_positivo(self):
		ctx = _make_ctx()
		nov = _enrich(ctx, "1001", "BONIFICACION_CP", valor=200_000, unidad="cop")
		compute.compute_novedad(nov, ctx.params)
		self.assertEqual(nov.computed_amount, 200_000)


class ExportSingleSheetTests(unittest.TestCase):
	def test_export_writes_one_row_per_employee(self):
		ctx = _make_ctx()
		# Empleado 1: HD 100h + descanso 2d + libranza 50k.
		# Empleado 2: HEN 4h.
		novs = [
			_enrich(ctx, "1001", "HD", cantidad=100, unidad="horas"),
			_enrich(ctx, "1001", "DESCANSO", cantidad=2, unidad="dias"),
			_enrich(ctx, "1001", "LIBRANZA_FINCOMERCIO", valor=50_000, unidad="cop"),
			_enrich(ctx, "1002", "HEN", cantidad=4, unidad="horas"),
		]
		for n in novs:
			compute.compute_novedad(n, ctx.params)

		blob = build_single_sheet(novs, ctx.params, period_label="2026-02")
		self.assertGreater(len(blob), 1000)

		wb = load_workbook(io.BytesIO(blob))
		ws = wb.active
		# 1 header + 2 empleados + 1 totales = 4 filas
		self.assertEqual(ws.max_row, 4)

		# Localiza la columna "$ HD" y la fila del empleado 1001
		headers = [c.value for c in ws[1]]
		col_hd_amt = headers.index("$ HD") + 1
		col_cedula = headers.index("Cédula") + 1
		row_emp1 = next(
			r for r in range(2, ws.max_row)
			if ws.cell(row=r, column=col_cedula).value == "1001"
		)
		expected_amount = round(100 * (1_400_000 / 220) * 1.0, 2)
		self.assertAlmostEqual(
			ws.cell(row=row_emp1, column=col_hd_amt).value, expected_amount, places=1
		)

		# El total descontado del empleado 1 incluye la libranza
		col_total_desc = headers.index("Total Descontado") + 1
		self.assertEqual(ws.cell(row=row_emp1, column=col_total_desc).value, -50_000)

		# Fila TOTAL final tiene fórmula SUM
		total_cell = ws.cell(row=ws.max_row, column=col_hd_amt).value
		self.assertTrue(str(total_cell).startswith("=SUM("))


if __name__ == "__main__":
	unittest.main()
