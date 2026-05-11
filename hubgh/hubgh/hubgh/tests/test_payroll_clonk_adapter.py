"""Tests del adapter CLONK.

No requieren Frappe levantado: el parser opera sobre un workbook de
openpyxl que se construye en memoria. Cubre:

- `matches()` con scoring multi-señal.
- `detect_period()` desde la columna Fecha de la hoja Detalle.
- `parse()` emitiendo NovedadCanonica de horas (Resumen) y de días
  (Novedades, con rango de fechas y mapeo de concepto a tipo canónico).
- Normalización de tipo de contrato del archivo (TC, TC - Admin,
  Aprendizaje → todos TC).
"""

from __future__ import annotations

import unittest
from datetime import date

from openpyxl import Workbook

from hubgh.hubgh.payroll.adapters import clonk
from hubgh.hubgh.jornada_utils import normalize_tipo_jornada, TIPO_JORNADA_FULL_TIME, TIPO_JORNADA_PART_TIME


def _build_clonk_workbook() -> Workbook:
	wb = Workbook()
	# Hoja Resumen
	ws_res = wb.active
	ws_res.title = "Resumen"
	ws_res.append(
		[
			"Empleado", "Documento", "Fecha de Ingreso", "Contrato", "Cargo", "Sucursal",
			"HD", "HN", "HFD", "HFN", "HED", "HEN", "HEFD", "HEFN", "NR", "NnR", "DnR",
		]
	)
	ws_res.append(
		[
			"Juan Perez", "1001", "01/01/2024", "Tiempo Completo", "AUX", "Home 1",
			120.5, 30.0, 0.0, 0.0, 5.0, 2.0, 0.0, 0.0, 0.0, 0.0, 0.0,
		]
	)
	ws_res.append(
		[
			"Ana Torres", "1002", "15/03/2025", "TC - Administración", "ADMIN", "Home 1",
			176.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 7.33, 0.0, 0.0,
		]
	)
	ws_res.append(
		[
			"Luz Aprendiz", "1003", "01/06/2025", "Aprendizaje", "APRENDIZ", "Home 2",
			88.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0,
		]
	)
	# Hoja Detalle de Tiempos (sólo para detect_period)
	ws_det = wb.create_sheet("Detalle de Tiempos")
	ws_det.append(
		[
			"Empleado", "Documento", "Fecha de Ingreso", "Contrato", "Cargo", "Sucursal",
			"Zona", "Fecha", "Entrada", "Salida",
			"HD", "HN", "HED", "HEN", "HFD", "HFN", "HEFD", "HEFN",
			"Total", "Diferencia", "NR", "NnR", "DnR", "Estado", "Comentarios",
		]
	)
	ws_det.append(
		[
			"Juan Perez", "1001", "01/01/2024", "Tiempo Completo", "AUX", "Home 1",
			"Operación", "16/01/2026", "06:00", "14:00",
			8.0, 0, 0, 0, 0, 0, 0, 0,
			8.0, "0.00", 0, 0, 0, "Revisado", "",
		]
	)
	ws_det.append(
		[
			"Juan Perez", "1001", "01/01/2024", "Tiempo Completo", "AUX", "Home 1",
			"Operación", "15/02/2026", "06:00", "14:00",
			8.0, 0, 0, 0, 0, 0, 0, 0,
			8.0, "0.00", 0, 0, 0, "Revisado", "",
		]
	)
	# Hoja Novedades
	ws_nov = wb.create_sheet("Novedades")
	ws_nov.append(
		[
			"Nombre", "Cédula", "Sede",
			"Incapacidad AT", "",
			"Vacaciones", "",
			"Incapacidad EG", "",
			"Descanso", "",
			"DIA CUMPLEAÑOS", "",
			"Maternidad", "",
			"AUSENTISMO", "",
			"INDUCCION", "",
			"DIA FAMILIA", "",
			"L. No Remunerada", "",
			"Suspensión", "",
		]
	)
	# Juan: descanso 26-27/01
	ws_nov.append(
		[
			"Juan Perez", "1001", "Home 1",
			"", "",
			"", "",
			"", "",
			date(2026, 1, 26), date(2026, 1, 27),
			"", "",
			"", "",
			"", "",
			"", "",
			"", "",
			"", "",
			"", "",
		]
	)
	# Ana: incapacidad EG 03-05/02
	ws_nov.append(
		[
			"Ana Torres", "1002", "Home 1",
			"", "",
			"", "",
			date(2026, 2, 3), date(2026, 2, 5),
			"", "",
			"", "",
			"", "",
			"", "",
			"", "",
			"", "",
			"", "",
			"", "",
		]
	)
	# Aprendiz: inducción 1 día
	ws_nov.append(
		[
			"Luz Aprendiz", "1003", "Home 2",
			"", "",
			"", "",
			"", "",
			"", "",
			"", "",
			"", "",
			"", "",
			date(2026, 1, 16), date(2026, 1, 16),
			"", "",
			"", "",
			"", "",
		]
	)
	return wb


class JornadaUtilsTests(unittest.TestCase):
	def test_tc_admin_normalizes_to_full_time(self):
		self.assertEqual(normalize_tipo_jornada("TC - Administración"), TIPO_JORNADA_FULL_TIME)
		self.assertEqual(normalize_tipo_jornada("TC-Admin"), TIPO_JORNADA_FULL_TIME)
		self.assertEqual(normalize_tipo_jornada("tc administración"), TIPO_JORNADA_FULL_TIME)

	def test_aprendizaje_normalizes_to_full_time(self):
		self.assertEqual(normalize_tipo_jornada("Aprendizaje"), TIPO_JORNADA_FULL_TIME)
		self.assertEqual(normalize_tipo_jornada("Aprendiz SENA"), TIPO_JORNADA_FULL_TIME)

	def test_tiempo_parcial_still_works(self):
		self.assertEqual(normalize_tipo_jornada("Tiempo Parcial"), TIPO_JORNADA_PART_TIME)
		self.assertEqual(normalize_tipo_jornada("TP"), TIPO_JORNADA_PART_TIME)

	def test_unknown_returns_empty(self):
		self.assertEqual(normalize_tipo_jornada("Vendedor freelance"), "")
		self.assertEqual(normalize_tipo_jornada(None), "")
		self.assertEqual(normalize_tipo_jornada(""), "")


class ClonkAdapterTests(unittest.TestCase):
	def setUp(self) -> None:
		self.wb = _build_clonk_workbook()
		self.file_meta = {
			"filename": "Toda la empresa - Jan 16 - Feb 15, 2026.xlsx",
			"sheets": list(self.wb.sheetnames),
		}

	def test_matches_full_score_with_filename_and_sheets(self):
		score = clonk.matches(self.file_meta)
		self.assertEqual(score, 3)

	def test_matches_partial_score_without_filename(self):
		meta = {"filename": "report.xlsx", "sheets": list(self.wb.sheetnames)}
		self.assertGreaterEqual(clonk.matches(meta), 2)

	def test_matches_zero_when_unrelated(self):
		meta = {"filename": "fincomercio.xlsx", "sheets": ["Hoja1"]}
		self.assertEqual(clonk.matches(meta), 0)

	def test_detect_period_from_detalle(self):
		self.assertEqual(clonk.detect_period(self.wb), (2026, 2))

	def test_parse_emits_hour_novedades_from_resumen(self):
		results = list(clonk.parse(self.wb))
		hour_results = [n for n in results if n.unidad == "horas"]
		# Juan: HD, HN, HED, HEN (4 columnas > 0). Ana: HD (1). Luz: HD (1). Total 6.
		self.assertEqual(len(hour_results), 6)
		hd_juan = next(
			n for n in hour_results
			if n.documento_identidad == "1001" and n.tipo_novedad == "HD"
		)
		self.assertAlmostEqual(hd_juan.cantidad, 120.5, places=4)
		self.assertEqual(hd_juan.raw_payload["contrato_text"], "Tiempo Completo")

	def test_parse_emits_concept_novedades_from_novedades_sheet(self):
		results = list(clonk.parse(self.wb))
		concept_results = [n for n in results if n.unidad == "dias"]
		self.assertEqual(len(concept_results), 3)
		# Juan: descanso 26-27/01 → 2 días, tipo DESCANSO
		descanso = next(n for n in concept_results if n.tipo_novedad == "DESCANSO")
		self.assertEqual(descanso.documento_identidad, "1001")
		self.assertEqual(descanso.fecha_desde, "2026-01-26")
		self.assertEqual(descanso.fecha_hasta, "2026-01-27")
		self.assertEqual(descanso.cantidad, 2.0)
		# Ana: incapacidad EG 03-05/02 → 3 días, tipo canónico
		incap = next(
			n for n in concept_results if n.tipo_novedad == "INCAPACIDAD_ENFERMEDAD_GENERAL"
		)
		self.assertEqual(incap.documento_identidad, "1002")
		self.assertEqual(incap.cantidad, 3.0)

	def test_ausentismo_maps_to_ausencia_injustificada(self):
		# Agrego una fila de AUSENTISMO para validar el alias del mapping
		ws_nov = self.wb["Novedades"]
		ws_nov.append(
			[
				"Otro", "1004", "Home 3",
				"", "",
				"", "",
				"", "",
				"", "",
				"", "",
				"", "",
				date(2026, 2, 8), date(2026, 2, 8),
				"", "",
				"", "",
				"", "",
				"", "",
			]
		)
		results = list(clonk.parse(self.wb))
		ausencia = next(
			n for n in results
			if n.documento_identidad == "1004" and n.tipo_novedad == "AUSENCIA_INJUSTIFICADA"
		)
		self.assertEqual(ausencia.cantidad, 1.0)


class DetectorTests(unittest.TestCase):
	def test_clonk_wins_when_signature_matches(self):
		from hubgh.hubgh.payroll.adapters import _detect

		meta = {
			"filename": "Toda la empresa - Jan 16 - Feb 15, 2026.xlsx",
			"sheets": ["Resumen", "Detalle de Tiempos", "Novedades"],
		}
		self.assertEqual(_detect.detect_source(meta), "clonk")

	def test_unknown_when_no_match(self):
		from hubgh.hubgh.payroll.adapters import _detect

		meta = {"filename": "random.xlsx", "sheets": ["Hoja1"]}
		self.assertEqual(_detect.detect_source(meta), "unknown")


if __name__ == "__main__":
	unittest.main()
