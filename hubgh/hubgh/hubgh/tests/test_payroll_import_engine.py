"""
Integration tests for Payroll Import Engine - Sprint 2 functionality.

Tests CLONK parsing, deduplication logic, and batch processing flow.
"""

import json
import os
import tempfile
from unittest.mock import patch, MagicMock

import frappe
from frappe.tests.utils import FrappeTestCase
from openpyxl import Workbook

from hubgh.hubgh.payroll_import_engine import (
	parse_clonk_file,
	detect_source_type,
	process_import_batch,
	process_import_run,
	check_duplicate_line,
	generate_dedup_hash,
	get_source_adapter,
	register_source_adapter,
	SOURCE_ADAPTERS,
)


class TestPayrollImportEngine(FrappeTestCase):
	"""Test suite for payroll import engine functionality."""
	
	def setUp(self):
		"""Set up test data and mocks."""
		self.cleanup_test_data()
		self.create_test_catalog_data()
		
	def tearDown(self):
		"""Clean up after tests."""
		self.cleanup_test_data()
		
	def cleanup_test_data(self):
		"""Remove any test data created during tests."""
		frappe.db.delete("Payroll Import Line", {"batch": ["like", "TEST-%"]})
		frappe.db.delete("Payroll Import Batch", {"name": ["like", "TEST-%"]})
		frappe.db.commit()
		
	def create_test_catalog_data(self):
		"""Create necessary catalog entries for testing."""
		# Create test source catalog entry
		if not frappe.db.exists("Payroll Source Catalog", "CLONK"):
			frappe.get_doc({
				"doctype": "Payroll Source Catalog",
				"nombre_fuente": "CLONK",
				"tipo_fuente": "clonk",
				"mapeo_columnas": json.dumps({
					"documento": ["cedula", "documento"],
					"nombre": ["nombre", "empleado"]
				}),
				"periodicidad": "Mensual",
				"status": "Activa"
			}).insert(ignore_permissions=True)
			
		# Create test period config
		if not frappe.db.exists("Payroll Period Config", "TEST-2026-03"):
			frappe.get_doc({
				"doctype": "Payroll Period Config",
				"name": "TEST-2026-03",
				"period_label": "Marzo 2026 TEST",
				"year": 2026,
				"month": 3,
				"status": "Activo"
			}).insert(ignore_permissions=True)
			
		frappe.db.commit()
		
	def create_mock_clonk_file(self):
		"""Create a mock CLONK Excel file for testing."""
		wb = Workbook()
		
		# Remove default sheet
		wb.remove(wb.active)
		
		# Create "Resumen horas" sheet
		resumen_sheet = wb.create_sheet("Resumen horas")
		resumen_sheet.append(["Cedula", "Nombre", "HD", "HN", "HED", "HEN"])
		resumen_sheet.append(["12345678", "Juan Perez", 40, 8, 5, 2])
		resumen_sheet.append(["87654321", "Maria Lopez", 35, 0, 0, 0])
		resumen_sheet.append(["11111111", "Carlos Rodriguez", 42, 10, 8, 4])
		
		# Create "Ausentismos" sheet
		ausentismo_sheet = wb.create_sheet("Ausentismos por tipo")
		ausentismo_sheet.append(["Cedula", "Nombre", "Tipo Ausencia", "Dias"])
		ausentismo_sheet.append(["12345678", "Juan Perez", "Vacaciones", 5])
		ausentismo_sheet.append(["87654321", "Maria Lopez", "Incapacidad EG", 3])
		
		# Save to temporary file
		temp_file = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
		wb.save(temp_file.name)
		wb.close()
		
		return temp_file.name
		
	def test_detect_source_type(self):
		"""Test automatic source type detection from filename."""
		test_cases = [
			("CLONK_Marzo_2026.xlsx", "CLONK"),
			("Toda la empresa marzo.xlsx", "CLONK"),
			("payflow_resumen.xlsx", "Payflow Resumen"),
			("fincomercio_datos.xlsx", "Fincomercio"),
			("fondo_m_home.xlsx", "Fondo FONGIGA"),
			("libranzas_marzo.xlsx", "Libranzas Bancolombia"),
			("unknown_file.xlsx", None)
		]
		
		for filename, expected_type in test_cases:
			with self.subTest(filename=filename):
				result = detect_source_type(filename)
				self.assertEqual(result, expected_type)
				
	def test_parse_clonk_file(self):
		"""Test CLONK file parsing with mock data."""
		temp_file = self.create_mock_clonk_file()
		
		try:
			lines, errors = parse_clonk_file(temp_file, "TEST-BATCH")
			
			# Verify no errors
			self.assertEqual(len(errors), 0, f"Parsing errors: {errors}")
			
			# Verify line count - should have lines from both sheets
			self.assertGreater(len(lines), 0)
			
			# Verify hour types from resumen sheet
			hour_lines = [l for l in lines if l["novedad_type"] in ["HD", "HN", "HED", "HEN"]]
			self.assertGreater(len(hour_lines), 0)
			
			# Verify absence types from ausentismo sheet  
			absence_lines = [l for l in lines if l["novedad_type"] in ["VACACIONES", "INC-EG"]]
			self.assertEqual(len(absence_lines), 2)
			
			# Verify line structure
			for line in lines:
				self.assertIn("batch", line)
				self.assertIn("employee_id", line)
				self.assertIn("novedad_type", line)
				self.assertIn("quantity", line)
				self.assertIn("source_sheet", line)
				self.assertIn("source_row_data", line)
				
		finally:
			os.unlink(temp_file)

	def test_source_adapter_registry_supports_alias_lookup(self):
		"""Registry should resolve adapters by canonical code and alias."""
		class DummyAdapter:
			code = "dummy-source"
			parser_version = "dummy.v1"

			def parse(self, file_path, batch_doc):
				return [], []

		original_adapters = dict(SOURCE_ADAPTERS)
		try:
			SOURCE_ADAPTERS.clear()
			adapter = DummyAdapter()
			register_source_adapter(adapter, "DUMMY", "Dummy Alias")

			self.assertIs(get_source_adapter("dummy-source"), adapter)
			self.assertIs(get_source_adapter("ignored", "dummy"), adapter)
			self.assertIs(get_source_adapter("Dummy Alias"), adapter)
		finally:
			SOURCE_ADAPTERS.clear()
			SOURCE_ADAPTERS.update(original_adapters)

	def test_parse_clonk_file_keeps_run_provenance_in_canonical_payload(self):
		"""CLONK parsing should keep grouped run provenance in each canonical row."""
		temp_file = self.create_mock_clonk_file()
		try:
			lines, errors = parse_clonk_file(temp_file, "TEST-RUN-BATCH")
			self.assertEqual(errors, [])
			self.assertGreater(len(lines), 0)

			first_line = lines[0]
			self.assertEqual(first_line["run_id"], "TEST-RUN-BATCH")
			self.assertIn("source_file", first_line)
			self.assertIn("source_type_code", first_line)
			raw_payload = json.loads(first_line["raw_payload_json"])
			self.assertEqual(raw_payload["provenance"]["run_id"], "TEST-RUN-BATCH")
			self.assertEqual(raw_payload["provenance"]["batch"], "TEST-RUN-BATCH")
		finally:
			os.unlink(temp_file)

	@patch("hubgh.hubgh.payroll_import_engine.process_import_batch")
	@patch("hubgh.hubgh.payroll_import_engine.frappe.db.get_value")
	@patch("hubgh.hubgh.payroll_import_engine.frappe.get_all")
	def test_process_import_run_reuses_confirmed_batch_totals(self, mock_get_all, mock_get_value, mock_process_batch):
		"""Run processing should consolidate fresh and already-confirmed batches coherently."""
		mock_get_all.return_value = [
			{
				"name": "TEST-RUN-BATCH-001",
				"status": "Confirmado",
				"nomina_period": "2026-03",
				"run_label": "Marzo 2026 · RUN-001",
				"run_source_count": 2,
			},
			{
				"name": "TEST-RUN-BATCH-002",
				"status": "Pendiente",
				"nomina_period": "2026-03",
				"run_label": "Marzo 2026 · RUN-001",
				"run_source_count": 2,
			},
		]
		mock_get_value.side_effect = [12, 10, 2]
		mock_process_batch.return_value = {
			"status": "Completado con duplicados",
			"total_rows": 8,
			"valid_rows": 7,
			"error_rows": 0,
			"duplicate_rows": 1,
			"errors": [],
		}

		result = process_import_run("RUN-001")

		self.assertEqual(result["run_id"], "RUN-001")
		self.assertEqual(result["status"], "Completado con duplicados")
		self.assertEqual(result["total_rows"], 20)
		self.assertEqual(result["valid_rows"], 17)
		self.assertEqual(result["error_rows"], 2)
		self.assertEqual(result["duplicate_rows"], 1)
		self.assertEqual(len(result["batch_results"]), 2)
		mock_process_batch.assert_called_once_with("TEST-RUN-BATCH-002")
			
	def test_deduplication_logic(self):
		"""Test deduplication hash generation and checking."""
		# Test hash generation
		hash1 = generate_dedup_hash("2026-03", "12345678", "HD", "2026-03-15")
		hash2 = generate_dedup_hash("2026-03", "12345678", "HD", "2026-03-15") 
		hash3 = generate_dedup_hash("2026-03", "12345678", "HN", "2026-03-15")
		
		# Same parameters should generate same hash
		self.assertEqual(hash1, hash2)
		
		# Different parameters should generate different hash
		self.assertNotEqual(hash1, hash3)
		
		# Test duplicate checking with empty database
		duplicate = check_duplicate_line("2026-03", "12345678", "HD", "2026-03-15")
		self.assertIsNone(duplicate)
		
	def test_batch_processing_flow(self):
		"""Test end-to-end batch processing flow."""
		temp_file = self.create_mock_clonk_file()
		
		try:
			# Create a test batch
			batch = frappe.get_doc({
				"doctype": "Payroll Import Batch",
				"source_type": "CLONK",
				"period": "TEST-2026-03",
				"nomina_period": "Marzo 2026 TEST",
				"status": "Pendiente"
			}).insert(ignore_permissions=True)
			
			# Mock the file attachment
			with patch("frappe.get_doc") as mock_get_doc:
				# Mock the batch document
				mock_batch = MagicMock()
				mock_batch.source_file = "/test/file/path"
				mock_batch.source_type = "CLONK"
				mock_batch.period = "TEST-2026-03"
				mock_batch.nomina_period = "Marzo 2026 TEST"
				mock_batch.status = "Pendiente"
				mock_batch.name = batch.name
				mock_batch.save = MagicMock()
				mock_batch.reload = MagicMock()
				
				# Mock the file document
				mock_file_doc = MagicMock()
				mock_file_doc.get_full_path.return_value = temp_file
				
				# Mock the source catalog
				mock_source_doc = MagicMock()
				mock_source_doc.tipo_fuente = "clonk"
				
				def mock_get_doc_side_effect(*args, **kwargs):
					if args[0] == "Payroll Import Batch":
						return mock_batch
					elif args[0] == "File":
						return mock_file_doc
					elif args[0] == "Payroll Source Catalog":
						return mock_source_doc
					else:
						return MagicMock()
						
				mock_get_doc.side_effect = mock_get_doc_side_effect
				
				# Process the batch
				result = process_import_batch(batch.name)
				
				# Verify result structure
				self.assertIn("status", result)
				self.assertIn("total_rows", result)
				self.assertIn("valid_rows", result)
				self.assertIn("error_rows", result)
				self.assertIn("duplicate_rows", result)
				
				# Should have processed successfully
				self.assertIn(result["status"], ["Completado", "Completado con errores"])
				self.assertGreater(result["total_rows"], 0)
				
		finally:
			os.unlink(temp_file)
			
	def test_duplicate_detection_in_processing(self):
		"""Test that duplicate lines are properly detected during processing."""
		# Create first line manually
		first_line = frappe.get_doc({
			"doctype": "Payroll Import Line",
			"batch": "TEST-DEDUP-BATCH",
			"employee_id": "12345678",
			"novedad_type": "HD",
			"novedad_date": "2026-03-15",
			"quantity": 8,
			"dedup_hash": generate_dedup_hash("2026-03", "12345678", "HD", "2026-03-15"),
			"status": "Pendiente"
		}).insert(ignore_permissions=True)
		
		try:
			# Check for duplicate (should find the first line)
			duplicate = check_duplicate_line("2026-03", "12345678", "HD", "2026-03-15")
			self.assertEqual(duplicate, first_line.name)
			
			# Check for non-duplicate (different novelty type)
			no_duplicate = check_duplicate_line("2026-03", "12345678", "HN", "2026-03-15")
			self.assertIsNone(no_duplicate)
			
		finally:
			frappe.delete_doc("Payroll Import Line", first_line.name, force=True)
			
	def test_clonk_resumen_parsing_edge_cases(self):
		"""Test CLONK resumen sheet parsing with edge cases."""
		wb = Workbook()
		wb.remove(wb.active)
		
		# Create sheet with edge cases
		sheet = wb.create_sheet("Resumen horas")
		sheet.append(["Cedula", "Nombre", "HD", "HN", "Observaciones"])
		sheet.append(["12345678", "Juan Perez", 40, 8, "Normal"])  # Valid
		sheet.append([None, "Sin Cedula", 35, 0, "Missing ID"])  # Missing document
		sheet.append(["87654321", "Zero Hours", 0, 0, "No hours"])  # Zero hours
		sheet.append(["11111111", "Negative", -5, 2, "Negative"])  # Negative hours
		sheet.append(["22222222", "Text Hours", "abc", "def", "Invalid"])  # Text in hours
		
		temp_file = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
		wb.save(temp_file.name)
		wb.close()
		
		try:
			lines, errors = parse_clonk_file(temp_file, "TEST-EDGE-CASES")
			
			# Should only get lines for valid data
			valid_lines = [l for l in lines if l["employee_id"] == "12345678"]
			self.assertGreater(len(valid_lines), 0)
			
			# Zero and negative hours should be filtered out
			zero_negative_lines = [l for l in lines if l["employee_id"] in ["87654321", "11111111"]]
			self.assertEqual(len(zero_negative_lines), 0)
			
			# Text hours should be filtered out 
			text_lines = [l for l in lines if l["employee_id"] == "22222222"]
			self.assertEqual(len(text_lines), 0)
			
		finally:
			os.unlink(temp_file)
			
	def test_clonk_ausentismo_mapping(self):
		"""Test CLONK ausentismo sheet absence type mapping."""
		wb = Workbook()
		wb.remove(wb.active)
		
		# Create sheet with various absence types
		sheet = wb.create_sheet("Ausentismos")
		sheet.append(["Cedula", "Tipo Ausencia", "Dias"])
		sheet.append(["12345678", "Vacaciones", 5])
		sheet.append(["87654321", "Incapacidad EG", 3])
		sheet.append(["11111111", "Incapacidad AT", 2])
		sheet.append(["22222222", "Licencia Remunerada", 1])
		sheet.append(["33333333", "Calamidad Doméstica", 2])
		sheet.append(["44444444", "Ausentismo", 1])
		sheet.append(["55555555", "Tipo Desconocido", 1])  # Unknown type
		
		temp_file = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
		wb.save(temp_file.name)
		wb.close()
		
		try:
			lines, errors = parse_clonk_file(temp_file, "TEST-ABSENCE-MAPPING")
			
			# Verify absence type mappings
			type_mapping = {l["employee_id"]: l["novedad_type"] for l in lines}
			
			self.assertEqual(type_mapping.get("12345678"), "VACACIONES")
			self.assertEqual(type_mapping.get("87654321"), "INC-EG")
			self.assertEqual(type_mapping.get("11111111"), "INC-AT")
			self.assertEqual(type_mapping.get("22222222"), "LIC-REM")
			self.assertEqual(type_mapping.get("33333333"), "CALAMIDAD")
			self.assertEqual(type_mapping.get("44444444"), "AUSENTISMO")
			# Unknown type should default to AUSENTISMO
			self.assertEqual(type_mapping.get("55555555"), "AUSENTISMO")
			
		finally:
			os.unlink(temp_file)


if __name__ == "__main__":
	frappe.init("hubgh.test")
	frappe.connect()
	
	# Run specific test
	import unittest
	unittest.main()
