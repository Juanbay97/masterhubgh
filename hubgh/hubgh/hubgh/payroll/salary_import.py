"""Import de salarios desde un Excel del operador.

Formato esperado del archivo (header en fila 1):
  Empleado | Nombre del empleado | Salario | Descripcion del cargo

Estrategia (best-effort, configurable):
1. Para cada fila, canonicaliza el cargo del archivo y busca el `Cargo`
   del catálogo Hubgh que mejor matchee.
2. Crea o actualiza una `Ficha Empleado` con cédula y nombre parseado.
3. Crea o actualiza un `Contrato` Activo asociado al empleado con el
   salario y el cargo canónico.
4. (Opcional) actualiza `Cargo.salario_base_tc` con el salario más
   común por cargo, para que empleados sin Ficha en DB sigan resolviendo
   por catálogo.

Uso:
  bench --site hubgh.local execute hubgh.hubgh.payroll.salary_import.import_from_xlsx \\
    --kwargs '{"path":"/tmp/nomina_grid.xlsx", "set_cargo_default": true}'
"""

from __future__ import annotations

import os
from collections import Counter, defaultdict
from typing import Iterator

import frappe

from hubgh.hubgh.payroll.enrichment import canonicalize_cargo


def _split_name(full: str) -> tuple[str, str]:
	"""Parsea el nombre del archivo. Heurística simple: las primeras dos
	palabras suelen ser apellidos (en hojas tipo NomProCon...) y el resto
	nombres. Si hay menos de 3 palabras, todo va a `nombres`.
	"""
	parts = (full or "").strip().split()
	if len(parts) >= 4:
		apellidos = " ".join(parts[:2])
		nombres = " ".join(parts[2:])
	elif len(parts) == 3:
		apellidos = parts[0]
		nombres = " ".join(parts[1:])
	else:
		apellidos = ""
		nombres = " ".join(parts) or ""
	return nombres, apellidos


def _iter_rows(path: str) -> Iterator[dict]:
	import openpyxl

	wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
	ws = wb[wb.sheetnames[0]]
	rows = ws.iter_rows(min_row=1, values_only=True)
	header = next(rows, None)
	if not header:
		return
	header_idx = {str(h).strip().lower(): i for i, h in enumerate(header) if h}

	def col(*aliases):
		for a in aliases:
			if a in header_idx:
				return header_idx[a]
		return None

	idx_doc = col("empleado", "cedula", "documento")
	idx_nombre = col("nombre del empleado", "nombre", "nombres")
	idx_salario = col("salario")
	idx_cargo = col("descripcion del cargo", "cargo")
	if idx_doc is None or idx_salario is None:
		raise ValueError("El archivo debe traer columnas 'Empleado' y 'Salario'.")
	for row in rows:
		doc_raw = row[idx_doc] if idx_doc < len(row) else None
		if doc_raw is None:
			continue
		documento = str(int(doc_raw)) if isinstance(doc_raw, (int, float)) and float(doc_raw).is_integer() else str(doc_raw).strip()
		if not documento:
			continue
		yield {
			"documento": documento,
			"nombre": str(row[idx_nombre]).strip() if idx_nombre is not None and row[idx_nombre] else "",
			"salario": float(row[idx_salario] or 0),
			"cargo_archivo": str(row[idx_cargo]).strip() if idx_cargo is not None and row[idx_cargo] else "",
		}


def _resolve_canonical_cargo(cargo_archivo: str, canonical_index: dict[str, str]) -> str | None:
	"""Mapea el string del archivo al `Cargo.name` canónico vía
	canonicalize_cargo. Devuelve None si no hay match.
	"""
	if not cargo_archivo:
		return None
	canon = canonicalize_cargo(cargo_archivo)
	return canonical_index.get(canon)


def _build_canonical_index() -> dict[str, str]:
	"""Construye {canonical_nombre: Cargo.name} desde todos los Cargos
	activos. Si dos cargos canonicalizan igual, gana el primero.
	"""
	rows = frappe.get_all(
		"Cargo",
		filters={"activo": 1},
		fields=["name", "nombre"],
		limit_page_length=0,
	)
	idx: dict[str, str] = {}
	for r in rows:
		canon = canonicalize_cargo(r.get("nombre")) or canonicalize_cargo(r.get("name"))
		if canon and canon not in idx:
			idx[canon] = r["name"]
	return idx


def import_from_xlsx(path: str, set_cargo_default: bool = True, dry_run: bool = False) -> dict:
	"""Importa el mapa salarial.

	`set_cargo_default`: si True, después de procesar todas las filas,
	setea `Cargo.salario_base_tc` con el salario más común para cada
	cargo canónico (modo). Útil cuando el operador no quiere crear
	Ficha+Contrato pero sí tener fallback por catálogo.

	`dry_run`: si True, devuelve el reporte sin escribir.
	"""
	if not os.path.exists(path):
		return {"error": f"No existe el archivo: {path}"}

	canonical_index = _build_canonical_index()
	cargos_existentes = set(canonical_index.values())

	stats = {
		"total_filas": 0,
		"empleados_creados": 0,
		"empleados_actualizados": 0,
		"contratos_creados": 0,
		"contratos_actualizados": 0,
		"sin_cargo_match": [],
		"sin_cedula": 0,
		"errores": [],
	}
	# Para set_cargo_default
	salaries_per_cargo: dict[str, list[float]] = defaultdict(list)

	for row in _iter_rows(path):
		stats["total_filas"] += 1
		documento = row["documento"]
		if not documento:
			stats["sin_cedula"] += 1
			continue

		cargo_canonico = _resolve_canonical_cargo(row["cargo_archivo"], canonical_index)
		if not cargo_canonico:
			stats["sin_cargo_match"].append({
				"documento": documento,
				"cargo_archivo": row["cargo_archivo"],
				"canonical": canonicalize_cargo(row["cargo_archivo"]),
			})

		if cargo_canonico and row["salario"] > 0:
			salaries_per_cargo[cargo_canonico].append(row["salario"])

		if dry_run:
			continue

		try:
			# 1. Ficha Empleado.
			emp_name = frappe.db.get_value("Ficha Empleado", {"cedula": documento}, "name")
			nombres, apellidos = _split_name(row["nombre"])
			if emp_name:
				updates = {}
				if nombres:
					updates["nombres"] = nombres
				if apellidos:
					updates["apellidos"] = apellidos
				if updates:
					frappe.db.set_value("Ficha Empleado", emp_name, updates, update_modified=False)
				stats["empleados_actualizados"] += 1
			else:
				doc = frappe.get_doc({
					"doctype": "Ficha Empleado",
					"cedula": documento,
					"nombres": nombres or "Sin nombre",
					"apellidos": apellidos or "—",
					"estado": "Activo",
				}).insert(ignore_permissions=True, ignore_mandatory=True)
				emp_name = doc.name
				stats["empleados_creados"] += 1

			# 2. Contrato Activo.
			contract = frappe.db.get_value(
				"Contrato",
				{"empleado": emp_name, "estado_contrato": "Activo"},
				["name", "salario", "cargo"],
				as_dict=True,
			)
			if contract:
				updates = {}
				if row["salario"] and float(contract.salario or 0) != row["salario"]:
					updates["salario"] = row["salario"]
				if cargo_canonico and contract.cargo != cargo_canonico:
					updates["cargo"] = cargo_canonico
				if updates:
					frappe.db.set_value("Contrato", contract.name, updates, update_modified=False)
					stats["contratos_actualizados"] += 1
			else:
				contract_doc = frappe.get_doc({
					"doctype": "Contrato",
					"empleado": emp_name,
					"numero_documento": documento,
					"nombres": nombres or "Sin nombre",
					"apellidos": apellidos or "—",
					"tipo_contrato": "Indefinido",
					"tipo_jornada": "Tiempo Completo",  # default; CLONK puede sobrescribir.
					"fecha_ingreso": "2024-01-01",
					"salario": row["salario"],
					"horas_trabajadas_mes": 220,
					"cargo": cargo_canonico,
					"estado_contrato": "Pendiente",
				}).insert(ignore_permissions=True, ignore_mandatory=True)
				frappe.db.set_value(
					"Contrato",
					contract_doc.name,
					"estado_contrato",
					"Activo",
					update_modified=False,
				)
				stats["contratos_creados"] += 1
		except Exception as exc:
			stats["errores"].append({"documento": documento, "error": str(exc)})

	# 3. Setear Cargo.salario_base_tc con el modo del salario.
	cargo_defaults_set = []
	if set_cargo_default and not dry_run:
		for cargo_name, salaries in salaries_per_cargo.items():
			if not salaries:
				continue
			counter = Counter(salaries)
			modal = counter.most_common(1)[0][0]
			frappe.db.set_value(
				"Cargo", cargo_name, "salario_base_tc", modal, update_modified=False
			)
			cargo_defaults_set.append({"cargo": cargo_name, "salario_modal": modal, "n": len(salaries)})
	stats["cargo_defaults_set"] = cargo_defaults_set

	if not dry_run:
		frappe.db.commit()
	# Limitar la lista de sin_cargo_match en el reporte para no saturar.
	stats["sin_cargo_match_total"] = len(stats["sin_cargo_match"])
	stats["sin_cargo_match"] = stats["sin_cargo_match"][:20]
	return stats
