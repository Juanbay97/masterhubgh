"""Excel single-sheet aplanado por empleado.

Layout de columnas (en orden):

  Identificación
    Cédula | Nombres | Apellidos | Jornada | Salario

  Horas (cantidades)
    HD h | HN h | HFD h | HFN h | HED h | HEN h | HEFD h | HEFN h

  Importes de horas
    $ HD | $ HN | $ HFD | $ HFN | $ HED | $ HEN | $ HEFD | $ HEFN

  Beneficios remunerados (días + importe agregado)
    Vacaciones d | Descanso d | Día Familia d | Día Cumple d |
    L. Maternidad d | L. Remunerada d | L. Luto d | Permiso Calamidad d

  Ausentismos no pagados (días)
    L. No Remunerada d | Suspensión d | Ausencia Injust d

  Incapacidades (días)
    Incap. EG d | Incap. AT d | Incap. Empresa d | Incap. >180 d

  Otros pagos
    Inducción $ | Auxilio Movilización $ | Auxilio Rodamiento $ |
    Bonificación CP $ | Auxilio Transporte $

  Descuentos
    Adelanto Payflow $ | Sanitas $ | Gafas $ | FONGIGA Fondo $ |
    Libranza Comfenalco $ | Libranza Fincomercio $ |
    Libranza Davivienda $ | Libranza Compensar $ |
    Préstamo Empresa $ | Préstamo FONGIGA $ | Pérdida Bonif. $

  Totales
    Total Devengado | Total Descontado | Neto a Pagar
"""

from __future__ import annotations

import io
from collections import defaultdict
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from hubgh.hubgh.payroll.compute import auxilio_transporte


_IDENT_COLUMNS = [
	("Cédula", "id_cedula", "first"),
	("Nombres", "id_nombres", "first"),
	("Apellidos", "id_apellidos", "first"),
	("Jornada", "id_jornada", "first"),
	("Cargo", "id_cargo", "first"),
	("Sucursal", "id_sucursal", "first"),
]


# Hoja "Hechos": cantidades crudas del archivo, sin importes calculados.
# Lo que el equipo de nómina necesita para correr su propio sistema oficial.
COLUMN_SPEC_HECHOS: list[tuple[str, str, str]] = _IDENT_COLUMNS + [
	# Horas (cantidad cruda del CLONK)
	("HD h", "h_qty:HD", "sum"),
	("HN h", "h_qty:HN", "sum"),
	("HFD h", "h_qty:HFD", "sum"),
	("HFN h", "h_qty:HFN", "sum"),
	("HED h", "h_qty:HED", "sum"),
	("HEN h", "h_qty:HEN", "sum"),
	("HEFD h", "h_qty:HEFD", "sum"),
	("HEFN h", "h_qty:HEFN", "sum"),
	# Días remunerados (cantidad)
	("Vacaciones d", "d_qty:VACACIONES", "sum"),
	("Descanso d", "d_qty:DESCANSO", "sum"),
	("Día Familia d", "d_qty:DIA_FAMILIA", "sum"),
	("Día Cumple d", "d_qty:DIA_CUMPLEANOS", "sum"),
	("L. Maternidad d", "d_qty:LICENCIA_MATERNIDAD", "sum"),
	("L. Remunerada d", "d_qty:LICENCIA_REMUNERADA", "sum"),
	("L. Luto d", "d_qty:LICENCIA_LUTO", "sum"),
	("Permiso Calamidad d", "d_qty:PERMISO_CALAMIDAD", "sum"),
	# Días no remunerados (cantidad)
	("L. No Remunerada d", "d_qty:LICENCIA_NO_REMUNERADA", "sum"),
	("Suspensión d", "d_qty:SUSPENSION_CONTRATO", "sum"),
	("Ausencia Injust d", "d_qty:AUSENCIA_INJUSTIFICADA", "sum"),
	# Incapacidades (cantidad)
	("Incap. EG d", "d_qty:INCAPACIDAD_ENFERMEDAD_GENERAL", "sum"),
	("Incap. AT d", "d_qty:INCAPACIDAD_ACCIDENTE_TRABAJO", "sum"),
	("Incap. Empresa d", "d_qty:INCAPACIDAD_PAGADA_EMPRESA", "sum"),
	("Incap. >180 d", "d_qty:INCAPACIDAD_MAYOR_180_DIAS", "sum"),
	# Inducción (días)
	("Inducción d", "d_qty:INDUCCION", "sum"),
	# Días trabajados (CLONK Detalle, para auditoría del prorrateo aux. transporte)
	("Días trabajados", "id_dias_trabajados", "first"),
	# Valores literales de descuentos (del archivo, NO calculado)
	("Adelanto Payflow $", "amt:ADELANTO_NOMINA_PAYFLOW", "sum"),
	("Sanitas $", "amt:DESCUENTO_SANITAS_PREMIUM", "sum"),
	("Gafas $", "amt:DESCUENTO_GAFAS", "sum"),
	("FONGIGA Fondo $", "amt:FONDO_EMPLEADOS_FONGIGA", "sum"),
	("Libranza Comfenalco $", "amt:LIBRANZA_COMFENALCO", "sum"),
	("Libranza Fincomercio $", "amt:LIBRANZA_FINCOMERCIO", "sum"),
	("Libranza Davivienda $", "amt:LIBRANZA_DAVIVIENDA", "sum"),
	("Libranza Compensar $", "amt:LIBRANZA_COMPENSAR", "sum"),
	("Préstamo Empresa $", "amt:PRESTAMO_EMPRESA", "sum"),
	("Préstamo FONGIGA $", "amt:PRESTAMO_FONGIGA", "sum"),
	("Pérdida Bonif. $", "amt:PERDIDA_BONIFICACION", "sum"),
	# Auxilios y bonificaciones pactados (también valor literal del archivo)
	("Auxilio Movilización $", "amt:AUXILIO_MOVILIZACION_DOM_FEST", "sum"),
	("Auxilio Rodamiento $", "amt:AUXILIO_RODAMIENTO", "sum"),
	("Bonificación CP $", "amt:BONIFICACION_CP", "sum"),
]


# Hoja "Cálculos": importes derivados + totales y neto.
# Lo que el operador de payroll usa como guía / chequeo.
COLUMN_SPEC_CALCULOS: list[tuple[str, str, str]] = _IDENT_COLUMNS + [
	("Salario base", "id_salario", "first"),
	# Importes por hora
	("$ HD", "h_amt:HD", "sum"),
	("$ HN", "h_amt:HN", "sum"),
	("$ HFD", "h_amt:HFD", "sum"),
	("$ HFN", "h_amt:HFN", "sum"),
	("$ HED", "h_amt:HED", "sum"),
	("$ HEN", "h_amt:HEN", "sum"),
	("$ HEFD", "h_amt:HEFD", "sum"),
	("$ HEFN", "h_amt:HEFN", "sum"),
	# Beneficios y ausentismos agregados
	("Beneficios remunerados $", "d_amt_remunerados", "sum"),
	("Ausencias descuento $", "d_amt_no_remunerados", "sum"),
	("Incapacidades $", "d_amt_incapacidades", "sum"),
	# Otros pagos calculados
	("Inducción $", "amt:INDUCCION", "sum"),
	("Auxilio Transporte $", "auxilio_t", "first"),
	# Descuentos literales (réplica con importe firmado para el neto)
	("Adelanto Payflow $", "amt:ADELANTO_NOMINA_PAYFLOW", "sum"),
	("Sanitas $", "amt:DESCUENTO_SANITAS_PREMIUM", "sum"),
	("Gafas $", "amt:DESCUENTO_GAFAS", "sum"),
	("FONGIGA Fondo $", "amt:FONDO_EMPLEADOS_FONGIGA", "sum"),
	("Libranza Fincomercio $", "amt:LIBRANZA_FINCOMERCIO", "sum"),
	("Libranza Davivienda $", "amt:LIBRANZA_DAVIVIENDA", "sum"),
	("Libranza Compensar $", "amt:LIBRANZA_COMPENSAR", "sum"),
	("Libranza Comfenalco $", "amt:LIBRANZA_COMFENALCO", "sum"),
	("Préstamo Empresa $", "amt:PRESTAMO_EMPRESA", "sum"),
	("Préstamo FONGIGA $", "amt:PRESTAMO_FONGIGA", "sum"),
	("Pérdida Bonif. $", "amt:PERDIDA_BONIFICACION", "sum"),
	# Auxilios y bonificaciones pactados (réplica)
	("Auxilio Movilización $", "amt:AUXILIO_MOVILIZACION_DOM_FEST", "sum"),
	("Auxilio Rodamiento $", "amt:AUXILIO_RODAMIENTO", "sum"),
	("Bonificación CP $", "amt:BONIFICACION_CP", "sum"),
	# Totales
	("Total Devengado", "total_devengado", "sum"),
	("Total Descontado", "total_descontado", "sum"),
	("Neto a Pagar", "neto", "sum"),
]


# Alias para compat con código viejo que aún importa COLUMN_SPEC.
COLUMN_SPEC = COLUMN_SPEC_CALCULOS


REMUNERADOS_DIAS = (
	"VACACIONES", "DESCANSO", "DIA_FAMILIA", "DIA_CUMPLEANOS",
	"LICENCIA_MATERNIDAD", "LICENCIA_REMUNERADA", "LICENCIA_LUTO",
	"PERMISO_CALAMIDAD",
)
NO_REMUNERADOS_DIAS = (
	"LICENCIA_NO_REMUNERADA", "SUSPENSION_CONTRATO", "AUSENCIA_INJUSTIFICADA",
)
INCAPACIDADES = (
	"INCAPACIDAD_ENFERMEDAD_GENERAL", "INCAPACIDAD_ACCIDENTE_TRABAJO",
	"INCAPACIDAD_PAGADA_EMPRESA", "INCAPACIDAD_MAYOR_180_DIAS",
)


def _auxilio_for(rec: dict, params) -> float:
	"""Calcula el auxilio del empleado consolidando dias_trabajados
	(prorrateo sobre 24 días), salario y devengado del periodo (TP).
	"""
	dias_trab = float(rec.get("id_dias_trabajados") or 0)
	dias_no_rem = sum(rec["d_qty"].get(t, 0.0) for t in NO_REMUNERADOS_DIAS)
	if dias_trab <= 0:
		dias_trab = max(0.0, 24 - dias_no_rem)
	return auxilio_transporte.compute_for_period(
		rec.get("id_salario", 0.0),
		params,
		dias_trabajados=dias_trab,
		dias_no_remunerados=dias_no_rem,
		ingresos_periodo=rec["total_devengado"],
	)


def build_single_sheet(
	novedades: Iterable,
	params,
	employees_meta: dict[str, dict] | None = None,
	period_label: str = "",
) -> bytes:
	"""Devuelve el .xlsx en memoria (bytes).

	`novedades` es un iterable de objetos con los atributos de
	`EnrichedNovedad` (computed_amount/quantity ya seteados).
	`employees_meta` es un dict opcional `empleado_name -> {"nombres":..., "apellidos":..., "cedula":...}`
	para enriquecer las filas. Si falta, se usa el `documento_identidad`.
	"""
	by_employee = _aggregate(novedades, employees_meta or {})
	wb = Workbook()
	# Hoja 1 — HECHOS (cantidades crudas del archivo, sin importes
	# derivados). El operador de nómina la usa para alimentar su
	# sistema oficial.
	ws_hechos = wb.active
	_write_sheet(ws_hechos, by_employee, params, COLUMN_SPEC_HECHOS,
	             title=f"Hechos {period_label}" if period_label else "Hechos")
	# Hoja 2 — CÁLCULOS (importes derivados + totales + neto). Vista
	# de auditoría / guía para comparar contra el sistema oficial.
	ws_calc = wb.create_sheet()
	_write_sheet(ws_calc, by_employee, params, COLUMN_SPEC_CALCULOS,
	             title=f"Cálculos {period_label}" if period_label else "Cálculos")

	buf = io.BytesIO()
	wb.save(buf)
	return buf.getvalue()


def _write_sheet(ws, by_employee: dict, params, spec: list, title: str) -> None:
	"""Escribe una hoja completa con la spec dada: header, filas por
	empleado, fila TOTAL con fórmulas SUM y formatos de moneda / horas.
	"""
	ws.title = title[:31]

	headers = [s[0] for s in spec]
	ws.append(headers)
	header_fill = PatternFill("solid", fgColor="305496")
	header_font = Font(bold=True, color="FFFFFF")
	for cell in ws[1]:
		cell.fill = header_fill
		cell.font = header_font
		cell.alignment = Alignment(horizontal="center", wrap_text=True)
	ws.row_dimensions[1].height = 32

	for emp_id in sorted(by_employee.keys()):
		rec = by_employee[emp_id]
		row = [_resolve_value(rec, source, params) for _h, source, _agg in spec]
		ws.append(row)

	total_row_idx = ws.max_row + 1
	ws.cell(row=total_row_idx, column=1, value="TOTAL")
	for col_idx, (_h, source, agg) in enumerate(spec, start=1):
		if agg != "sum" or source.startswith("id_") or source == "auxilio_t":
			continue
		col_letter = get_column_letter(col_idx)
		ws.cell(
			row=total_row_idx,
			column=col_idx,
			value=f"=SUM({col_letter}2:{col_letter}{total_row_idx - 1})",
		)
	for cell in ws[total_row_idx]:
		cell.font = Font(bold=True)
		cell.fill = PatternFill("solid", fgColor="D9E1F2")

	for col_idx, (header, _src, _agg) in enumerate(spec, start=1):
		ws.column_dimensions[get_column_letter(col_idx)].width = max(
			14, min(28, len(header) + 2)
		)
		if "$" in header or "Devengado" in header or "Descontado" in header or "Neto" in header:
			for row_idx in range(2, ws.max_row + 1):
				ws.cell(row=row_idx, column=col_idx).number_format = '"$"#,##0.00'
		elif " h" in header or " d" in header:
			for row_idx in range(2, ws.max_row + 1):
				ws.cell(row=row_idx, column=col_idx).number_format = "0.00"


# ──────────────────────────────────────────────────────────────────────
# Agregación por empleado
# ──────────────────────────────────────────────────────────────────────

def _aggregate(novedades: Iterable, employees_meta: dict[str, dict]) -> dict[str, dict]:
	by_emp: dict[str, dict] = defaultdict(lambda: _empty_emp_record())
	for nov in novedades:
		key = nov.empleado or nov.documento_identidad or "SIN_EMPLEADO"
		rec = by_emp[key]
		# Identidad — first-non-empty: cada novedad del empleado puede
		# completar campos que la primera no traía (ej. HD trae cargo,
		# DESCANSO no; el aggregate las recibe en orden indefinido del
		# get_all).
		meta = employees_meta.get(key, {})
		payload = nov.raw_payload or {}

		def _fill(field, *candidates):
			if rec.get(field):
				return
			for c in candidates:
				if c:
					rec[field] = c
					return

		_fill("id_cedula", meta.get("cedula"), nov.documento_identidad)
		_fill("id_nombres", meta.get("nombres"), payload.get("empleado_nombre"))
		_fill("id_apellidos", meta.get("apellidos"))
		_fill("id_jornada", nov.tipo_jornada_snapshot)
		_fill("id_cargo", payload.get("cargo"))
		_fill("id_sucursal", payload.get("sucursal"), payload.get("sede"))
		if not rec.get("id_salario") and nov.salario_mensual:
			rec["id_salario"] = float(nov.salario_mensual)
		# dias_trabajados: tomar el máximo (todas las novedades del
		# empleado deberían reportar el mismo desde CLONK).
		dt = float(payload.get("dias_trabajados") or 0)
		if dt > rec.get("id_dias_trabajados", 0.0):
			rec["id_dias_trabajados"] = dt
		rec["id_set"] = True
		# Cantidades por tipo
		tipo = nov.tipo_novedad
		amount = float(nov.computed_amount or 0.0)
		qty = float(nov.computed_quantity or nov.cantidad or 0.0)
		if nov.unidad == "horas":
			rec["h_qty"][tipo] += qty
			rec["h_amt"][tipo] += amount
		elif nov.unidad == "dias":
			rec["d_qty"][tipo] += qty
			rec["d_amt"][tipo] += amount
		else:
			# cop / unidades
			rec["amt"][tipo] += amount
		# Totales por categoría agregada
		if tipo in REMUNERADOS_DIAS or tipo in INCAPACIDADES:
			rec["total_devengado"] += amount
		elif tipo in NO_REMUNERADOS_DIAS:
			rec["total_descontado"] += amount  # ausencia_injustificada ya negativo
		else:
			# horas, induccion, auxilios, bonificaciones suman al devengado.
			# Descuentos literales suman al descontado.
			from hubgh.hubgh.payroll.compute.literal import DESCUENTO_TYPES

			if tipo in DESCUENTO_TYPES:
				rec["total_descontado"] += amount
			else:
				rec["total_devengado"] += amount
	return dict(by_emp)


def _empty_emp_record() -> dict:
	return {
		"id_set": False,
		"id_cedula": "", "id_nombres": "", "id_apellidos": "",
		"id_jornada": "", "id_cargo": "", "id_sucursal": "",
		"id_salario": 0.0,
		"id_dias_trabajados": 0.0,
		"h_qty": defaultdict(float), "h_amt": defaultdict(float),
		"d_qty": defaultdict(float), "d_amt": defaultdict(float),
		"amt": defaultdict(float),
		"total_devengado": 0.0,
		"total_descontado": 0.0,
	}


def _resolve_value(rec: dict, source: str, params):
	if source.startswith("id_"):
		return rec.get(source, "")
	if source == "auxilio_t":
		return _auxilio_for(rec, params)
	if source.startswith("h_qty:"):
		return rec["h_qty"].get(source.split(":", 1)[1], 0.0)
	if source.startswith("h_amt:"):
		return rec["h_amt"].get(source.split(":", 1)[1], 0.0)
	if source.startswith("d_qty:"):
		return rec["d_qty"].get(source.split(":", 1)[1], 0.0)
	if source.startswith("amt:"):
		return rec["amt"].get(source.split(":", 1)[1], 0.0)
	if source == "d_amt_remunerados":
		return sum(rec["d_amt"].get(t, 0.0) for t in REMUNERADOS_DIAS)
	if source == "d_amt_no_remunerados":
		return sum(rec["d_amt"].get(t, 0.0) for t in NO_REMUNERADOS_DIAS)
	if source == "d_amt_incapacidades":
		return sum(rec["d_amt"].get(t, 0.0) for t in INCAPACIDADES)
	if source == "total_devengado":
		return rec["total_devengado"] + _auxilio_for(rec, params)
	if source == "total_descontado":
		return rec["total_descontado"]
	if source == "neto":
		return rec["total_devengado"] + _auxilio_for(rec, params) + rec["total_descontado"]
	return ""
