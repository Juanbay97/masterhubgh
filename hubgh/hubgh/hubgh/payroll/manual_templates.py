"""Plantillas manuales descargables desde el workspace.

Cuando los datos llegan en imagen, PDF o papel (descuentos casuales,
pérdida de bonificación, ascensos, traslados), el operador llena una
plantilla xlsx y la sube al Run como un archivo más. El adapter
`manual` la lee y emite las NovedadCanonica correspondientes.

Hay un template por concepto:

  descuentos.xlsx       → cedula | nombre | tipo | valor | motivo
  perdida_bonif.xlsx    → cedula | nombre | valor | motivo
  ascensos.xlsx         → cedula | nombre | cargo_nuevo | salario_nuevo | fecha_efectiva
  movimientos.xlsx      → cedula | nombre | nueva_sucursal | fecha_efectiva

El service expone `download_manual_template(template_id)` que devuelve
el xlsx generado al vuelo. El adapter manual reconoce el template por
el nombre del archivo y la firma de columnas.
"""

from __future__ import annotations

import io
from typing import Iterator

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


# ──────────────────────────────────────────────────────────────────────
# Definiciones declarativas de cada template
# ──────────────────────────────────────────────────────────────────────

# Cada plantilla: id → {label, sheet_title, headers, sample_rows, help}
TEMPLATES: dict[str, dict] = {
	"descuentos": {
		"label": "Descuentos manuales",
		"sheet_title": "Descuentos",
		"headers": [
			"Cédula",
			"Nombre (opcional)",
			"Tipo de descuento",
			"Valor a descontar",
			"Motivo / Observación",
		],
		"help": (
			"Plantilla para descuentos casuales que no llegan en archivo del proveedor. "
			"El sistema persiste cada fila como una novedad de tipo descuento manual."
		),
		"sample_rows": [
			["1019058261", "EJEMPLO PEREZ", "DOTACION", 35000, "Faltante de dotación julio"],
			["1017127331", "OTRO EJEMPLO", "PRESTAMO_EMPRESA", 150000, "Cuota préstamo abril"],
		],
		"valid_tipos": (
			"DESCUENTO_GAFAS",
			"DESCUENTO_SANITAS_PREMIUM",
			"PRESTAMO_EMPRESA",
			"PRESTAMO_FONGIGA",
			"DOTACION",
			"OTRO_DESCUENTO",
		),
	},
	"perdida_bonificacion": {
		"label": "Bonificación PDV (1/0)",
		"sheet_title": "Bonificación PDV",
		"headers": [
			"Cédula",
			"Nombre (opcional)",
			"Tiene bonificación (1=sí, 0=la pierde)",
			"Motivo (si pierde)",
		],
		"help": (
			"Sólo aplica a empleados de punto de venta (cargo Operativo). Marcá "
			"1 si la conserva y 0 si la pierde. El sistema NO calcula el valor de "
			"la bonificación — eso lo hace contabilidad. Acá sólo registramos el "
			"flag para que aparezca como columna 'Tiene bonif.' en la hoja Hechos."
		),
		"sample_rows": [
			["1019058261", "EJEMPLO PEREZ", 1, ""],
			["1017127331", "OTRO EJEMPLO", 0, "No cumplió KPI servicio"],
		],
	},
	"ascensos": {
		"label": "Ascensos",
		"sheet_title": "Ascensos",
		"headers": [
			"Cédula",
			"Nombre (opcional)",
			"Cargo nuevo",
			"Salario nuevo",
			"Fecha efectiva (YYYY-MM-DD)",
		],
		"help": (
			"Cambios de cargo con impacto en salario. El sistema actualiza el Contrato "
			"y aplica el nuevo salario para los cálculos del Run actual y siguientes."
		),
		"sample_rows": [
			["1019058261", "EJEMPLO PEREZ", "ANALISTA DE GESTION HUMANA", 2300000, "2026-04-15"],
		],
	},
	"maestro_empleados": {
		"label": "Maestro de empleados activos",
		"sheet_title": "Maestro empleados",
		"headers": [
			"Cédula",
			"Nombre (opcional)",
			"Cargo",
			"Sucursal (opcional)",
			"Tipo (Operativo/Administrativo)",
		],
		"help": (
			"Lista de TODOS los empleados activos en este periodo (operativos y "
			"administrativos). El sistema los incluye en la prenómina aunque NO "
			"aparezcan en el CLONK (típico de personal administrativo): a esos "
			"empleados se les asigna salario base por cargo y sólo se procesan "
			"sus descuentos / novedades manuales. El cargo debe existir en el "
			"catálogo de Cargos para que se les resuelva el salario."
		),
		"sample_rows": [
			["1019058261", "EJEMPLO PEREZ", "ANALISTA DE GESTION HUMANA", "Bogotá", "Administrativo"],
			["1128478178", "OTRO EJEMPLO", "AUXILIAR DE COCINA PDV", "Home 6", "Operativo"],
		],
	},
	"movimientos": {
		"label": "Movimientos / Traslados",
		"sheet_title": "Movimientos",
		"headers": [
			"Cédula",
			"Nombre (opcional)",
			"Nueva sucursal / PDV",
			"Fecha efectiva (YYYY-MM-DD)",
			"Observación",
		],
		"help": (
			"Traslados de sucursal o PDV. Informativo: NO afecta cálculo de pago, sólo "
			"actualiza el campo Sucursal del empleado para reportes."
		),
		"sample_rows": [
			["1019058261", "EJEMPLO PEREZ", "Home 5", "2026-04-15", "Cobertura licencia"],
		],
	},
}


def list_templates() -> list[dict]:
	"""Lista para la UI: id, label, help."""
	return [
		{"id": k, "label": v["label"], "help": v["help"]}
		for k, v in TEMPLATES.items()
	]


def build_template(template_id: str) -> bytes:
	"""Genera el xlsx de la plantilla con headers, fila de ayuda y 1-2
	ejemplos demostrativos en gris claro (el operador los reemplaza).
	"""
	spec = TEMPLATES.get(template_id)
	if not spec:
		raise ValueError(f"Template '{template_id}' no existe.")
	wb = Workbook()
	ws = wb.active
	ws.title = spec["sheet_title"]

	# Banner / ayuda en la fila 1.
	ws.cell(row=1, column=1, value=spec["help"])
	ws.cell(row=1, column=1).fill = PatternFill("solid", fgColor="FEF3C7")
	ws.cell(row=1, column=1).font = Font(italic=True, color="92400E")
	ws.cell(row=1, column=1).alignment = Alignment(wrap_text=True, vertical="center")
	ws.merge_cells(
		start_row=1, start_column=1, end_row=1, end_column=len(spec["headers"])
	)
	ws.row_dimensions[1].height = 36

	# Headers en fila 2.
	header_fill = PatternFill("solid", fgColor="305496")
	header_font = Font(bold=True, color="FFFFFF")
	for col_idx, header in enumerate(spec["headers"], start=1):
		cell = ws.cell(row=2, column=col_idx, value=header)
		cell.fill = header_fill
		cell.font = header_font
		cell.alignment = Alignment(horizontal="center", wrap_text=True)
		ws.column_dimensions[ws.cell(row=2, column=col_idx).column_letter].width = max(
			18, min(34, len(str(header)) + 4)
		)
	ws.row_dimensions[2].height = 28

	# Filas de ejemplo en gris claro.
	sample_fill = PatternFill("solid", fgColor="F1F5F9")
	sample_font = Font(italic=True, color="64748B")
	for row_offset, sample in enumerate(spec.get("sample_rows", []), start=3):
		for col_idx, value in enumerate(sample, start=1):
			cell = ws.cell(row=row_offset, column=col_idx, value=value)
			cell.fill = sample_fill
			cell.font = sample_font

	# Hoja secundaria con la lista de valores válidos si existe.
	if spec.get("valid_tipos"):
		ws_v = wb.create_sheet("Tipos válidos")
		ws_v.cell(row=1, column=1, value="Tipo de descuento")
		ws_v.cell(row=1, column=1).fill = header_fill
		ws_v.cell(row=1, column=1).font = header_font
		for i, t in enumerate(spec["valid_tipos"], start=2):
			ws_v.cell(row=i, column=1, value=t)
		ws_v.column_dimensions["A"].width = 32

	buf = io.BytesIO()
	wb.save(buf)
	return buf.getvalue()


# ──────────────────────────────────────────────────────────────────────
# Reconocimiento del template subido (lo usa adapters/manual.py)
# ──────────────────────────────────────────────────────────────────────

# Mapa header normalizado → template_id, para reconocer un archivo
# subido aunque el operador haya renombrado el archivo.
_FIRST_HEADER_TO_TEMPLATE: dict[str, str] = {
	"descuentos": "descuentos",
	"bonificacion pdv": "perdida_bonificacion",
	"bonificación pdv": "perdida_bonificacion",
	"pérdida bonificación": "perdida_bonificacion",
	"perdida bonificacion": "perdida_bonificacion",
	"ascensos": "ascensos",
	"movimientos": "movimientos",
	"maestro empleados": "maestro_empleados",
	"maestro de empleados": "maestro_empleados",
}


def identify_template_from_sheet(sheet_title: str) -> str | None:
	"""Mapea el título de la hoja al template_id si matcheamos. None si no."""
	if not sheet_title:
		return None
	import unicodedata

	key = unicodedata.normalize("NFKD", sheet_title).strip().lower()
	key = "".join(c for c in key if not unicodedata.combining(c))
	for needle, tid in _FIRST_HEADER_TO_TEMPLATE.items():
		if needle in key:
			return tid
	return None
