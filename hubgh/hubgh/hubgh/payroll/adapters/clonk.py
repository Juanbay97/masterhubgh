"""Adapter CLONK — control de horas y novedades en hojas Excel.

CLONK exporta tres hojas:
  - "Resumen": una fila por empleado con totales agregados de horas
    (HD, HN, HFD, HFN, HED, HEN, HEFD, HEFN) + NR/NnR/DnR.
  - "Detalle de Tiempos": una fila por empleado-día con marcaciones.
  - "Novedades": una fila por evento, conceptos en columnas pareadas
    (fecha_inicio | fecha_fin) por concepto.

Este adapter lee `Resumen` para emitir novedades canónicas de horas y
`Novedades` para emitir novedades canónicas de ausentismos/beneficios.
La hoja `Detalle` se ignora en v1 (redundante para totales).

Período: inferido del rango de fechas presente en `Detalle de Tiempos`
o del nombre del archivo; el corte CLONK estándar para TC es 16-15.
"""

from __future__ import annotations

import re
from datetime import date, datetime
from pathlib import Path
from typing import Iterator

from hubgh.hubgh.payroll import catalogs
from hubgh.hubgh.payroll.adapters import NovedadCanonica


SOURCE_ID = "clonk"

REQUIRED_SHEETS = ("Resumen", "Detalle de Tiempos", "Novedades")
RESUMEN_HOUR_COLUMNS = ("HD", "HN", "HFD", "HFN", "HED", "HEN", "HEFD", "HEFN")

# Mapa columna → tipo canónico de novedad para los 11 conceptos pareados
# de la hoja Novedades. La clave es el header literal del archivo;
# el valor es el id de NOVEDAD_TYPES.
NOVEDADES_CONCEPT_TO_TYPE: dict[str, str] = {
	"Incapacidad AT": "INCAPACIDAD_ACCIDENTE_TRABAJO",
	"Vacaciones": "VACACIONES",
	"Incapacidad EG": "INCAPACIDAD_ENFERMEDAD_GENERAL",
	"Descanso": "DESCANSO",
	"DIA CUMPLEAÑOS": "DIA_CUMPLEANOS",
	"DIA CUMPLEANOS": "DIA_CUMPLEANOS",
	"Maternidad": "LICENCIA_MATERNIDAD",
	"AUSENTISMO": "AUSENCIA_INJUSTIFICADA",
	"INDUCCION": "INDUCCION",
	"INDUCCIÓN": "INDUCCION",
	"DIA FAMILIA": "DIA_FAMILIA",
	"L. No Remunerada": "LICENCIA_NO_REMUNERADA",
	"Suspensión": "SUSPENSION_CONTRATO",
	"Suspension": "SUSPENSION_CONTRATO",
}

_FILENAME_PATTERN = re.compile(r"clonk|toda\s+la\s+empresa", re.IGNORECASE)


# ──────────────────────────────────────────────────────────────────────
# matches: scoring por archivo
# ──────────────────────────────────────────────────────────────────────

def matches(file_meta) -> int:
	"""Devuelve un score 0..3 sobre qué tan probable es que sea CLONK.

	`file_meta` es un dict con al menos `filename` y opcionalmente
	`sheets` (lista de nombres de hojas).
	"""
	score = 0
	filename = (file_meta or {}).get("filename", "") or ""
	if _FILENAME_PATTERN.search(filename):
		score += 1
	sheets = set((file_meta or {}).get("sheets") or [])
	if all(s in sheets for s in REQUIRED_SHEETS):
		score += 2
	elif "Resumen" in sheets and "Novedades" in sheets:
		score += 1
	return score


# ──────────────────────────────────────────────────────────────────────
# detect_period: deduce (year, month) del archivo
# ──────────────────────────────────────────────────────────────────────

def detect_period(workbook) -> tuple[int, int] | None:
	"""Devuelve el (year, month) lógico del Run a partir de las fechas
	presentes en la hoja `Detalle de Tiempos`.

	El mes lógico = mes del último día del rango (fecha de pago).
	Si no se puede leer el rango, devuelve None.
	"""
	if "Detalle de Tiempos" not in workbook.sheetnames:
		return None
	ws = workbook["Detalle de Tiempos"]
	# La columna Fecha es la 8 (índice 7, 1-based H).
	max_dt = None
	for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
		if idx > 200:  # tomar muestra para no recorrer 11k filas
			break
		fecha = row[7] if len(row) > 7 else None
		if not fecha:
			continue
		dt = _coerce_date(fecha)
		if dt and (max_dt is None or dt > max_dt):
			max_dt = dt
	if max_dt is None:
		# Fallback: leer fechas hasta el final
		for row in ws.iter_rows(min_row=2, values_only=True):
			fecha = row[7] if len(row) > 7 else None
			dt = _coerce_date(fecha)
			if dt and (max_dt is None or dt > max_dt):
				max_dt = dt
	if max_dt is None:
		return None
	return max_dt.year, max_dt.month


# ──────────────────────────────────────────────────────────────────────
# parse: emite NovedadCanonica por cada fila válida
# ──────────────────────────────────────────────────────────────────────

def parse(workbook) -> Iterator[NovedadCanonica]:
	"""Itera novedades canónicas leyendo Resumen + Novedades.

	Antes de iterar Novedades construye un índice por documento desde
	Resumen + Detalle (para `dias_trabajados`). El índice se inyecta
	en `_parse_resumen` y `_parse_novedades` para enriquecer cada
	NovedadCanonica con jornada / cargo / sucursal / dias_trabajados.
	"""
	emp_index = _build_employee_index(workbook)
	# Lo cuelgo del workbook para que `_parse_resumen` lo lea sin cambiar
	# su firma (workaround: openpyxl Workbook acepta atributos).
	workbook._pwsp_emp_index = emp_index  # noqa: SLF001
	yield from _parse_resumen(workbook)
	yield from _parse_novedades(workbook, emp_index)


def _build_employee_index(workbook) -> dict[str, dict]:
	"""Mapa documento → {contrato_text, cargo, sucursal, dias_trabajados}.

	Los primeros tres salen de Resumen. `dias_trabajados` se cuenta del
	Detalle de Tiempos: días con horas > 0 (HD/HN/HFD/HFN/HE*) en alguna
	columna H*. Se usa para prorratear el auxilio de transporte sobre 24
	días/mes.
	"""
	idx: dict[str, dict] = {}
	if "Resumen" in workbook.sheetnames:
		ws = workbook["Resumen"]
		rows = ws.iter_rows(min_row=1, values_only=True)
		header = next(rows, None)
		if header:
			col = {str(h).strip(): i for i, h in enumerate(header) if h is not None}
			doc_idx = col.get("Documento")
			if doc_idx is not None:
				for row in rows:
					documento = _str_id(row[doc_idx] if doc_idx < len(row) else None)
					if not documento or documento in idx:
						continue
					idx[documento] = {
						"contrato_text": str(row[col["Contrato"]]).strip() if "Contrato" in col and row[col["Contrato"]] else "",
						"cargo": row[col["Cargo"]] if "Cargo" in col else None,
						"sucursal": row[col["Sucursal"]] if "Sucursal" in col else None,
						"dias_trabajados": 0.0,
					}

	# Contar dias_trabajados desde Detalle de Tiempos.
	if "Detalle de Tiempos" in workbook.sheetnames:
		ws = workbook["Detalle de Tiempos"]
		rows = ws.iter_rows(min_row=1, values_only=True)
		header = next(rows, None)
		if header:
			col = {str(h).strip(): i for i, h in enumerate(header) if h is not None}
			doc_idx = col.get("Documento")
			fecha_idx = col.get("Fecha")
			h_cols = [col[h] for h in RESUMEN_HOUR_COLUMNS if h in col]
			if doc_idx is not None and fecha_idx is not None and h_cols:
				dias_set: dict[str, set] = {}
				for row in rows:
					documento = _str_id(row[doc_idx] if doc_idx < len(row) else None)
					if not documento:
						continue
					fecha = row[fecha_idx] if fecha_idx < len(row) else None
					if not fecha or fecha == "-":
						continue
					# Día con al menos 1h en cualquier columna H*.
					tiene_horas = False
					for c in h_cols:
						val = row[c] if c < len(row) else None
						try:
							if float(val or 0) > 0:
								tiene_horas = True
								break
						except (TypeError, ValueError):
							continue
					if not tiene_horas:
						continue
					dias_set.setdefault(documento, set()).add(str(fecha))
				for documento, fechas in dias_set.items():
					if documento not in idx:
						idx[documento] = {"contrato_text": "", "cargo": None, "sucursal": None, "dias_trabajados": 0.0}
					idx[documento]["dias_trabajados"] = float(len(fechas))
	return idx


def _parse_resumen(workbook) -> Iterator[NovedadCanonica]:
	if "Resumen" not in workbook.sheetnames:
		return
	ws = workbook["Resumen"]
	rows = ws.iter_rows(min_row=1, values_only=True)
	header = next(rows, None)
	if not header:
		return
	col = {str(h).strip(): i for i, h in enumerate(header) if h is not None}
	doc_idx = col.get("Documento")
	contrato_idx = col.get("Contrato")
	if doc_idx is None:
		return
	hour_idx = {h: col[h] for h in RESUMEN_HOUR_COLUMNS if h in col}
	for row in rows:
		documento = _str_id(row[doc_idx] if doc_idx is not None else None)
		if not documento:
			continue
		contrato_text = (
			str(row[contrato_idx]).strip() if contrato_idx is not None and row[contrato_idx] else ""
		)
		emp_meta = (workbook._pwsp_emp_index or {}).get(documento, {}) if hasattr(workbook, "_pwsp_emp_index") else {}
		raw_emp = {
			"empleado_nombre": row[col["Empleado"]] if "Empleado" in col else None,
			"contrato_text": contrato_text,
			"cargo": row[col["Cargo"]] if "Cargo" in col else None,
			"sucursal": row[col["Sucursal"]] if "Sucursal" in col else None,
			"dias_trabajados": emp_meta.get("dias_trabajados", 0.0),
			"sheet": "Resumen",
		}
		for hour_key, idx in hour_idx.items():
			value = row[idx]
			if value is None:
				continue
			try:
				cantidad = float(value)
			except (TypeError, ValueError):
				continue
			if cantidad <= 0:
				continue
			yield NovedadCanonica(
				documento_identidad=documento,
				tipo_novedad=hour_key,  # HD, HN, HFD, ... ya son ids canónicos
				cantidad=round(cantidad, 4),
				unidad="horas",
				raw_payload={**raw_emp, "campo": hour_key, "valor_crudo": value},
			)


def _parse_novedades(workbook, emp_index: dict[str, dict] | None = None) -> Iterator[NovedadCanonica]:
	if "Novedades" not in workbook.sheetnames:
		return
	ws = workbook["Novedades"]
	emp_index = emp_index or {}
	rows = ws.iter_rows(min_row=1, values_only=True)
	header = next(rows, None)
	if not header:
		return
	# Identificar pares de columnas (concepto, vacío) → cada concepto ocupa
	# 2 columnas: la de inicio (con el header del concepto) y la de fin
	# (header vacío contiguo).
	concept_columns: list[tuple[int, str]] = []
	for i, name in enumerate(header):
		if not name:
			continue
		clean = str(name).strip()
		if clean in NOVEDADES_CONCEPT_TO_TYPE:
			concept_columns.append((i, clean))
	# Documento puede estar bajo "Cédula", "Cedula" o "Documento".
	doc_idx = None
	nombre_idx = None
	sede_idx = None
	for i, name in enumerate(header):
		if name is None:
			continue
		clean_lower = str(name).strip().lower()
		if doc_idx is None and clean_lower in {"cédula", "cedula", "documento"}:
			doc_idx = i
		if nombre_idx is None and clean_lower in {"nombre", "empleado"}:
			nombre_idx = i
		if sede_idx is None and clean_lower in {"sede", "sucursal", "pdv"}:
			sede_idx = i
	if doc_idx is None:
		return
	for row in rows:
		documento = _str_id(row[doc_idx])
		if not documento:
			continue
		emp_meta = emp_index.get(documento, {}) or {}
		raw_emp = {
			"empleado_nombre": row[nombre_idx] if nombre_idx is not None else None,
			"sede": row[sede_idx] if sede_idx is not None else None,
			# Heredados del Resumen / Detalle para enrichment + auxilio.
			"contrato_text": emp_meta.get("contrato_text", ""),
			"cargo": emp_meta.get("cargo"),
			"sucursal": emp_meta.get("sucursal"),
			"dias_trabajados": emp_meta.get("dias_trabajados", 0.0),
			"sheet": "Novedades",
		}
		for col_idx, concept_label in concept_columns:
			start = row[col_idx] if col_idx < len(row) else None
			end = row[col_idx + 1] if col_idx + 1 < len(row) else None
			if not start and not end:
				continue
			fecha_desde = _coerce_date(start)
			fecha_hasta = _coerce_date(end) or fecha_desde
			if fecha_desde is None:
				continue
			dias = (fecha_hasta - fecha_desde).days + 1 if fecha_hasta else 1
			tipo_canonico = NOVEDADES_CONCEPT_TO_TYPE.get(concept_label, "OTRO")
			yield NovedadCanonica(
				documento_identidad=documento,
				tipo_novedad=tipo_canonico,
				cantidad=float(max(dias, 1)),
				unidad="dias",
				fecha_desde=fecha_desde.isoformat(),
				fecha_hasta=fecha_hasta.isoformat() if fecha_hasta else None,
				raw_payload={**raw_emp, "concepto_clonk": concept_label},
			)


# ──────────────────────────────────────────────────────────────────────
# Helpers privados
# ──────────────────────────────────────────────────────────────────────

def _str_id(value) -> str:
	if value is None:
		return ""
	if isinstance(value, float) and value.is_integer():
		return str(int(value))
	return str(value).strip()


def _coerce_date(value) -> date | None:
	if value is None or value == "" or value == "-":
		return None
	if isinstance(value, datetime):
		return value.date()
	if isinstance(value, date):
		return value
	if isinstance(value, str):
		text = value.strip()
		# Formatos típicos del CLONK: dd/mm/yyyy o yyyy-mm-dd
		for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
			try:
				return datetime.strptime(text, fmt).date()
			except ValueError:
				continue
	return None


# ──────────────────────────────────────────────────────────────────────
# API conveniencia para abrir desde path
# ──────────────────────────────────────────────────────────────────────

def open_workbook(path: str | Path):
	"""Abre el archivo CLONK con openpyxl en modo read-only."""
	import openpyxl

	return openpyxl.load_workbook(path, data_only=True, read_only=True)
