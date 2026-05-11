"""Detector de fuente por scoring multi-señal.

Cada adapter expone `matches(file_meta) -> int` con score 0..3. El detector
toma metadata del archivo (filename, sheets, columns_first_sheet) y le
pregunta a cada adapter en orden. Gana el primer adapter con score >= 2;
en empate, el primero registrado en `ALL_ADAPTERS` toma prioridad.

Si ningún adapter llega a 2, devuelve `unknown` y el usuario selecciona
manualmente.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any


def _registered_adapters():
	"""Importa diferido para evitar ciclos al cargar Frappe."""
	from hubgh.hubgh.payroll.adapters import (
		clonk,
		fincomercio,
		fongiga,
		libranza_compensar,
		libranza_davivienda,
		manual,
		payflow,
	)

	return [
		clonk,
		payflow,
		fincomercio,
		fongiga,
		libranza_davivienda,
		libranza_compensar,
		manual,
	]


def file_meta_from_path(path: str | Path) -> dict[str, Any]:
	"""Extrae metadata barata de un archivo Excel sin parsear el contenido."""
	from openpyxl import load_workbook

	p = Path(path)
	wb = load_workbook(p, data_only=True, read_only=True)
	try:
		sheets = list(wb.sheetnames)
		columns: list[str] = []
		if sheets:
			ws = wb[sheets[0]]
			first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
			if first_row:
				columns = [str(c).strip() for c in first_row if c]
		return {
			"filename": p.name,
			"sheets": sheets,
			"columns_first_sheet": columns,
		}
	finally:
		wb.close()


def detect_source(file_meta: dict[str, Any]) -> str:
	"""Devuelve el id de la fuente con mayor score >= 2, o `unknown`."""
	best_id = "unknown"
	best_score = 1  # umbral mínimo: hay que superar 1
	for adapter in _registered_adapters():
		score = int(adapter.matches(file_meta) or 0)
		if score > best_score:
			best_id = getattr(adapter, "SOURCE_ID", "unknown")
			best_score = score
	return best_id


def detect_period(workbook, source_id: str) -> tuple[int, int] | None:
	"""Delega la detección de periodo al adapter correspondiente."""
	for adapter in _registered_adapters():
		if getattr(adapter, "SOURCE_ID", None) == source_id:
			return adapter.detect_period(workbook)
	return None
