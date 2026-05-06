import base64
from io import BytesIO

import frappe
from frappe.utils import nowdate

from hubgh.hubgh.page.seleccion_documentos.seleccion_documentos import (
	_has_medical_exam_access,
	list_medical_exam_candidates,
	list_medical_exam_history,
	set_medical_concept,
	upload_medical_exam_document,
)


__all__ = [
	"export_proximos_examenes_xlsx",
	"list_medical_exam_candidates",
	"list_medical_exam_history",
	"set_medical_concept",
	"upload_medical_exam_document",
]


@frappe.whitelist()
def export_proximos_examenes_xlsx():
	"""Genera un xlsx con los exámenes médicos agendados desde hoy en adelante.

	Columnas: Fecha, Hora, Cédula, Nombre completo, Ciudad, Sede, Cargo,
	Tipo cargo (Operativo/Administrativo).

	Solo usuarios con acceso SST/Selección/GH lo pueden descargar.

	Returns:
		dict con `filename` (sugerido) y `content_b64` (xlsx base64 para que
		el JS lo decodifique y dispare la descarga via Blob).
	"""
	if frappe.session.user != "Administrator" and not _has_medical_exam_access(frappe.session.user):
		frappe.throw("No autorizado")

	import openpyxl
	from openpyxl.styles import Font, PatternFill, Alignment

	today = nowdate()
	rows = frappe.db.sql(
		"""
		SELECT
			cita.fecha_cita,
			cita.hora_cita,
			cita.sede_seleccionada,
			cita.cargo_al_enviar,
			cita.candidato,
			cand.nombres,
			cand.primer_apellido,
			cand.segundo_apellido,
			cand.numero_documento,
			cand.ciudad,
			cand.celular,
			cand.email,
			cargo.nombre AS cargo_nombre,
			cargo.tipo_cargo
		FROM `tabCita Examen Medico` cita
		LEFT JOIN `tabCandidato` cand ON cita.candidato = cand.name
		LEFT JOIN `tabCargo` cargo ON cita.cargo_al_enviar = cargo.name
		WHERE cita.estado = 'Agendada' AND cita.fecha_cita >= %s
		ORDER BY cita.fecha_cita ASC, cita.hora_cita ASC
		""",
		(today,),
		as_dict=True,
	)

	wb = openpyxl.Workbook()
	ws = wb.active
	ws.title = "Próximos exámenes"

	headers = [
		"Fecha",
		"Hora",
		"Cédula",
		"Nombre completo",
		"Celular",
		"Email",
		"Ciudad",
		"Sede",
		"Cargo",
		"Tipo cargo",
	]
	ws.append(headers)

	# Style del header
	header_fill = PatternFill(start_color="1D4ED8", end_color="1D4ED8", fill_type="solid")
	header_font = Font(bold=True, color="FFFFFF")
	for col_idx in range(1, len(headers) + 1):
		cell = ws.cell(row=1, column=col_idx)
		cell.fill = header_fill
		cell.font = header_font
		cell.alignment = Alignment(horizontal="center", vertical="center")

	for r in rows:
		nombre_completo = " ".join(
			str(p).strip()
			for p in (r.nombres, r.primer_apellido, r.segundo_apellido)
			if p and str(p).strip()
		)
		ws.append(
			[
				str(r.fecha_cita) if r.fecha_cita else "",
				str(r.hora_cita) if r.hora_cita else "",
				r.numero_documento or r.candidato or "",
				nombre_completo or r.candidato or "",
				r.celular or "",
				r.email or "",
				r.ciudad or "",
				r.sede_seleccionada or "",
				r.cargo_nombre or r.cargo_al_enviar or "",
				r.tipo_cargo or "Operativo",
			]
		)

	# Anchos de columna razonables
	col_widths = [12, 10, 16, 38, 16, 30, 14, 28, 36, 16]
	for idx, w in enumerate(col_widths, start=1):
		ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = w

	output = BytesIO()
	wb.save(output)
	return {
		"filename": f"proximos_examenes_{today}.xlsx",
		"content_b64": base64.b64encode(output.getvalue()).decode("ascii"),
		"count": len(rows),
	}
