# Copyright (c) 2026, Antigravity and contributors
# For license information, please see license.txt

"""
sst_seguimiento_examenes — Bandeja SST de seguimiento de citas activas.

Expone endpoints para listar, agendar, anotar observaciones, cambiar estado
y exportar citas de examen médico que aún no tienen concepto_resultado.

Solo accesible para: System Manager, HR SST, HR Selection, Gestión Humana.
"""

from __future__ import annotations

import base64
from io import BytesIO

import frappe
from frappe import _
from frappe.utils import nowdate

# ---------------------------------------------------------------------------
# Constantes
# ---------------------------------------------------------------------------

_ALLOWED_ROLES = {"System Manager", "HR SST", "HR Selection", "Gestión Humana"}
_ESTADOS_VALIDOS = {"Realizada", "Aplazada", "No Asistió", "Cancelada"}
_CONCEPTOS_VALIDOS = {"Favorable", "Desfavorable", "Aplazado"}
_OBS_MAX_CHARS = 5000

# ---------------------------------------------------------------------------
# Helpers de acceso
# ---------------------------------------------------------------------------


def _has_seguimiento_access(user: str | None = None) -> bool:
    """Retorna True si el usuario tiene acceso a la bandeja de seguimiento.

    Administrator siempre tiene acceso. Para el resto, verifica intersección
    con los 4 roles autorizados de la bandeja.
    """
    user = user or frappe.session.user
    if user == "Administrator":
        return True
    roles = set(frappe.get_roles(user) or [])
    return bool(roles & _ALLOWED_ROLES)


def _require_access() -> None:
    """Lanza PermissionError si el usuario no tiene acceso a la bandeja."""
    if not _has_seguimiento_access(frappe.session.user):
        frappe.throw(
            _("No tienes permisos para acceder a la bandeja de seguimiento de exámenes."),
            frappe.PermissionError,
        )


# ---------------------------------------------------------------------------
# Helper de query
# ---------------------------------------------------------------------------


def _query_seguimiento_examenes(
    filters: dict | None,
    limit: int | None = 100,
    offset: int = 0,
) -> tuple[list[dict], int]:
    """Ejecuta la query base de citas activas con filtros opcionales.

    Excluye citas con concepto_resultado registrado y citas Canceladas.
    'No Asistió' permanece en bandeja hasta que GH registre el concepto.

    Retorna (rows, total_count).
    """
    filters = filters or {}

    # Condiciones base — se construyen con placeholders %s para evitar inyección SQL
    where_clauses = [
        "(cita.concepto_resultado IS NULL OR cita.concepto_resultado = '')",
        "cita.estado NOT IN ('Cancelada')",
    ]
    params: list = []

    # Filtro: rango de fechas (se ignora si solo_manuales_sin_datos está activo)
    solo_manuales = frappe.parse_json(filters.get("solo_manuales_sin_datos", False))
    if not solo_manuales:
        fecha_desde = filters.get("fecha_desde")
        fecha_hasta = filters.get("fecha_hasta")
        if fecha_desde:
            where_clauses.append("(cita.fecha_cita IS NULL OR cita.fecha_cita >= %s)")
            params.append(fecha_desde)
        if fecha_hasta:
            where_clauses.append("(cita.fecha_cita IS NULL OR cita.fecha_cita <= %s)")
            params.append(fecha_hasta)

    # Filtro: toggle "solo manuales sin datos"
    if solo_manuales:
        where_clauses.append("cand.modo_agendamiento_examen = 'Manual'")
        where_clauses.append("cita.fecha_cita IS NULL")

    # Filtro: estados (lista)
    estados = filters.get("estados")
    if estados:
        if isinstance(estados, str):
            try:
                import json
                estados = json.loads(estados)
            except Exception:
                estados = [estados]
        if estados:
            placeholders = ", ".join(["%s"] * len(estados))
            where_clauses.append(f"cita.estado IN ({placeholders})")
            params.extend(estados)

    # Filtro: modo (Manual / Autogestionado)
    modo = filters.get("modo")
    if modo and modo != "Ambos":
        where_clauses.append("cand.modo_agendamiento_examen = %s")
        params.append(modo)

    # Filtro: ciudad
    ciudad = filters.get("ciudad")
    if ciudad:
        where_clauses.append("cand.ciudad = %s")
        params.append(ciudad)

    # Filtro: tipo_cargo
    tipo_cargo = filters.get("tipo_cargo")
    if tipo_cargo:
        where_clauses.append("cargo.tipo_cargo = %s")
        params.append(tipo_cargo)

    # Filtro: sede
    sede = filters.get("sede")
    if sede:
        where_clauses.append("cita.sede_seleccionada = %s")
        params.append(sede)

    # Filtro: búsqueda libre (nombre o documento)
    search = (filters.get("search") or "").strip()
    if search:
        like_val = f"%{search}%"
        where_clauses.append(
            "(cand.nombres LIKE %s OR cand.primer_apellido LIKE %s OR cand.numero_documento LIKE %s)"
        )
        params.extend([like_val, like_val, like_val])

    where_sql = " AND ".join(where_clauses)

    select_sql = """
        SELECT
            cita.name,
            cita.estado,
            cita.fecha_cita,
            cita.hora_cita,
            cita.sede_seleccionada,
            cita.cargo_al_enviar,
            cita.ips,
            cita.observaciones_sst,
            cand.name AS candidato,
            cand.nombres,
            cand.primer_apellido,
            cand.segundo_apellido,
            cand.numero_documento,
            cand.ciudad,
            cand.celular,
            cand.email,
            cand.modo_agendamiento_examen AS modo,
            cargo.nombre AS cargo_nombre,
            cargo.tipo_cargo
        FROM `tabCita Examen Medico` cita
        LEFT JOIN `tabCandidato` cand ON cita.candidato = cand.name
        LEFT JOIN `tabCargo` cargo ON cita.cargo_al_enviar = cargo.name
        WHERE {where}
        ORDER BY (cita.fecha_cita IS NULL) DESC, cita.fecha_cita ASC, cita.hora_cita ASC
    """.format(where=where_sql)

    count_sql = """
        SELECT COUNT(*) AS total
        FROM `tabCita Examen Medico` cita
        LEFT JOIN `tabCandidato` cand ON cita.candidato = cand.name
        LEFT JOIN `tabCargo` cargo ON cita.cargo_al_enviar = cargo.name
        WHERE {where}
    """.format(where=where_sql)

    # total_count (mismos params, sin limit/offset)
    count_result = frappe.db.sql(count_sql, params, as_dict=True)
    total = (count_result[0].get("total") or 0) if count_result else 0

    # Paginación
    if limit is not None:
        select_sql += " LIMIT %s OFFSET %s"
        params_paged = params + [limit, offset]
    else:
        params_paged = params

    rows_raw = frappe.db.sql(select_sql, params_paged, as_dict=True)

    rows = []
    for r in rows_raw:
        row = dict(r)
        # Nombre completo
        row["nombre_completo"] = " ".join(
            str(p).strip()
            for p in (row.get("nombres"), row.get("primer_apellido"), row.get("segundo_apellido"))
            if p and str(p).strip()
        )
        # Preview de observaciones (80 chars)
        obs = row.get("observaciones_sst") or ""
        row["observaciones_preview"] = obs[:80] + ("..." if len(obs) > 80 else "")
        # Indicador de datos faltantes (modo Manual sin fecha agendada)
        row["datos_faltantes"] = bool(
            row.get("modo") == "Manual" and not row.get("fecha_cita")
        )
        rows.append(row)

    return rows, int(total)


# ---------------------------------------------------------------------------
# Endpoints públicos
# ---------------------------------------------------------------------------


@frappe.whitelist()
def list_seguimiento_examenes(filters=None, limit=100, offset=0):
    """Lista citas activas sin concepto_resultado registrado.

    Args:
        filters: dict con claves opcionales: fecha_desde, fecha_hasta, estados,
                 modo, ciudad, tipo_cargo, sede, solo_manuales_sin_datos, search.
        limit: máx 500, default 100.
        offset: desplazamiento para paginación, default 0.

    Returns:
        {"rows": [...], "total": N, "limit": N, "offset": N}
    """
    _require_access()

    filters = frappe.parse_json(filters) if filters else {}
    limit = max(1, min(int(limit or 100), 500))
    offset = max(0, int(offset or 0))

    rows, total = _query_seguimiento_examenes(filters, limit=limit, offset=offset)
    return {"rows": rows, "total": total, "limit": limit, "offset": offset}


@frappe.whitelist()
def set_cita_schedule(cita_name, fecha, hora, sede=None, force_pasado=False):
    """Agenda o reagenda una cita (fecha, hora, sede).

    Transiciona el estado a 'Agendada'. Usa frappe.db.set_value para bypass
    del read_only=1 en sede_seleccionada.

    Args:
        cita_name: name del documento Cita Examen Medico.
        fecha: str en formato YYYY-MM-DD.
        hora: str en formato HH:MM o HH:MM:SS.
        sede: nombre de la sede (opcional).
        force_pasado: bool — si True, permite fechas pasadas sin error.

    Returns:
        {"ok": True, "cita": cita_name, "fecha": fecha, "hora": hora}
    """
    _require_access()

    if not cita_name:
        frappe.throw(_("Falta cita_name"), frappe.ValidationError)
    if not frappe.db.exists("Cita Examen Medico", cita_name):
        frappe.throw(_("Cita no encontrada: {0}").format(cita_name), frappe.ValidationError)

    from frappe.utils import getdate

    fecha_parsed = getdate(fecha)
    today = getdate(nowdate())

    force = frappe.parse_json(force_pasado) if isinstance(force_pasado, str) else bool(force_pasado)

    if fecha_parsed < today and not force:
        frappe.throw(
            _("La fecha {0} es anterior a hoy ({1}). Si querés agendar igualmente, "
              "confirmá con force_pasado=True.").format(fecha, today),
            frappe.ValidationError,
        )

    # Normalizar hora a HH:MM:SS
    if hora and len(str(hora).split(":")) == 2:
        hora = str(hora) + ":00"

    vals = {
        "estado": "Agendada",
        "fecha_cita": fecha,
        "hora_cita": hora,
    }
    if sede is not None:
        vals["sede_seleccionada"] = sede

    frappe.db.set_value("Cita Examen Medico", cita_name, vals, update_modified=True)

    return {"ok": True, "cita": cita_name, "fecha": fecha, "hora": hora}


@frappe.whitelist()
def list_sedes_for_cita(cita_name):
    """Devuelve sedes activas de la IPS para la ciudad del candidato de la cita.

    Lo usa el dialog "Editar agendamiento" en la bandeja para autocompletar
    el dropdown de sedes válidas. Si la cita no tiene candidato resoluble o
    no hay IPS activa para la ciudad, retorna lista vacía.

    Returns:
        dict con `ciudad` y `sedes` (lista de dicts con nombre_sede,
        ciudad, direccion).
    """
    _require_access()
    if not frappe.db.exists("Cita Examen Medico", cita_name):
        frappe.throw("Cita no existe.", frappe.ValidationError)
    candidato = frappe.db.get_value("Cita Examen Medico", cita_name, "candidato") or ""
    if not candidato:
        return {"ciudad": "", "sedes": []}
    ciudad = frappe.db.get_value("Candidato", candidato, "ciudad") or ""
    if not ciudad:
        return {"ciudad": "", "sedes": []}
    from hubgh.hubgh.examen_medico.cita_service import (
        _get_sedes_for_ciudad,
        _resolve_active_ips_for_ciudad,
    )
    ips_name = _resolve_active_ips_for_ciudad(ciudad)
    if not ips_name:
        return {"ciudad": ciudad, "sedes": []}
    ips_doc = frappe.get_doc("IPS", ips_name).as_dict()
    sedes = _get_sedes_for_ciudad(ips_doc, ciudad)
    return {"ciudad": ciudad, "sedes": sedes}


@frappe.whitelist()
def set_cita_observaciones(cita_name, texto):
    """Guarda observaciones SST en la cita.

    Args:
        cita_name: name del documento Cita Examen Medico.
        texto: contenido (str, puede ser vacío para limpiar notas).

    Returns:
        {"ok": True}
    """
    _require_access()

    if not cita_name:
        frappe.throw(_("Falta cita_name"), frappe.ValidationError)
    if not frappe.db.exists("Cita Examen Medico", cita_name):
        frappe.throw(_("Cita no encontrada: {0}").format(cita_name), frappe.ValidationError)

    texto = texto or ""
    if len(texto) > _OBS_MAX_CHARS:
        frappe.throw(
            _("El texto excede el límite de {0} caracteres.").format(_OBS_MAX_CHARS),
            frappe.ValidationError,
        )

    frappe.db.set_value(
        "Cita Examen Medico",
        cita_name,
        "observaciones_sst",
        texto,
        update_modified=True,
    )
    return {"ok": True}


@frappe.whitelist()
def set_cita_outcome(cita_name, estado, concepto=None, motivo=None, instrucciones=None):
    """Registra el resultado de una cita desde la bandeja de seguimiento.

    Delega en cita_service.set_exam_outcome (con el bug fix aplicado).

    Args:
        cita_name: name del documento.
        estado: uno de Realizada / Aplazada / No Asistió / Cancelada.
        concepto: requerido si estado=Realizada (Favorable / Desfavorable / Aplazado).
        motivo: requerido si estado=Aplazada (≥5 chars) o Cancelada.
        instrucciones: opcional, para Aplazada.

    Returns:
        {"ok": True, "cita": cita_name, "estado": estado}
    """
    _require_access()

    if not cita_name:
        frappe.throw(_("Falta cita_name"), frappe.ValidationError)
    if not frappe.db.exists("Cita Examen Medico", cita_name):
        frappe.throw(_("Cita no encontrada: {0}").format(cita_name), frappe.ValidationError)

    if estado not in _ESTADOS_VALIDOS:
        frappe.throw(
            _("Estado inválido: {0}. Valores permitidos: {1}.").format(
                estado, ", ".join(sorted(_ESTADOS_VALIDOS))
            ),
            frappe.ValidationError,
        )

    if estado == "Realizada":
        if not concepto or concepto not in _CONCEPTOS_VALIDOS:
            frappe.throw(
                _("Para estado Realizada se requiere concepto: {0}.").format(
                    ", ".join(sorted(_CONCEPTOS_VALIDOS))
                ),
                frappe.ValidationError,
            )

    if estado == "Aplazada":
        motivo_str = (motivo or "").strip()
        if len(motivo_str) < 5:
            frappe.throw(
                _("Para estado Aplazada se requiere un motivo de al menos 5 caracteres."),
                frappe.ValidationError,
            )

    if estado == "Cancelada":
        if not (motivo or "").strip():
            frappe.throw(
                _("Para cancelar la cita se requiere un motivo."),
                frappe.ValidationError,
            )

    from hubgh.hubgh.examen_medico.cita_service import set_exam_outcome

    set_exam_outcome(
        cita_name=cita_name,
        estado=estado,
        concepto=concepto,
        motivo=motivo,
        instrucciones=instrucciones,
    )

    return {"ok": True, "cita": cita_name, "estado": estado}


@frappe.whitelist()
def export_seguimiento_examenes_xlsx(filters=None):
    """Exporta las citas activas de seguimiento a un archivo Excel (xlsx).

    Misma query y filtros que list_seguimiento_examenes, sin paginación.
    Columnas: Nombre completo, Cédula, Celular, Email, Ciudad, Cargo,
              Tipo cargo, IPS, Sede, Fecha cita, Hora cita, Estado, Modo,
              Observaciones SST (completa, sin truncar).

    Returns:
        {"filename": "...", "content_b64": "<base64>", "count": N}
    """
    _require_access()

    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    filters = frappe.parse_json(filters) if filters else {}
    rows, _ = _query_seguimiento_examenes(filters, limit=None, offset=0)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Seguimiento exámenes"

    headers = [
        "Nombre completo",
        "Cédula",
        "Celular",
        "Email",
        "Ciudad",
        "Cargo",
        "Tipo cargo",
        "IPS",
        "Sede",
        "Fecha cita",
        "Hora cita",
        "Estado",
        "Modo",
        "Observaciones SST",
    ]
    ws.append(headers)

    # Estilo del encabezado (mismo patrón que export_proximos_examenes_xlsx)
    header_fill = PatternFill(start_color="1D4ED8", end_color="1D4ED8", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for r in rows:
        ws.append(
            [
                r.get("nombre_completo") or "",
                r.get("numero_documento") or r.get("candidato") or "",
                r.get("celular") or "",
                r.get("email") or "",
                r.get("ciudad") or "",
                r.get("cargo_nombre") or r.get("cargo_al_enviar") or "",
                r.get("tipo_cargo") or "",
                r.get("ips") or "",
                r.get("sede_seleccionada") or "",
                str(r.get("fecha_cita")) if r.get("fecha_cita") else "",
                str(r.get("hora_cita")) if r.get("hora_cita") else "",
                r.get("estado") or "",
                r.get("modo") or "",
                r.get("observaciones_sst") or "",  # sin truncar
            ]
        )

    col_widths = [38, 16, 16, 30, 14, 36, 14, 20, 28, 12, 10, 20, 16, 50]
    for idx, w in enumerate(col_widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = w

    output = BytesIO()
    wb.save(output)
    today = nowdate()

    return {
        "filename": f"seguimiento_examenes_{today}.xlsx",
        "content_b64": base64.b64encode(output.getvalue()).decode("ascii"),
        "count": len(rows),
    }
