"""
Payroll Import Engine - exogenous-first parser registry and run orchestrator.
"""

from __future__ import annotations

import hashlib
import json
import os
from datetime import datetime
from typing import Protocol

import frappe
from frappe.utils import now_datetime

from hubgh.hubgh.payroll_novedad_service import PayrollNovedadService
from hubgh.hubgh.payroll_permissions import enforce_payroll_access


CLONK_RESUMEN_COLUMNS = {
	"documento": ["cedula", "documento", "doc", "id"],
	"nombre": ["nombre", "empleado", "name"],
	"hd": ["hd", "hora diurna", "horas diurnas"],
	"hn": ["hn", "hora nocturna", "horas nocturnas"],
	"hfd": ["hfd", "hora festivo diurna"],
	"hfn": ["hfn", "hora festivo nocturna"],
	"hed": ["hed", "hora extra diurna", "horas extras diurnas"],
	"hen": ["hen", "hora extra nocturna", "horas extras nocturnas"],
	"hefd": ["hefd", "hora extra festivo diurna"],
	"hefn": ["hefn", "hora extra festivo nocturna"],
	"nr": ["nr", "novedad registrada"],
	"nnr": ["nnr", "novedad no registrada"],
	"dnr": ["dnr", "dia no registrado", "días no registrados"],
}

CLONK_AUSENTISMO_COLUMNS = {
	"documento": ["cedula", "documento", "doc", "id"],
	"nombre": ["nombre", "empleado"],
	"tipo_ausencia": ["tipo", "ausencia", "tipo ausencia", "motivo"],
	"fecha_inicio": ["fecha inicio", "inicio", "desde"],
	"fecha_fin": ["fecha fin", "fin", "hasta"],
	"dias": ["dias", "días", "cantidad"],
}

ABSENCE_TYPE_MAP = {
	"descanso": "DESCANSO",
	"vacaciones": "VACACIONES",
	"incapacidad eg": "INC-EG",
	"incapacidad enfermedad general": "INC-EG",
	"incapacidad at": "INC-AT",
	"incapacidad accidente": "INC-AT",
	"enfermedad general": "ENF-GENERAL",
	"ausentismo": "AUSENTISMO",
	"licencia remunerada": "LIC-REM",
	"l. remunerada": "LIC-REM",
	"licencia no remunerada": "LIC-NO-REM",
	"l. no remunerada": "LIC-NO-REM",
	"calamidad": "CALAMIDAD",
	"calamidad doméstica": "CALAMIDAD",
	"maternidad": "MATERNIDAD",
	"licencia maternidad": "MATERNIDAD",
	"día familia": "DIA-FAMILIA",
	"dia familia": "DIA-FAMILIA",
	"cumpleaños": "CUMPLEANOS",
	"cumpleanos": "CUMPLEANOS",
	"inducción": "INDUCCION",
	"induccion": "INDUCCION",
	"luto": "LUTO",
	"licencia luto": "LUTO",
}


class PayrollSourceAdapter(Protocol):
	code: str
	parser_version: str

	def parse(self, file_path: str, batch_doc) -> tuple[list[dict], list[dict]]:
		...


SOURCE_ADAPTERS: dict[str, PayrollSourceAdapter] = {}


def normalize_column_name(col):
	if col is None:
		return ""
	return str(col).strip().lower().replace("_", " ")


def find_column_index(headers, possible_names):
	for idx, header in enumerate(headers):
		normalized = normalize_column_name(header)
		for name in possible_names:
			if name in normalized or normalized in name:
				return idx
	return None


def _safe_json(value):
	return json.dumps(value or {}, ensure_ascii=False, default=str)


def _build_run_label(batch_doc):
	period_label = batch_doc.nomina_period or batch_doc.period or "Sin periodo"
	return f"{period_label} · {batch_doc.run_id}"


def _build_canonical_record(batch_doc, *, row_number, employee_id, employee_name, novelty_code, quantity=None, amount=None, novelty_date=None, source_sheet=None, source_concept=None, raw_payload=None):
	return {
		"batch": batch_doc.name,
		"run_id": batch_doc.run_id,
		"row_number": row_number,
		"employee_id": employee_id,
		"employee_name": employee_name,
		"novedad_type": novelty_code,
		"novedad_date": novelty_date,
		"quantity": quantity,
		"amount": amount,
		"status": "Pendiente",
		"source_sheet": source_sheet,
		"source_file": batch_doc.source_file,
		"source_type_code": batch_doc.source_type,
		"source_row_number": row_number,
		"source_concept_code": source_concept or novelty_code,
		"parser_version": "clonk.v2",
		"source_row_data": _safe_json(raw_payload),
		"raw_payload_json": _safe_json(
			{
				"employee_ref": employee_id,
				"employee_name": employee_name,
				"novelty_code": novelty_code,
				"quantity": quantity,
				"amount": amount,
				"effective_date": novelty_date,
				"source_concept": source_concept or novelty_code,
				"source_sheet": source_sheet,
				"source_row": row_number,
				"raw_payload": raw_payload or {},
				"provenance": {
					"run_id": batch_doc.run_id,
					"batch": batch_doc.name,
					"file_url": batch_doc.source_file,
					"source_type": batch_doc.source_type,
					"parser_version": "clonk.v2",
					"imported_at": now_datetime().isoformat(),
				},
			}
		),
	}


class ClonkSourceAdapter:
	code = "clonk"
	parser_version = "clonk.v2"

	def parse(self, file_path: str, batch_doc) -> tuple[list[dict], list[dict]]:
		try:
			import openpyxl
		except ImportError:
			frappe.throw("openpyxl no está instalado. Ejecute: pip install openpyxl")

		lines = []
		errors = []
		try:
			workbook = openpyxl.load_workbook(file_path, data_only=True)
		except Exception as exc:
			frappe.throw(f"Error al abrir archivo CLONK: {exc}")

		for sheet_name in workbook.sheetnames:
			sheet = workbook[sheet_name]
			sheet_lower = sheet_name.lower()
			if "resumen" in sheet_lower or "horas" in sheet_lower:
				sheet_lines, sheet_errors = self._parse_resumen_sheet(sheet, sheet_name, batch_doc)
			elif "ausent" in sheet_lower or "novedad" in sheet_lower:
				sheet_lines, sheet_errors = self._parse_ausentismo_sheet(sheet, sheet_name, batch_doc)
			else:
				continue
			lines.extend(sheet_lines)
			errors.extend(sheet_errors)

		workbook.close()
		return lines, errors

	def _parse_resumen_sheet(self, sheet, sheet_name, batch_doc):
		lines = []
		errors = []
		rows = list(sheet.iter_rows(values_only=True))
		if len(rows) < 2:
			return lines, errors

		headers = rows[0]
		col_map = {}
		for field, possible_names in CLONK_RESUMEN_COLUMNS.items():
			idx = find_column_index(headers, possible_names)
			if idx is not None:
				col_map[field] = idx

		if "documento" not in col_map:
			errors.append({"sheet": sheet_name, "error": "No se encontró columna de documento"})
			return lines, errors

		for row_num, row in enumerate(rows[1:], start=2):
			doc_value = row[col_map["documento"]] if col_map.get("documento") is not None else None
			if not doc_value:
				continue

			employee_id = str(doc_value).strip()
			employee_name = str(row[col_map.get("nombre", 0)] or "").strip() if col_map.get("nombre") is not None else ""
			for hour_type in ["hd", "hn", "hfd", "hfn", "hed", "hen", "hefd", "hefn", "nr", "nnr", "dnr"]:
				if hour_type not in col_map:
					continue
				value = row[col_map[hour_type]]
				if not value or value == 0:
					continue
				try:
					quantity = float(value)
				except (TypeError, ValueError):
					continue
				if quantity <= 0:
					continue

				raw_payload = {
					"row": row_num,
					"documento": employee_id,
					"nombre": employee_name,
					"tipo": hour_type,
					"valor": quantity,
				}
				lines.append(
					_build_canonical_record(
						batch_doc,
						row_number=row_num,
						employee_id=employee_id,
						employee_name=employee_name,
						novelty_code=hour_type.upper(),
						quantity=quantity,
						source_sheet=sheet_name,
						source_concept=hour_type.upper(),
						raw_payload=raw_payload,
					)
				)

		return lines, errors

	def _parse_ausentismo_sheet(self, sheet, sheet_name, batch_doc):
		lines = []
		errors = []
		rows = list(sheet.iter_rows(values_only=True))
		if len(rows) < 2:
			return lines, errors

		headers = rows[0]
		col_map = {}
		for field, possible_names in CLONK_AUSENTISMO_COLUMNS.items():
			idx = find_column_index(headers, possible_names)
			if idx is not None:
				col_map[field] = idx

		if "documento" not in col_map:
			errors.append({"sheet": sheet_name, "error": "No se encontró columna de documento"})
			return lines, errors

		for row_num, row in enumerate(rows[1:], start=2):
			doc_value = row[col_map["documento"]] if col_map.get("documento") is not None else None
			if not doc_value:
				continue

			employee_id = str(doc_value).strip()
			employee_name = str(row[col_map.get("nombre", 0)] or "").strip() if col_map.get("nombre") is not None else ""
			tipo_ausencia_raw = str(row[col_map.get("tipo_ausencia", 0)] or "").strip().lower() if col_map.get("tipo_ausencia") is not None else ""
			novedad_code = ABSENCE_TYPE_MAP.get(tipo_ausencia_raw, "AUSENTISMO")
			fecha_inicio = row[col_map.get("fecha_inicio")] if col_map.get("fecha_inicio") is not None else None
			dias = row[col_map.get("dias")] if col_map.get("dias") is not None else None
			novedad_date = None
			if fecha_inicio:
				if isinstance(fecha_inicio, datetime):
					novedad_date = fecha_inicio.strftime("%Y-%m-%d")
				else:
					novedad_date = str(fecha_inicio)
			quantity = None
			if dias not in (None, ""):
				try:
					quantity = float(dias)
				except (TypeError, ValueError):
					quantity = 1

			raw_payload = {
				"row": row_num,
				"documento": employee_id,
				"nombre": employee_name,
				"tipo_ausencia": tipo_ausencia_raw,
				"fecha_inicio": str(fecha_inicio) if fecha_inicio else None,
				"dias": dias,
			}
			lines.append(
				_build_canonical_record(
					batch_doc,
					row_number=row_num,
					employee_id=employee_id,
					employee_name=employee_name,
					novelty_code=novedad_code,
					quantity=quantity,
					novelty_date=novedad_date,
					source_sheet=sheet_name,
					source_concept=tipo_ausencia_raw or novedad_code,
					raw_payload=raw_payload,
				)
			)

		return lines, errors


def register_source_adapter(adapter: PayrollSourceAdapter, *aliases: str):
	keys = {adapter.code, *(alias for alias in aliases if alias)}
	for key in keys:
		SOURCE_ADAPTERS[str(key).strip().lower()] = adapter


def get_source_adapter(source_type: str | None = None, tipo_fuente: str | None = None):
	for candidate in [tipo_fuente, source_type]:
		if not candidate:
			continue
		adapter = SOURCE_ADAPTERS.get(str(candidate).strip().lower())
		if adapter:
			return adapter
	return None


def ensure_default_adapters():
	if SOURCE_ADAPTERS:
		return
	register_source_adapter(ClonkSourceAdapter(), "CLONK")


def parse_clonk_file(file_path, batch_name):
	batch_doc = type("BatchDoc", (), {
		"name": batch_name,
		"run_id": batch_name,
		"source_file": file_path,
		"source_type": "CLONK",
	})()
	adapter = ClonkSourceAdapter()
	return adapter.parse(file_path, batch_doc)


def generate_dedup_hash(periodo, employee_id, novelty_type, novelty_date=None):
	key_parts = [str(periodo or ""), str(employee_id or ""), str(novelty_type or ""), str(novelty_date or "")]
	return hashlib.md5("|".join(key_parts).encode()).hexdigest()


def check_duplicate_line(periodo, employee_id, novelty_type, novelty_date=None):
	dedup_hash = generate_dedup_hash(periodo, employee_id, novelty_type, novelty_date)
	existing = frappe.db.sql(
		"""
			SELECT name
			FROM `tabPayroll Import Line`
			WHERE dedup_hash = %s
			LIMIT 1
		""",
		[dedup_hash],
	)
	return existing[0][0] if existing else None


def detect_source_type(file_path):
	filename = os.path.basename(file_path).lower()
	if "clonk" in filename or "toda la empresa" in filename:
		return "CLONK"
	if "payflow" in filename:
		return "Payflow Resumen"
	if "fincomercio" in filename:
		return "Fincomercio"
	if "fondo" in filename or "fongiga" in filename or "m home" in filename:
		return "Fondo FONGIGA"
	if "libranza" in filename:
		return "Libranzas Bancolombia"
	return None


def _get_batch_file_path(batch_doc):
	if not batch_doc.source_file:
		raise ValueError("No se encontró archivo fuente.")
	file_doc = frappe.get_doc("File", {"file_url": batch_doc.source_file})
	return file_doc.get_full_path()


def _mark_unsupported_batch(batch_doc, source_doc=None):
	tipo_fuente = getattr(source_doc, "tipo_fuente", None) or batch_doc.source_type
	batch_doc.reload()
	batch_doc.status = "Fuente no soportada"
	batch_doc.processed_on = now_datetime()
	batch_doc.processing_log = _safe_json(
		{
			"unsupported_source": True,
			"source_type": batch_doc.source_type,
			"tipo_fuente": tipo_fuente,
			"source_file": batch_doc.source_file,
		}
	)
	batch_doc.save(ignore_permissions=True)
	frappe.db.commit()
	return {
		"status": batch_doc.status,
		"total_rows": 0,
		"valid_rows": 0,
		"error_rows": 0,
		"duplicate_rows": 0,
		"errors": [{"error": f"Fuente no soportada: {tipo_fuente}"}],
	}


def _persist_import_lines(batch_doc, lines, errors):
	service = PayrollNovedadService()
	prepared_lines = service.prepare_import_lines(lines, batch_doc)
	valid_count = 0
	error_count = 0
	duplicate_count = 0
	for line_data in prepared_lines:
		try:
			period = batch_doc.nomina_period or batch_doc.period
			duplicate_line = check_duplicate_line(
				period,
				line_data.get("employee_id"),
				line_data.get("novedad_type"),
				line_data.get("novedad_date"),
			)
			if duplicate_line:
				line_data["status"] = "Duplicado"
				line_data["validation_errors"] = f"Duplicado de línea existente: {duplicate_line}"
				duplicate_count += 1

			line_data["dedup_hash"] = generate_dedup_hash(
				period,
				line_data.get("employee_id"),
				line_data.get("novedad_type"),
				line_data.get("novedad_date"),
			)
			frappe.get_doc({"doctype": "Payroll Import Line", **line_data}).insert(ignore_permissions=True)
			valid_count += 1
		except Exception as exc:
			error_count += 1
			errors.append({"row": line_data.get("row_number"), "error": str(exc)})

	batch_doc.reload()
	batch_doc.total_rows = len(prepared_lines)
	batch_doc.valid_rows = max(valid_count - duplicate_count, 0)
	batch_doc.error_rows = error_count
	batch_doc.processed_on = now_datetime()
	batch_doc.processing_log = _safe_json({"errors": errors, "duplicates_found": duplicate_count, "total_processed": len(prepared_lines)}) if (errors or duplicate_count) else None
	if error_count == 0:
		batch_doc.status = "Completado con duplicados" if duplicate_count else "Completado"
	elif valid_count > 0:
		batch_doc.status = "Completado con errores"
	else:
		batch_doc.status = "Fallido"
	batch_doc.save(ignore_permissions=True)
	frappe.db.commit()
	return {
		"status": batch_doc.status,
		"total_rows": batch_doc.total_rows,
		"valid_rows": batch_doc.valid_rows,
		"error_rows": batch_doc.error_rows,
		"duplicate_rows": duplicate_count,
		"errors": errors,
	}


def process_import_batch(batch_name):
	ensure_default_adapters()
	batch_doc = frappe.get_doc("Payroll Import Batch", batch_name)
	if batch_doc.status not in ("Pendiente", "Fallido", "Fuente no soportada"):
		frappe.throw(f"El lote {batch_name} ya fue procesado.")

	batch_doc.status = "Procesando"
	batch_doc.run_label = batch_doc.run_label or _build_run_label(batch_doc)
	batch_doc.save(ignore_permissions=True)
	frappe.db.commit()

	try:
		file_path = _get_batch_file_path(batch_doc)
		source_doc = frappe.get_doc("Payroll Source Catalog", batch_doc.source_type)
		adapter = get_source_adapter(batch_doc.source_type, getattr(source_doc, "tipo_fuente", None))
		if not adapter:
			return _mark_unsupported_batch(batch_doc, source_doc)

		lines, errors = adapter.parse(file_path, batch_doc)
		return _persist_import_lines(batch_doc, lines, errors)
	except Exception as exc:
		batch_doc.reload()
		batch_doc.status = "Fallido"
		batch_doc.processed_on = now_datetime()
		batch_doc.processing_log = _safe_json({"error": str(exc)})
		batch_doc.save(ignore_permissions=True)
		frappe.db.commit()
		raise


def process_import_run(run_id):
	batches = frappe.get_all(
		"Payroll Import Batch",
		filters={"run_id": run_id},
		fields=["name", "status", "nomina_period", "run_label", "run_source_count"],
		order_by="creation asc",
	)
	if not batches:
		frappe.throw("No se encontraron lotes para el run indicado.")

	results = []
	totals = {"total_rows": 0, "valid_rows": 0, "error_rows": 0, "duplicate_rows": 0}
	for batch in batches:
		if batch.get("status") in ("Confirmado", "Aprobado TC", "Rechazado TC"):
			result = {
				"status": batch.get("status"),
				"total_rows": frappe.db.get_value("Payroll Import Batch", batch["name"], "total_rows") or 0,
				"valid_rows": frappe.db.get_value("Payroll Import Batch", batch["name"], "valid_rows") or 0,
				"error_rows": frappe.db.get_value("Payroll Import Batch", batch["name"], "error_rows") or 0,
				"duplicate_rows": 0,
				"errors": [],
			}
		else:
			result = process_import_batch(batch["name"])
		results.append({"batch": batch["name"], **result})
		for key in totals:
			totals[key] += int(result.get(key) or 0)

	statuses = [row["status"] for row in results]
	if any(status == "Fallido" for status in statuses):
		status = "Fallido"
	elif any(status == "Fuente no soportada" for status in statuses):
		status = "Fuente no soportada"
	elif any(status == "Completado con errores" for status in statuses):
		status = "Completado con errores"
	elif any(status == "Completado con duplicados" for status in statuses):
		status = "Completado con duplicados"
	else:
		status = "Completado"

	first_batch = batches[0]
	return {
		"run_id": run_id,
		"run_label": first_batch.get("run_label") or run_id,
		"source_count": first_batch.get("run_source_count") or len(batches),
		"status": status,
		"batch_results": results,
		**totals,
	}


@frappe.whitelist()
def process_batch(batch_name):
	enforce_payroll_access("import_batches")
	return process_import_batch(batch_name)


@frappe.whitelist()
def process_run(run_id):
	enforce_payroll_access("import_batches")
	return process_import_run(run_id)
