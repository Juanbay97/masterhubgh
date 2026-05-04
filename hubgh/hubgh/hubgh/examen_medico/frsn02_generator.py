# Copyright (c) 2026, Antigravity and contributors
# For license information, please see license.txt

"""
frsn02_generator — Generador de formulario FRSN-02 (Orden de Servicio Médico).

Abre el template xlsx configurado en la IPS (openpyxl), lo clona en memoria,
llena las celdas con los datos del candidato y retorna los bytes del archivo.
No modifica el template original.

Mapa de celdas canónico (FRSN-02 v01):
  B6  = fecha_solicitud (hoy)
  B8  = candidato.nombre
  B9  = candidato.cedula
  B10 = candidato.cargo_al_enviar
  B11 = candidato.ciudad
  Celdas de exámenes = tomadas de IPS Examen Estandar Por Cargo.celda_excel
"""

from __future__ import annotations

# Default cell map for FRSN-02 template (Zonamedica FRSN-02 v01).
# Coordinates verified against the real template at
# `Correos SST/FRSN-02-ORDEN DE SERVICIO DE EXAMENES MEDICOS OCUPACIONALES - laura 1 (27).xlsx`.
# Admin can override via IPS fields in future versions.
DEFAULT_CELL_MAP = {
	"fecha_solicitud": "D10",
	"nombre": "E13",
	"cedula": "E14",
	"cargo": "N13",
	"ciudad": "M14",
}

# Default cell for the "Ingreso" tipo de examen mark (column H of row 16,
# next to the "Ingreso" label at G16).
DEFAULT_TIPO_EXAMEN_INGRESO_CELL = "H16"


def generate_frsn02(
	ips: dict,
	candidato: dict,
	fecha_examen: str | None = None,
) -> bytes:
	"""
	Abre el template xlsx de la IPS, llena celdas del candidato y retorna bytes.

	Args:
		ips: Documento IPS como dict con template_orden_servicio y examenes_estandar.
		candidato: Dict con {nombre, cedula, cargo, ciudad}.
		fecha_examen: (Opcional, no usado actualmente) fecha del examen "YYYY-MM-DD".
		              Se mantiene por compatibilidad con el contrato original.

	Returns:
		Bytes del archivo xlsx generado (para adjuntar en email).

	Note:
		Celdas de exámenes estándar se toman de IPS Examen Estandar Por Cargo.celda_excel.
		La celda de tipo_examen_ingreso es configurable en IPS o default "F14".

	Raises:
		frappe.ValidationError: Si template_orden_servicio no está configurado en la IPS.
	"""
	import frappe
	from io import BytesIO
	import openpyxl
	from datetime import date

	template_url = ips.get("template_orden_servicio")
	if not template_url:
		frappe.throw(
			"La IPS no tiene template de orden de servicio configurado.",
			frappe.ValidationError,
		)

	# Fetch the file content from Frappe's file system
	file_doc = frappe.get_doc("File", {"file_url": template_url})
	file_path = file_doc.get_full_path()

	with open(file_path, "rb") as f:
		template_bytes = f.read()

	# Load workbook in memory (do NOT write to the original)
	wb = openpyxl.load_workbook(BytesIO(template_bytes))
	ws = wb.active

	# Fill candidate data using default cell map
	cell_map = DEFAULT_CELL_MAP
	today_str = date.today().isoformat()

	ws[cell_map["fecha_solicitud"]] = today_str
	ws[cell_map["nombre"]] = candidato.get("nombre", "")
	ws[cell_map["cedula"]] = candidato.get("cedula", "")
	ws[cell_map["cargo"]] = candidato.get("cargo", "")
	ws[cell_map["ciudad"]] = candidato.get("ciudad", "")

	# Mark tipo_examen_ingreso cell
	tipo_cell = ips.get("celda_tipo_examen_ingreso") or DEFAULT_TIPO_EXAMEN_INGRESO_CELL
	ws[tipo_cell] = "X"

	# Mark per-exam cells for this candidate's cargo
	candidato_cargo = candidato.get("cargo", "")
	for row in ips.get("examenes_estandar") or []:
		if isinstance(row, dict):
			row_cargo = row.get("cargo", "")
			celda = row.get("celda_excel", "")
		else:
			row_cargo = getattr(row, "cargo", "")
			celda = getattr(row, "celda_excel", "")

		if row_cargo == candidato_cargo and celda:
			ws[celda] = "X"

	# Save to BytesIO and return bytes
	output = BytesIO()
	wb.save(output)
	return output.getvalue()
