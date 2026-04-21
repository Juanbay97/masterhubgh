# Copyright (c) 2026, Antigravity and contributors
# For license information, please see license.txt

"""
Tests for frsn02_generator.generate_frsn02.

Strategy (Batch 3 — GREEN):
- generate_frsn02 is fully implemented.
- Tests verify actual xlsx output by loading the returned bytes with openpyxl.
- Uses FrappeTestCase for bench harness.
- frappe.get_doc("File", ...) is patched to avoid real file system access.
- openpyxl creates a minimal in-memory template for testing.

REQ refs: REQ-23 (FRSN-02 xlsx, celdas candidato, tipo_examen, examenes_estandar).
"""

from io import BytesIO
from unittest.mock import patch, MagicMock
from frappe.tests.utils import FrappeTestCase
import frappe

from hubgh.hubgh.examen_medico import frsn02_generator


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _make_minimal_xlsx() -> bytes:
	"""Create a minimal xlsx workbook in memory for testing."""
	import openpyxl
	wb = openpyxl.Workbook()
	ws = wb.active
	output = BytesIO()
	wb.save(output)
	return output.getvalue()


def _make_ips(cargo="Auxiliar", examenes=None):
	return {
		"name": "IPS-TEST",
		"template_orden_servicio": "/private/files/frsn02.xlsx",
		"celda_tipo_examen_ingreso": "F14",
		"examenes_estandar": examenes or [
			{"cargo": cargo, "nombre_examen": "Hemograma", "celda_excel": "E18"},
			{"cargo": cargo, "nombre_examen": "Glicemia", "celda_excel": "E19"},
		],
	}


def _make_candidato(cargo="Auxiliar"):
	return {
		"nombre": "Ana Gómez",
		"cedula": "12345",
		"cargo": cargo,
		"ciudad": "Cartagena",
	}


def _patch_frappe_file(xlsx_bytes: bytes):
	"""Return a context manager that patches frappe.get_doc('File', ...) to return a fake file."""
	import tempfile, os

	# Write bytes to a temp file and return its path via get_full_path()
	tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
	tmp.write(xlsx_bytes)
	tmp.flush()
	tmp.close()

	fake_file_doc = MagicMock()
	fake_file_doc.get_full_path.return_value = tmp.name

	return patch.object(frappe, "get_doc", return_value=fake_file_doc), tmp.name


# ---------------------------------------------------------------------------
# Tests — all GREEN after Batch 3 implementation
# ---------------------------------------------------------------------------

class TestFrsn02Generator(FrappeTestCase):

	def _generate(self, ips=None, candidato=None, fecha="2026-08-10"):
		"""Helper: patches file access, calls generate_frsn02, returns (wb, tmp_path)."""
		import openpyxl, tempfile, os

		if ips is None:
			ips = _make_ips()
		if candidato is None:
			candidato = _make_candidato()

		xlsx_bytes = _make_minimal_xlsx()
		tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
		tmp.write(xlsx_bytes)
		tmp.flush()
		tmp.close()

		fake_file_doc = MagicMock()
		fake_file_doc.get_full_path.return_value = tmp.name

		with patch.object(frappe, "get_doc", return_value=fake_file_doc):
			result = frsn02_generator.generate_frsn02(ips, candidato, fecha)

		os.unlink(tmp.name)

		wb = openpyxl.load_workbook(BytesIO(result))
		return wb

	def test_fills_candidato_cells_nombre_cedula_cargo_ciudad(self):
		"""REQ-23: Celdas D13/D14/P13/P14 tienen nombre, cédula, cargo, ciudad del candidato."""
		wb = self._generate()
		ws = wb.active

		self.assertEqual(ws["D13"].value, "Ana Gómez", "D13 debe tener el nombre")
		self.assertEqual(ws["D14"].value, "12345", "D14 debe tener la cédula")
		self.assertEqual(ws["P13"].value, "Auxiliar", "P13 debe tener el cargo")
		self.assertEqual(ws["P14"].value, "Cartagena", "P14 debe tener la ciudad")

	def test_marks_x_in_tipo_examen_ingreso_cell(self):
		"""REQ-23: Celda de tipo_examen_ingreso (F14 por defecto) marcada con 'X'."""
		wb = self._generate()
		ws = wb.active
		self.assertEqual(ws["F14"].value, "X", "F14 debe tener 'X' para tipo_examen_ingreso")

	def test_marks_x_per_examen_estandar_por_cargo_row(self):
		"""REQ-23: Para cada examen del cargo, la celda_excel correspondiente recibe 'X'."""
		wb = self._generate()
		ws = wb.active
		self.assertEqual(ws["E18"].value, "X", "E18 debe tener 'X' para Hemograma")
		self.assertEqual(ws["E19"].value, "X", "E19 debe tener 'X' para Glicemia")

	def test_saves_to_private_file_linked_to_cita(self):
		"""REQ-23: generate_frsn02 retorna bytes (no None, no vacío)."""
		import openpyxl, tempfile, os

		ips = _make_ips()
		candidato = _make_candidato()
		xlsx_bytes = _make_minimal_xlsx()
		tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
		tmp.write(xlsx_bytes)
		tmp.flush()
		tmp.close()

		fake_file_doc = MagicMock()
		fake_file_doc.get_full_path.return_value = tmp.name

		with patch.object(frappe, "get_doc", return_value=fake_file_doc):
			result = frsn02_generator.generate_frsn02(ips, candidato, "2026-08-10")

		os.unlink(tmp.name)

		self.assertIsNotNone(result, "Resultado no debe ser None")
		self.assertIsInstance(result, bytes, "Resultado debe ser bytes")
		self.assertGreater(len(result), 0, "Resultado no debe estar vacío")

	def test_raises_when_template_missing(self):
		"""REQ-23: Sin template_orden_servicio en IPS, debe lanzar excepción."""
		ips_sin_template = {
			"name": "IPS-NO-TEMPLATE",
			"template_orden_servicio": None,
			"examenes_estandar": [],
		}
		with self.assertRaises(Exception):
			frsn02_generator.generate_frsn02(ips_sin_template, _make_candidato(), "2026-08-10")
